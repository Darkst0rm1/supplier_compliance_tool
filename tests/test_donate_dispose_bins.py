"""Tests for the Donate / Dispose list's optional EWM storage-bin lookup.

Self-contained: every fixture workbook is built in memory, so these tests do not
depend on the supplied sample files.

The lookup reproduces the manual Excel workflow it replaces
(``VLOOKUP(Material&Batch, EWM!I:J, 2, 0)``), so the behaviours pinned here are
the ones that workflow guaranteed: first match wins, the key ignores padding,
and a missing bin never removes a row.
"""
from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
import pandas as pd
import pytest

from src.donate_dispose_engine import (
    OUTPUT_COLUMNS,
    EWM_BATCH,
    EWM_BIN,
    EWM_OWNER,
    EWM_PRODUCT,
    MAT_BATCH,
    MAT_SLED,
    NO_LOOKUP_MARKER,
    OUT_BIN,
    OUT_MATERIALBATCH,
    DonateDisposeError,
    build_donate_dispose,
    generate_excel,
    load_ewm_bins,
)

SLED_CUTOFF = date(2026, 8, 5)


def _ewm_book(rows, owner="BP2910"):
    """An EWM dispose export as bytes. ``rows`` are (product, batch, bin)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = [EWM_PRODUCT, EWM_BATCH, EWM_BIN, EWM_OWNER]
    ws.append(headers)
    for product, batch, bin_ in rows:
        ws.append([product, batch, bin_, owner])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _materials(rows):
    """A loaded-Materials frame. ``rows`` are (material, plant, batch)."""
    return pd.DataFrame([
        {
            "Material": material,
            "Material Description": "TEST ITEM",
            "Plant": plant,
            "Plant Name": "Test Whse",
            "Storage Location": "1000",
            "Description of Storage Location": "Main Whse",
            "Batch": batch,
            "Shelf Life Expiration Date": pd.Timestamp(2026, 8, 1),
            "Special Stock Type Description": None,
            "Unrestricted Stock": 10,
            "Stock in Quality Inspection": 0,
            "Blocked Stock": 0,
        }
        for material, plant, batch in rows
    ])


def _master(materials, last_sell_day=30):
    return pd.DataFrame([
        {
            "Product Number": m,
            "Brand Manager Name": "Test BDM",
            "Last Sell Day": last_sell_day,
        }
        for m in materials
    ])


def _build(mat_rows, ewm_bins=None):
    materials = _materials(mat_rows)
    master = _master({m for m, _, _ in mat_rows})
    return build_donate_dispose(
        materials, master, sled_cutoff=SLED_CUTOFF, ewm_bins=ewm_bins,
    )


# ---------------------------------------------------------------------------
# The lookup table
# ---------------------------------------------------------------------------
def test_key_is_material_and_batch_concatenated():
    lut = load_ewm_bins(_ewm_book([("10026065", "2609191170", "FDA62F")]))
    assert lut["100260652609191170"] == "FDA62F"


def test_first_match_wins_for_a_repeated_product_batch():
    """EWM lists a batch once per bin it occupies. The VLOOKUP this replaces
    returned the first row in file order, so file order is load-bearing."""
    lut = load_ewm_bins(_ewm_book([
        ("10026065", "2609191170", "FDA62F"),
        ("10026065", "2609191170", "FDB08F"),
        ("10026065", "2609191170", "FDA35F"),
    ]))
    assert lut["100260652609191170"] == "FDA62F"
    assert len(lut) == 1


@pytest.mark.parametrize("pad", [" ", " "], ids=["space", "nbsp"])
def test_padding_does_not_break_the_key(pad):
    r"""Batches arrive padded differently in the two exports, and the padding is
    often a **non-breaking** space — ``load_materials`` preserves those on
    purpose. A ``\s+`` regex would not strip U+00A0 here (pandas runs
    ``str.replace(regex=True)`` on RE2, whose ``\s`` is ASCII-only), so this
    guards the key against silently missing every padded batch."""
    lut = load_ewm_bins(_ewm_book([
        (f"10026065{pad}", f"{pad}2609191170{pad}", "FDA62F"),
    ]))
    assert lut["100260652609191170"] == "FDA62F"


def test_rows_without_a_bin_are_dropped_from_the_lookup():
    lut = load_ewm_bins(_ewm_book([
        ("10026065", "2609191170", None),
        ("10026067", "2608301150", "FDA61F"),
    ]))
    assert "100260652609191170" not in lut
    assert lut["100260672608301150"] == "FDA61F"


def test_file_uploaded_for_the_wrong_plant_is_rejected():
    with pytest.raises(DonateDisposeError, match="uploaded in the 2910 box"):
        load_ewm_bins(_ewm_book([("1", "A", "X")], owner="BP2920"),
                      expected_plant="2910")


def test_matching_plant_is_accepted():
    lut = load_ewm_bins(_ewm_book([("1", "A", "X")], owner="BP2910"),
                        expected_plant="2910")
    assert lut["1A"] == "X"


def test_a_non_ewm_file_is_rejected_by_name():
    wb = openpyxl.Workbook()
    wb.active.append(["Something", "Else"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(DonateDisposeError, match="EWM dispose"):
        load_ewm_bins(buf, expected_plant="2910")


# ---------------------------------------------------------------------------
# Applying bins to the report
# ---------------------------------------------------------------------------
def test_bin_is_filled_from_the_matching_plant_only():
    """2920 and 2910 share a sheet-free namespace: a key present in one plant's
    export must not leak into another plant's rows."""
    sheets = _build(
        [("10026065", "2910", "B1"), ("10026065", "2920", "B1")],
        ewm_bins={"2910": load_ewm_bins(_ewm_book([("10026065", "B1", "WDA58D")]))},
    )
    assert sheets["Mississauga"][OUT_BIN].tolist() == ["WDA58D"]
    assert sheets["Calgary"][OUT_BIN].tolist() == [NO_LOOKUP_MARKER]


def test_plants_with_no_ewm_export_keep_the_na_marker():
    """2925 and 2935 have no EWM extract at all, so nothing was searched for
    them. That must stay #N/A — blank would claim we looked and found no bin."""
    sheets = _build(
        [("10026065", "2920", "B1"), ("10026065", "2925", "B1")],
        ewm_bins={"2920": load_ewm_bins(_ewm_book([("10026065", "B1", "GA60E")]))},
    )
    calgary = sheets["Calgary"]
    by_plant = dict(zip(calgary["Plant"], calgary[OUT_BIN]))
    assert by_plant["2920"] == "GA60E"
    assert by_plant["2925"] == NO_LOOKUP_MARKER


def test_a_skipped_upload_leaves_that_plant_as_na_not_blank():
    """Same rule for 2910/2920/2930 when the user doesn't upload that plant's
    file — nothing was searched, so it is #N/A, not a blank."""
    sheets = _build(
        [("10026065", "2910", "B1"), ("10026065", "2920", "B1")],
        ewm_bins={"2910": load_ewm_bins(_ewm_book([("10026065", "B1", "WDA58D")]))},
    )
    assert sheets["Mississauga"][OUT_BIN].tolist() == ["WDA58D"]
    assert sheets["Calgary"][OUT_BIN].tolist() == [NO_LOOKUP_MARKER]


def test_na_marker_is_written_as_a_real_excel_error():
    """The golden's #N/A came from a VLOOKUP, i.e. a genuine error value.
    openpyxl binds the string to data type 'e', so it renders identically."""
    sheets = _build([("10026065", "2925", "B1")])
    ws = openpyxl.load_workbook(io.BytesIO(generate_excel(sheets)))["Calgary"]
    bin_col = [c.value for c in ws[1]].index(OUT_BIN) + 1
    cell = ws.cell(row=2, column=bin_col)
    assert cell.value == NO_LOOKUP_MARKER
    assert cell.data_type == "e"


def test_an_unmatched_batch_is_blank_not_dropped():
    sheets = _build(
        [("10026065", "2910", "MISSING")],
        ewm_bins={"2910": load_ewm_bins(_ewm_book([("10026065", "OTHER", "WDA58D")]))},
    )
    assert len(sheets["Mississauga"]) == 1
    # Searched and not found -> blank, which is NOT the same as #N/A.
    assert pd.isna(sheets["Mississauga"][OUT_BIN].iloc[0])


def test_bins_never_change_row_selection():
    rows = [("10026065", "2910", "B1"), ("10026067", "2920", "B2"),
            ("10026068", "2930", "B3")]
    bins = {"2910": load_ewm_bins(_ewm_book([("10026065", "B1", "WDA58D")]))}
    without, with_ = _build(rows), _build(rows, ewm_bins=bins)
    for sheet in without:
        pd.testing.assert_frame_equal(
            without[sheet].drop(columns=[OUT_BIN]),
            with_[sheet].drop(columns=[OUT_BIN]),
        )


def test_report_builds_with_no_ewm_files_at_all():
    sheets = _build([("10026065", "2910", "B1")])
    assert len(sheets["Mississauga"]) == 1
    assert (sheets["Mississauga"][OUT_BIN] == NO_LOOKUP_MARKER).all()


# ---------------------------------------------------------------------------
# Layout — the business signs off on these files, so column positions matter
# ---------------------------------------------------------------------------
def test_materialbatch_is_the_key_and_appears_once():
    """One key column, matching the Overstock report's final layout."""
    sheets = _build([("10026065", "2910", "2609191170")])
    cols = list(sheets["Mississauga"].columns)
    assert cols.count(OUT_MATERIALBATCH) == 1
    assert sheets["Mississauga"].iloc[0][OUT_MATERIALBATCH] == "100260652609191170"


def test_a_batch_less_row_still_gets_a_key_and_survives():
    """A blank Batch must degrade to the bare Material, not blank the whole key
    (``astype(str)`` leaves NA as NA on the string dtype)."""
    sheets = _build([("10026065", "2910", None)])
    assert sheets["Mississauga"].iloc[0][OUT_MATERIALBATCH] == "10026065"


def test_new_columns_sit_after_storage_location():
    """Same position as the Overstock report, so the two read alike."""
    assert OUTPUT_COLUMNS[6:8] == [OUT_MATERIALBATCH, OUT_BIN]
    assert OUTPUT_COLUMNS[9:11] == [MAT_BATCH, MAT_SLED]


def test_excel_headers_match_the_layout():
    sheets = _build([("10026065", "2910", "2609191170")])
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(sheets)))
    headers = [c.value for c in wb["Mississauga"][1]]
    assert headers[6] == "Materialbatch"
    assert headers[7] == "Bin"
    assert headers.count("Materialbatch") == 1
    assert headers[9:11] == ["Batch", "Shelf Life Expiration Date"]


def test_long_numeric_keys_are_written_as_text():
    """A 18-digit Material+Batch would lose precision if Excel stored it as a
    number."""
    sheets = _build([("10026065", "2910", "2609191170")])
    ws = openpyxl.load_workbook(io.BytesIO(generate_excel(sheets)))["Mississauga"]
    assert ws.cell(row=2, column=7).value == "100260652609191170"
    assert ws.cell(row=2, column=7).number_format == "@"


def test_autofilter_covers_the_widened_layout():
    sheets = _build([("10026065", "2910", "B1")])
    ws = openpyxl.load_workbook(io.BytesIO(generate_excel(sheets)))["Mississauga"]
    assert ws.auto_filter.ref == "A1:Q2"
