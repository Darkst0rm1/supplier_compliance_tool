"""Tests for the Daily Short Report Excel export."""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import load_workbook

from src.daily_short_engine import build_kpis, generate_excel_report, load_daily_short


def _sample_xlsx() -> io.BytesIO:
    raw = pd.DataFrame({
        "Sales Order": ["3001", "3001", "3002"],
        "Material": ["100", "200", "300"],
        "TOL Material Description": ["Widget A", "Widget B", "Widget C"],
        "Plant": ["2910", "2910", "2920"],
        "Sold To Name": ["ACME", "ACME", "BETA"],
        "Order Quantity": [10, 5, 8],
        "Confirmed Quantity (CS)": [8, 5, 0],
        "Total Delivery (Sales) Quantity (CS)": [6, 5, 0],
        "Invoice Quantity (CS)": [6, 4, 0],
        "Reason for Rejection Desc.": ["", "", "No stock"],
    })
    buf = io.BytesIO()
    raw.to_excel(buf, index=False)
    buf.seek(0)
    return buf


def test_excel_has_raw_data_sheet():
    df = load_daily_short(_sample_xlsx())
    assert not df.empty
    xlsx = generate_excel_report(df, build_kpis(df), top_n=5)

    wb = load_workbook(io.BytesIO(xlsx))
    assert "Raw Data" in wb.sheetnames

    ws = wb["Raw Data"]
    headers = [c.value for c in ws[1]]
    # source columns present; internal helper columns (e.g. _short_*) excluded
    assert "Sales Order" in headers
    assert "Order Quantity" in headers
    assert not any(str(h).startswith("_short_") for h in headers)
    # one header row + one row per order line
    assert ws.max_row == len(df) + 1
