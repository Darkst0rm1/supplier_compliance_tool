"""Tests for the Risky Inventory engine.

Self-contained: every fixture workbook is built in memory, so these tests do not
depend on the supplied sample files.
"""
from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
import pytest

from src.risky_inventory_engine import (
    BUCKET_0_90,
    BUCKET_91_180,
    BUCKET_NONE,
    DETAIL_HEADERS,
    GRAND_TOTAL_LABEL,
    SUMMARY_HEADER,
    RiskyInventoryError,
    assign_buckets,
    build_summary,
    bucket_for,
    compute_cutoff,
    generate_excel,
    load_detail,
)


def _make_row(**over):
    """A full 20-column detail row keyed by header name; override any field."""
    base = {
        "Material": "10001334",
        "Material Description": "LC WHITE HOMINY",
        "Material Group": "D28",
        "Material Group Desc.": "LA COSTENA",
        "Purchasing Group": "104",
        "Description p. group": "Bita Farahani",
        "Brand Manager": "10",
        "Brand Manager Desc": "NANCY QUISPE PT",
        "MRP Area": "2920",
        "Batch": "OPAUG2726",
        "SLED Offset in days": -60,
        "Batch Expiry Date": datetime(2026, 8, 27),
        "MRP Last Sell Date": datetime(2026, 6, 28),
        "Quantity": 10,
        "Base Unit of Measure": "CS",
        "Qty Val. UoM": 10,
        "Total Stock": 10,
        "Moving price": 32.59,
        "Value": 325.90,
        "Batch Comment": None,
    }
    base.update(over)
    return [base[h] for h in DETAIL_HEADERS]


# Per-column number formats matching the real SAP export, keyed by header.
_FIXTURE_FORMATS = {
    "SLED Offset in days": "#,##0",
    "Batch Expiry Date": "mm-dd-yy",
    "MRP Last Sell Date": "mm-dd-yy",
    "Quantity": "#,##0.000",
    "Qty Val. UoM": "#,##0.000",
    "Total Stock": "#,##0.000",
    "Moving price": "#,##0.00",
    "Value": "#,##0.00",
}


def _make_xlsx(rows, *, sheet="Sheet1", headers=None):
    headers = headers if headers is not None else DETAIL_HEADERS
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for r in rows:
        ws.append(r)
    # Mirror the real export's column number formats so format passthrough is
    # exercised faithfully.
    for c, name in enumerate(headers, 1):
        fmt = _FIXTURE_FORMATS.get(name)
        if fmt:
            for r in range(2, ws.max_row + 1):
                ws.cell(r, c).number_format = fmt
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ---------------------------------------------------------------------------
# load_detail
# ---------------------------------------------------------------------------
def test_load_detail_reads_headers_and_rows():
    d = load_detail(_make_xlsx([_make_row(), _make_row(Batch="X2")]))
    assert d.headers == DETAIL_HEADERS
    assert len(d) == 2


def test_load_detail_rejects_missing_sheet1():
    wb = openpyxl.Workbook()
    wb.active.title = "Other"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    with pytest.raises(RiskyInventoryError):
        load_detail(buf)


def test_load_detail_rejects_wrong_headers():
    bad = ["Material", "Oops"] + DETAIL_HEADERS[2:]
    with pytest.raises(RiskyInventoryError):
        load_detail(_make_xlsx([_make_row()], headers=bad))


def test_load_detail_skips_blank_rows():
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Sheet1"
    ws.append(DETAIL_HEADERS)
    ws.append(_make_row())
    ws.append([None] * len(DETAIL_HEADERS))  # trailing blank
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    assert len(load_detail(buf)) == 1


# ---------------------------------------------------------------------------
# build_summary
# ---------------------------------------------------------------------------
def test_summary_structure_grouping_sort_and_grand_total():
    rows = [
        _make_row(**{"Material Group Desc.": "ALPHA", "Quantity": 5, "Total Stock": 5, "Value": 100}),
        _make_row(**{"Material Group Desc.": "ALPHA", "Quantity": 5, "Total Stock": 5, "Value": 50}),
        _make_row(**{"Material Group Desc.": "BETA", "Quantity": 1, "Total Stock": 2, "Value": 300}),
    ]
    grid = build_summary(load_detail(_make_xlsx(rows)))

    # Three filter rows, blank, header row
    assert grid[0][0] == "Description p. group" and grid[0][1] == "(All)"
    assert grid[3] == [None] * len(SUMMARY_HEADER)
    assert grid[4] == SUMMARY_HEADER

    body = grid[5:-1]
    labels = [r[0] for r in body]
    assert labels == ["BETA", "ALPHA"]              # sorted by Sum of Value desc
    beta = body[0]
    assert (beta[5], beta[6], beta[7]) == (1, 2, 300)
    alpha = body[1]
    assert (alpha[5], alpha[6], alpha[7]) == (10, 10, 150)  # summed

    total = grid[-1]
    assert total[0] == GRAND_TOTAL_LABEL
    assert (total[5], total[6], total[7]) == (11, 12, 450)


def test_summary_value_tie_breaks_by_group_name_ascending():
    rows = [
        _make_row(**{"Material Group Desc.": "WESTKEY", "Value": 0, "Quantity": 1, "Total Stock": 1}),
        _make_row(**{"Material Group Desc.": "CARRS", "Value": 0, "Quantity": 1, "Total Stock": 1}),
    ]
    grid = build_summary(load_detail(_make_xlsx(rows)))
    labels = [r[0] for r in grid[5:-1]]
    assert labels == ["CARRS", "WESTKEY"]


# ---------------------------------------------------------------------------
# generate_excel
# ---------------------------------------------------------------------------
def test_generate_excel_sheet_order_and_detail_preserved():
    d90 = load_detail(_make_xlsx([_make_row(Material="A")]))
    d180_clean = load_detail(_make_xlsx([_make_row(Material="Z")]))

    data = generate_excel(d90, d180_clean)
    wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    assert wb.sheetnames == ["90D Detail", "90D Summary", "180D Detail", "180D Summary"]

    det = wb["180D Detail"]
    assert [det.cell(1, c).value for c in range(1, len(DETAIL_HEADERS) + 1)] == DETAIL_HEADERS
    # only the Z row is in the 180D detail
    assert det.max_row == 2
    assert det.cell(2, 1).value == "Z"


def test_generate_excel_applies_number_formats():
    d = load_detail(_make_xlsx([_make_row()]))
    data = generate_excel(d, d)
    wb = openpyxl.load_workbook(io.BytesIO(data))
    det = wb["90D Detail"]
    # Batch Expiry Date / Quantity / Value column formats
    assert det.cell(2, DETAIL_HEADERS.index("Batch Expiry Date") + 1).number_format == "mm-dd-yy"
    assert det.cell(2, DETAIL_HEADERS.index("Quantity") + 1).number_format == "#,##0.000"
    assert det.cell(2, DETAIL_HEADERS.index("Value") + 1).number_format == "#,##0.00"
    summ = wb["90D Summary"]
    assert summ.cell(6, 7).number_format == "#,##0"      # Sum of Total Stock
    assert summ.cell(6, 8).number_format == '"$"#,##0'   # Sum of Value


# ---------------------------------------------------------------------------
# Bucketing
# ---------------------------------------------------------------------------
def test_compute_cutoff_is_run_date_plus_90():
    assert compute_cutoff(date(2026, 6, 24)) == date(2026, 9, 22)


def test_bucket_for_boundaries():
    cutoff = date(2026, 9, 22)
    assert bucket_for(datetime(2026, 9, 22), cutoff) == BUCKET_0_90   # inclusive
    assert bucket_for(datetime(2026, 9, 21), cutoff) == BUCKET_0_90
    assert bucket_for(datetime(2026, 9, 23), cutoff) == BUCKET_91_180
    assert bucket_for(None, cutoff) == BUCKET_NONE
    assert bucket_for("", cutoff) == BUCKET_NONE


def test_assign_buckets_appends_column_and_counts():
    rows = [
        _make_row(Material="A", **{"MRP Last Sell Date": datetime(2026, 8, 1)}),   # 0-90
        _make_row(Material="B", **{"MRP Last Sell Date": datetime(2026, 12, 1)}),  # 91-180
        _make_row(Material="C", **{"MRP Last Sell Date": None}),                   # none
    ]
    detail = load_detail(_make_xlsx(rows))
    bucketed, counts = assign_buckets(detail, compute_cutoff(date(2026, 6, 24)))
    assert bucketed.headers[-1] == "Bucket"
    assert [r[-1] for r in bucketed.rows] == [BUCKET_0_90, BUCKET_91_180, BUCKET_NONE]
    assert counts == {BUCKET_0_90: 1, BUCKET_91_180: 1, BUCKET_NONE: 1}
    assert len(bucketed.rows) == 3 and len(bucketed.rows[0]) == len(detail.headers) + 1
