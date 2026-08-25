"""Tests for the EWM Dispose list engine.

Self-contained: every fixture workbook is built in memory, so these tests do not
depend on the supplied sample files.

The engine deliberately applies only the one derivable rule (exclude Product
numbers starting "40"). The finished workbook it was reverse-engineered from was
hand-cleaned differently on each sheet — see the module docstring — so the tests
below pin the intended behaviour, not that workbook's inconsistencies.
"""
from __future__ import annotations

import io

import openpyxl
import pandas as pd
import pytest

from datetime import date

from src.dispose_ewm_engine import (
    DATE_COLS,
    EWM_BATCH,
    EWM_BIN,
    EWM_OWNER,
    EWM_PRODUCT,
    MATDISP_LAST_SELL_DATE,
    MATDISP_MATERIAL,
    MATDISP_PLANT,
    MATERIALS_PLANTS,
    OUT_BDM,
    PLANTS,
    SLED_CUTOFF_OFFSET_DAYS,
    DisposeEwmError,
    build_dispose_ewm,
    generate_excel,
    load_ewm,
    load_master,
    load_materials_dispose,
)

# A trimmed but faithful column set: the real export has 54, and the engine
# keeps whatever it is given rather than a fixed list.
EWM_HEADERS = [
    "Storage Type", EWM_BIN, EWM_PRODUCT, "Product Short Description",
    "Quantity", "Stock Type", "Description of Stock Type", EWM_BATCH,
    "Stock Segment", "Owner", EWM_OWNER, "Shelf Life Expiration Date",
    "Document Category", "Consolidation Group",
]


def _ewm_book(rows, owner="BP2910"):
    """An EWM dispose export as bytes. ``rows`` are (product, batch, bin)
    tuples or full dicts keyed by header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(EWM_HEADERS)
    for row in rows:
        data = {h: None for h in EWM_HEADERS}
        if isinstance(row, dict):
            data.update(row)
        else:
            product, batch, bin_ = row
            data.update({EWM_PRODUCT: product, EWM_BATCH: batch, EWM_BIN: bin_})
        data[EWM_OWNER] = data[EWM_OWNER] or owner
        ws.append([data[h] for h in EWM_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _master_book(rows):
    """A Last Sell / BDM master. ``rows`` are (product, brand manager) or
    (product, brand manager, last_sell_day)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Plant", "Product Number", "Brand Manager Name", "Last Sell Day"])
    for row in rows:
        product, bdm, *rest = row
        last_sell_day = rest[0] if rest else 30
        ws.append(["2910", product, bdm, last_sell_day])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


MATDISP_HEADERS = [
    "Material", "Material Description", "Plant", "Plant Name", "BDM",
    "Storage Location", "Description of Storage Location", "Batch",
    "Shelf Life Expiration Date", "Last sell day", "Last sell date",
    "Special Stock Type Description", "Unrestricted Stock",
    "Stock in Quality Inspection", "Blocked Stock",
]


def _materials_dispose_book(rows):
    """The 2925/2935 materials-based dispose export as bytes. ``rows`` are
    (material, plant, unrestricted_stock) tuples or full dicts keyed by
    header."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(MATDISP_HEADERS)
    for row in rows:
        data = {h: None for h in MATDISP_HEADERS}
        if isinstance(row, dict):
            data.update(row)
        else:
            material, plant, unrestricted = row
            data.update({
                "Material": material, "Plant": plant,
                "Unrestricted Stock": unrestricted,
                "Stock in Quality Inspection": 0, "Blocked Stock": 0,
            })
        ws.append([data[h] for h in MATDISP_HEADERS])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build(rows, master=None, plant="2910"):
    ewm = {plant: load_ewm(_ewm_book(rows, owner=f"BP{plant}"), expected_plant=plant)}
    lut = load_master(_master_book(master)) if master is not None else None
    return build_dispose_ewm(ewm, lut)


# ---------------------------------------------------------------------------
# The packaging exclusion — the only selection rule
# ---------------------------------------------------------------------------
def test_products_starting_40_are_excluded():
    sheets = _build([("10026065", "B1", "WDA58D"), ("40048660", "250907", "GLA16H")])
    assert sheets["2910"][EWM_PRODUCT].tolist() == ["10026065"]


def test_everything_else_is_kept():
    """No SLED window, no storage-location filter, no stock-type filter — the
    export is already scoped by whoever ran it."""
    rows = [
        {EWM_PRODUCT: "10015930", EWM_BATCH: "B1", EWM_BIN: "SHIPPING",
         "Storage Type": "9020", "Stock Type": "F2", "Document Category": "PDO",
         "Shelf Life Expiration Date": "2021-06-13"},
        {EWM_PRODUCT: "10026065", EWM_BATCH: "B2", EWM_BIN: "GA60E",
         "Storage Type": "1001", "Stock Type": "F3",
         "Shelf Life Expiration Date": "2026-07-26"},
    ]
    assert len(_build(rows)["2910"]) == 2


def test_shipping_allocated_rows_are_kept():
    """The finished workbook dropped these on 2920 but kept an identical one on
    2910. Inconsistent by hand, so the engine keeps them — see module docstring.
    If they should go, that is a new rule to add deliberately."""
    rows = [{EWM_PRODUCT: "10015930", EWM_BATCH: "B110325", EWM_BIN: "SHIPPING",
             "Storage Type": "9020", "Stock Type": "F2",
             "Document Category": "PDO", "Consolidation Group": "0005000493"}]
    assert len(_build(rows)["2910"]) == 1


def test_original_row_order_is_preserved():
    """The warehouse's ordering is meaningful; nothing here re-sorts it."""
    rows = [("10000003", "B", "Z9"), ("10000001", "A", "A1"), ("10000002", "C", "M5")]
    assert _build(rows)["2910"][EWM_PRODUCT].tolist() == [
        "10000003", "10000001", "10000002",
    ]


# ---------------------------------------------------------------------------
# The BDM column
# ---------------------------------------------------------------------------
def test_bdm_is_looked_up_from_the_master():
    sheets = _build([("10026065", "B1", "X")], master=[("10026065", "AMOL PRAKASH")])
    assert sheets["2910"][OUT_BDM].tolist() == ["AMOL PRAKASH"]


def test_bdm_sits_immediately_after_batch():
    sheets = _build([("10026065", "B1", "X")], master=[("10026065", "A")])
    cols = list(sheets["2910"].columns)
    assert cols[cols.index(EWM_BATCH) + 1] == OUT_BDM


def test_every_source_column_survives():
    sheets = _build([("10026065", "B1", "X")], master=[("10026065", "A")])
    cols = list(sheets["2910"].columns)
    assert all(h in cols for h in EWM_HEADERS)
    assert len(cols) == len(EWM_HEADERS) + 1


def test_product_missing_from_the_master_gets_a_blank_bdm():
    sheets = _build([("10026065", "B1", "X")], master=[("99999999", "SOMEONE")])
    assert pd.isna(sheets["2910"][OUT_BDM].iloc[0])


def test_master_repeating_a_product_across_vendors_takes_the_first():
    sheets = _build(
        [("10026065", "B1", "X")],
        master=[("10026065", "AMOL PRAKASH"), ("10026065", "AMOL PRAKASH")],
    )
    assert sheets["2910"][OUT_BDM].tolist() == ["AMOL PRAKASH"]


def test_bdm_column_exists_even_with_no_master():
    sheets = _build([("10026065", "B1", "X")])
    assert OUT_BDM in sheets["2910"].columns
    assert sheets["2910"][OUT_BDM].isna().all()


# ---------------------------------------------------------------------------
# Inputs
# ---------------------------------------------------------------------------
def test_file_uploaded_for_the_wrong_plant_is_rejected():
    with pytest.raises(DisposeEwmError, match="uploaded in the 2910 box"):
        load_ewm(_ewm_book([("1", "A", "X")], owner="BP2920"), expected_plant="2910")


def test_a_non_ewm_file_is_rejected_by_name():
    wb = openpyxl.Workbook()
    wb.active.append(["Something", "Else"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(DisposeEwmError, match="EWM dispose"):
        load_ewm(buf, expected_plant="2910")


def test_only_the_plants_supplied_get_a_sheet():
    ewm = {p: load_ewm(_ewm_book([("10026065", "B1", "X")], owner=f"BP{p}"),
                       expected_plant=p) for p in ("2910", "2930")}
    assert list(build_dispose_ewm(ewm)) == ["2910", "2930"]


def test_sheets_come_out_in_plant_order():
    ewm = {p: load_ewm(_ewm_book([("10026065", "B1", "X")], owner=f"BP{p}"),
                       expected_plant=p) for p in reversed(PLANTS)}
    assert list(build_dispose_ewm(ewm)) == PLANTS


def test_no_files_at_all_is_an_error():
    with pytest.raises(DisposeEwmError, match="at least one"):
        build_dispose_ewm({})


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
def test_sheets_are_named_by_plant_code():
    ewm = {p: load_ewm(_ewm_book([("10026065", "B1", "X")], owner=f"BP{p}"),
                       expected_plant=p) for p in PLANTS}
    wb = openpyxl.load_workbook(io.BytesIO(generate_excel(build_dispose_ewm(ewm))))
    assert wb.sheetnames == PLANTS


def test_dates_are_written_as_dates_not_text():
    rows = [{EWM_PRODUCT: "10026065", EWM_BATCH: "B1", EWM_BIN: "X",
             "Shelf Life Expiration Date": "2026-07-26"}]
    ws = openpyxl.load_workbook(io.BytesIO(generate_excel(_build(rows))))["2910"]
    headers = [c.value for c in ws[1]]
    cell = ws.cell(row=2, column=headers.index(DATE_COLS[0]) + 1)
    assert cell.is_date
    assert cell.number_format == "mm-dd-yy"


def test_ids_keep_their_leading_zeros():
    """Consolidation Group "0005000242" must not be coerced to a number."""
    rows = [{EWM_PRODUCT: "10026065", EWM_BATCH: "B1", EWM_BIN: "X",
             "Consolidation Group": "0005000242"}]
    ws = openpyxl.load_workbook(io.BytesIO(generate_excel(_build(rows))))["2910"]
    headers = [c.value for c in ws[1]]
    assert ws.cell(row=2, column=headers.index("Consolidation Group") + 1).value \
        == "0005000242"


def test_autofilter_covers_every_column():
    ws = openpyxl.load_workbook(
        io.BytesIO(generate_excel(_build([("10026065", "B1", "X")])))
    )["2910"]
    last = openpyxl.utils.get_column_letter(len(EWM_HEADERS) + 1)
    assert ws.auto_filter.ref == f"A1:{last}2"


# ---------------------------------------------------------------------------
# Materials-based dispose export — plants 2925 / 2935 (no EWM system)
# ---------------------------------------------------------------------------
def test_materials_dispose_splits_into_two_sheets_by_plant():
    md = load_materials_dispose(_materials_dispose_book([
        ("10076882", "2925", 28), ("10015386", "2935", 5),
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md)
    assert sheets["2925"][MATDISP_MATERIAL].tolist() == ["10076882"]
    assert sheets["2935"][MATDISP_MATERIAL].tolist() == ["10015386"]


def test_materials_dispose_both_plants_get_a_sheet_even_if_one_is_empty():
    md = load_materials_dispose(_materials_dispose_book([("10076882", "2925", 28)]))
    sheets = build_dispose_ewm({}, materials_dispose=md)
    assert list(sheets) == MATERIALS_PLANTS
    assert sheets["2935"].empty


def test_materials_dispose_packaging_exclusion_applies():
    md = load_materials_dispose(_materials_dispose_book([
        ("10076882", "2925", 28), ("40048660", "2925", 5),
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md)
    assert sheets["2925"][MATDISP_MATERIAL].tolist() == ["10076882"]


def test_materials_dispose_bdm_already_present_is_kept_as_is():
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925", "BDM": "AMOL PRAKASH",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0,
         "Blocked Stock": 0},
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md)
    assert sheets["2925"][OUT_BDM].tolist() == ["AMOL PRAKASH"]


def test_materials_dispose_alone_is_enough_to_build():
    md = load_materials_dispose(_materials_dispose_book([("10076882", "2925", 28)]))
    sheets = build_dispose_ewm({}, materials_dispose=md)
    assert sheets["2925"][MATDISP_MATERIAL].tolist() == ["10076882"]


def test_ewm_and_materials_dispose_sheets_combine_in_plant_order():
    ewm = {p: load_ewm(_ewm_book([("10026065", "B1", "X")], owner=f"BP{p}"),
                       expected_plant=p) for p in PLANTS}
    md = load_materials_dispose(
        _materials_dispose_book([("10076882", "2925", 28), ("10015386", "2935", 5)])
    )
    sheets = build_dispose_ewm(ewm, materials_dispose=md)
    assert list(sheets) == PLANTS + MATERIALS_PLANTS


def test_a_non_materials_dispose_file_is_rejected_by_name():
    wb = openpyxl.Workbook()
    wb.active.append(["Something", "Else"])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(DisposeEwmError, match="Materials dispose"):
        load_materials_dispose(buf)


# ---------------------------------------------------------------------------
# Materials-based dispose date filter (2925 / 2935 only)
# ---------------------------------------------------------------------------
def test_materials_dispose_cutoff_keeps_rows_with_both_dates_in_window():
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-20",
         "Last sell date": "2026-08-15",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md, materials_cutoff=date(2026, 8, 20))
    assert sheets["2925"][MATDISP_MATERIAL].tolist() == ["10076882"]


def test_materials_dispose_cutoff_drops_row_when_sled_is_past_the_window():
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-26",  # cutoff+4 = 08-24
         "Last sell date": "2026-08-15",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md, materials_cutoff=date(2026, 8, 20))
    assert sheets["2925"].empty


def test_materials_dispose_cutoff_requires_both_dates_not_either():
    """AND, not OR: Last sell date past the window drops the row even though
    SLED is well inside it."""
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-15",
         "Last sell date": "2026-08-26",  # past cutoff+4 = 08-24
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md, materials_cutoff=date(2026, 8, 20))
    assert sheets["2925"].empty


def test_materials_dispose_cutoff_offset_boundary_is_inclusive():
    assert SLED_CUTOFF_OFFSET_DAYS == 4
    boundary = date(2026, 8, 24)  # 2026-08-20 + 4
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "ON_BOUNDARY", "Plant": "2925",
         "Shelf Life Expiration Date": boundary.isoformat(),
         "Last sell date": boundary.isoformat(),
         "Unrestricted Stock": 1, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
        {"Material": "PAST_BOUNDARY", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-25",
         "Last sell date": "2026-08-25",
         "Unrestricted Stock": 1, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md, materials_cutoff=date(2026, 8, 20))
    assert sheets["2925"][MATDISP_MATERIAL].tolist() == ["ON_BOUNDARY"]


def test_materials_dispose_cutoff_drops_rows_missing_a_date():
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-15",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md, materials_cutoff=date(2026, 8, 20))
    assert sheets["2925"].empty


def test_materials_dispose_no_cutoff_means_no_date_filtering():
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2099-01-01",
         "Last sell date": "2099-01-01",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md)
    assert sheets["2925"][MATDISP_MATERIAL].tolist() == ["10076882"]


def test_materials_dispose_last_sell_date_is_computed_from_the_master_when_missing():
    """A raw export with no Last sell date column populated: computed as
    SLED - Last Sell Day from the master, then the cutoff applies to it."""
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-20",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    lut = load_master(_master_book([("10076882", "AMOL PRAKASH", 5)]))
    sheets = build_dispose_ewm(
        {}, lut, materials_dispose=md, materials_cutoff=date(2026, 8, 20)
    )
    assert sheets["2925"][MATDISP_MATERIAL].tolist() == ["10076882"]
    assert sheets["2925"][MATDISP_LAST_SELL_DATE].iloc[0] == pd.Timestamp("2026-08-15")
    assert sheets["2925"][OUT_BDM].iloc[0] == "AMOL PRAKASH"


def test_materials_dispose_computed_last_sell_date_can_fail_the_cutoff():
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-20",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    # Last Sell Day = -10 => Last sell date = 2026-08-30, past cutoff+4 = 08-24
    lut = load_master(_master_book([("10076882", "AMOL PRAKASH", -10)]))
    sheets = build_dispose_ewm(
        {}, lut, materials_dispose=md, materials_cutoff=date(2026, 8, 20)
    )
    assert sheets["2925"].empty


def test_materials_dispose_without_master_cannot_compute_last_sell_date():
    """No master + no Last sell date on the row + a cutoff => can't be
    evaluated, so the row is dropped. This is the scenario the page warns
    about when only the materials file is uploaded."""
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-15",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    sheets = build_dispose_ewm({}, materials_dispose=md, materials_cutoff=date(2026, 8, 20))
    assert sheets["2925"].empty


def test_materials_dispose_existing_last_sell_date_is_not_recomputed():
    """The file already carries Last sell date (older, pre-filtered shape) —
    that value wins over what the master's Last Sell Day would compute."""
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Shelf Life Expiration Date": "2026-08-20",
         "Last sell date": "2026-08-01",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    # Master's Last Sell Day=5 would compute 2026-08-15, but the existing
    # 2026-08-01 value must win.
    lut = load_master(_master_book([("10076882", "AMOL PRAKASH", 5)]))
    sheets = build_dispose_ewm(
        {}, lut, materials_dispose=md, materials_cutoff=date(2026, 8, 20)
    )
    assert sheets["2925"][MATDISP_LAST_SELL_DATE].iloc[0] == pd.Timestamp("2026-08-01")


def test_materials_dispose_bdm_is_filled_from_master_when_blank():
    md = load_materials_dispose(_materials_dispose_book([
        {"Material": "10076882", "Plant": "2925",
         "Unrestricted Stock": 28, "Stock in Quality Inspection": 0, "Blocked Stock": 0},
    ]))
    lut = load_master(_master_book([("10076882", "AMOL PRAKASH")]))
    sheets = build_dispose_ewm({}, lut, materials_dispose=md)
    assert sheets["2925"][OUT_BDM].tolist() == ["AMOL PRAKASH"]
