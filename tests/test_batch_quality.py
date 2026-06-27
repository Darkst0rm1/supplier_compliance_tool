"""Tests for the Batch Quality Analysis engine (reworked)."""
from __future__ import annotations

import importlib
import io
import py_compile
from datetime import datetime
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import Workbook, load_workbook

from src.batch_quality import ai_agent, exporter, loader, normalization, reviews, rules
from src.batch_quality.issue_groups import (
    GROUP_COLUMNS,
    _dedup_hits,
    analyze,
    build_related_records,
    build_results_table,
)
from src.batch_quality.loader import (
    COL_BATCH,
    COL_BATCH_SLED,
    COL_MATERIAL,
    COL_PLANT,
    COL_PO,
    COL_PURCH_GROUP,
    COL_QTY,
    COL_RECEIVED,
    COL_SUPPLIER,
    COL_SUPPLIER_NAME,
    COL_VENDOR,
    COL_VENDOR_NAME,
    NORMALIZED_BATCH,
)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------
def _xlsx(headers: list[str], rows: list[list]) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = loader.SAP_SHEET
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


_CANON_COLS = [
    COL_MATERIAL, "Material Desc", COL_BATCH, COL_BATCH_SLED, COL_PO,
    COL_PLANT, COL_SUPPLIER, COL_SUPPLIER_NAME, COL_VENDOR, COL_VENDOR_NAME,
    COL_PURCH_GROUP, COL_RECEIVED, COL_QTY,
]


def make_df(records: list[dict]) -> pd.DataFrame:
    """Build a DataFrame shaped like loader output from concise record dicts."""
    rows = []
    for i, rec in enumerate(records):
        rows.append({
            COL_MATERIAL: str(rec.get("material", "1000")),
            "Material Desc": rec.get("desc", "WIDGET"),
            COL_BATCH: rec.get("batch", "B1"),
            COL_BATCH_SLED: rec.get("sled"),
            COL_PO: str(rec.get("po", f"PO{i}")),
            COL_PLANT: str(rec.get("plant", "2910")),
            COL_SUPPLIER: str(rec.get("supplier", "S1")),
            COL_SUPPLIER_NAME: rec.get("supplier_name", "Supplier One"),
            COL_VENDOR: str(rec.get("vendor", "V1")),
            COL_VENDOR_NAME: rec.get("vendor_name", "Vendor One"),
            COL_PURCH_GROUP: str(rec.get("pgroup", "PG1")),
            COL_RECEIVED: rec.get("received"),
            COL_QTY: rec.get("qty", 1.0),
        })
    df = pd.DataFrame(rows, columns=_CANON_COLS)
    df[COL_BATCH_SLED] = pd.to_datetime(df[COL_BATCH_SLED], errors="coerce")
    df[COL_RECEIVED] = pd.to_datetime(df[COL_RECEIVED], errors="coerce")
    df[NORMALIZED_BATCH] = df[COL_BATCH].map(loader.normalize_batch)
    return df.reset_index(drop=True)


# ---------------------------------------------------------------------------
# 1-4: Loader / normalization
# ---------------------------------------------------------------------------
def test_load_workbook_and_headers_and_leading_zeros():
    buf = _xlsx(
        ["Material", "Material Description", "Batch", "Batch SLED",
         "Purchase Order", "Plant", "Qty Base"],
        [
            [10038128.0, "SCHAR", "00009926", datetime(2027, 1, 9), 1000006096, 2910, 675],
            [10038095.0, "HONEY", "310327A", datetime(2027, 3, 31), 1000006096, 2910, 507],
        ],
    )
    df = loader.load_batch_data(buf)
    # Header normalization (aliases mapped to canonical names)
    assert COL_PO in df.columns and "Material Desc" in df.columns
    # Leading zeros preserved, trailing .0 stripped
    assert df[COL_BATCH].iloc[0] == "00009926"
    assert df[COL_MATERIAL].iloc[0] == "10038128"
    # Dates parsed
    assert pd.api.types.is_datetime64_any_dtype(df[COL_BATCH_SLED])
    assert df[COL_BATCH_SLED].iloc[0] == pd.Timestamp(2027, 1, 9)
    assert df[NORMALIZED_BATCH].iloc[1] == "310327A"


def test_load_missing_columns_raises():
    buf = _xlsx(["Foo", "Bar"], [[1, 2]])
    with pytest.raises(loader.BatchQualityError):
        loader.load_batch_data(buf)


def test_normalize_batch_variants():
    assert normalization.normalize_batch("31-5357") == "315357"
    assert normalization.normalize_batch("31 5357") == "315357"
    assert normalization.normalize_batch("31/5357") == "315357"
    assert normalization.normalize_batch("31.53_57") == "315357"
    assert normalization.normalize_batch("  abc1  ") == "ABC1"
    assert normalization.normalize_batch(None) == ""
    assert normalization.normalize_batch(float("nan")) == ""
    # loader re-exports the same function
    assert loader.normalize_batch is normalization.normalize_batch


# ---------------------------------------------------------------------------
# 5-6: Format variation / conflicting expiry
# ---------------------------------------------------------------------------
def test_rule_format_variation():
    df = make_df([
        {"material": "M1", "batch": "31-5357", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "315357", "sled": "2027-01-09", "po": "PO2"},
    ])
    hits = rules.rule_batch_groups(df)
    assert len(hits) == 1
    assert hits[0].issue_type == rules.ISSUE_FORMAT_VARIATION
    assert hits[0].risk_level == rules.RISK_MEDIUM
    assert set(hits[0].indices) == {0, 1}


def test_rule_conflicting_expiry_takes_priority():
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
    ])
    hits = rules.rule_batch_groups(df)
    assert len(hits) == 1
    assert hits[0].issue_type == rules.ISSUE_CONFLICTING_EXPIRY
    assert hits[0].risk_level == rules.RISK_HIGH


# ---------------------------------------------------------------------------
# 7-9: Character-confusion detection
# ---------------------------------------------------------------------------
def test_rule_character_entry_basic():
    df = make_df([
        {"material": "M1", "batch": "NI8D61", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "N18D61", "sled": "2027-01-09", "po": "PO2"},
    ])
    hits = rules.rule_character_entry(df)
    assert len(hits) == 1
    assert hits[0].issue_type == rules.ISSUE_CHARACTER_ENTRY
    assert set(hits[0].indices) == {0, 1}


def test_character_entry_I_vs_1():
    assert normalization.is_confusable("I", "1")
    assert normalization.confusable_diff_positions("AI1", "A11") == [1]


def test_character_entry_O_vs_0():
    assert normalization.is_confusable("O", "0")
    df = make_df([
        {"material": "M2", "batch": "OO12", "po": "PO1", "sled": "2027-01-09"},
        {"material": "M2", "batch": "0O12", "po": "PO2", "sled": "2027-01-09"},
    ])
    hits = rules.rule_character_entry(df)
    assert len(hits) == 1
    assert hits[0].issue_type == rules.ISSUE_CHARACTER_ENTRY


def test_character_entry_ignores_unrelated_batches():
    # Different non-confusable characters must NOT be flagged as a mis-read.
    df = make_df([
        {"material": "M3", "batch": "ABCD", "po": "PO1", "sled": "2027-01-09"},
        {"material": "M3", "batch": "WXYZ", "po": "PO2", "sled": "2027-01-09"},
    ])
    assert rules.rule_character_entry(df) == []


# ---------------------------------------------------------------------------
# 10-11: Same material + supplier received close together, pattern conflict
# ---------------------------------------------------------------------------
def test_rule_pattern_conflict_flags_close_receipts():
    df = make_df([
        {"material": "10001310", "supplier": "S9", "batch": "OPOCT2227",
         "sled": "2029-08-13", "received": "2026-01-01", "po": "PO1", "plant": "2910"},
        {"material": "10001310", "supplier": "S9", "batch": "2027AU13",
         "sled": "2027-08-13", "received": "2026-01-05", "po": "PO2", "plant": "2920"},
    ])
    hits = rules.rule_pattern_conflict(df)
    assert len(hits) == 1
    assert hits[0].issue_type == rules.ISSUE_PATTERN_CONFLICT
    assert hits[0].risk_level == rules.RISK_HIGH  # ~2yr expiry gap
    assert set(hits[0].indices) == {0, 1}


def test_pattern_conflict_respects_window_and_expiry():
    # Same batch-structure difference but received far apart -> not flagged.
    df = make_df([
        {"material": "M1", "supplier": "S1", "batch": "OPOCT2227",
         "sled": "2029-08-13", "received": "2026-01-01", "po": "PO1"},
        {"material": "M1", "supplier": "S1", "batch": "2027AU13",
         "sled": "2027-08-13", "received": "2026-03-01", "po": "PO2"},
    ])
    assert rules.rule_pattern_conflict(df) == []
    # Close, but expiry nearly identical -> not significant.
    df2 = make_df([
        {"material": "M1", "supplier": "S1", "batch": "OPOCT2227",
         "sled": "2027-08-13", "received": "2026-01-01", "po": "PO1"},
        {"material": "M1", "supplier": "S1", "batch": "2027AU13",
         "sled": "2027-08-14", "received": "2026-01-05", "po": "PO2"},
    ])
    assert rules.rule_pattern_conflict(df2) == []


# ---------------------------------------------------------------------------
# 12-14: Duplicate receiving / exact duplicate / format concerns
# ---------------------------------------------------------------------------
def test_rule_duplicate_receiving():
    df = make_df([
        {"material": "M1", "batch": "B1", "sled": "2027-01-09", "po": "PO9", "qty": 5.0},
        {"material": "M1", "batch": "B1", "sled": "2027-01-09", "po": "PO9", "qty": 7.0},
    ])
    hits = rules.rule_duplicate_receiving(df)
    assert len(hits) == 1
    assert hits[0].issue_type == rules.ISSUE_DUPLICATE_RECEIVING


def test_rule_exact_duplicate():
    df = make_df([
        {"material": "M1", "batch": "B1", "sled": "2027-01-09", "po": "PO9", "qty": 5.0},
        {"material": "M1", "batch": "B1", "sled": "2027-01-09", "po": "PO9", "qty": 5.0},
    ])
    hits = rules.rule_exact_duplicate(df)
    assert len(hits) == 1
    assert hits[0].issue_type == rules.ISSUE_EXACT_DUPLICATE
    assert hits[0].risk_level == rules.RISK_HIGH


def test_rule_format_review():
    df = make_df([
        {"material": "M1", "batch": "B1 ", "sled": "2027-01-09", "po": "PO1"},   # trailing space
        {"material": "M2", "batch": "X", "sled": "2027-01-09", "po": "PO2"},      # too short
        {"material": "M3", "batch": "GOODBATCH", "sled": "2027-01-09", "po": "PO3"},
    ])
    hits = rules.rule_format_review(df)
    assert {h.issue_type for h in hits} == {rules.ISSUE_FORMAT_REVIEW}
    assert len(hits) == 2


# ---------------------------------------------------------------------------
# 15: Multiple-batches contextual table
# ---------------------------------------------------------------------------
def test_multiple_batches_summary():
    df = make_df([
        {"material": "M1", "batch": "B1", "sled": "2027-01-01"},
        {"material": "M1", "batch": "B2", "sled": "2027-02-01"},  # multi batch + multi date
        {"material": "M2", "batch": "B3", "sled": "2027-01-01"},
        {"material": "M2", "batch": "B4", "sled": "2027-01-01"},  # multi batch, single date
    ])
    mb = rules.build_multiple_batches(df)
    assert list(mb["Material"]) == ["M1"]
    assert mb.iloc[0]["Number of Batch Values"] == 2
    assert mb.iloc[0]["Number of Expiry Dates"] == 2


# ---------------------------------------------------------------------------
# 16-17: Issue-group dedup + linking original rows
# ---------------------------------------------------------------------------
def test_dedup_identical_hits():
    h1 = rules.RuleHit(rules.ISSUE_FORMAT_REVIEW, rules.RISK_LOW, "r", [0, 1])
    h2 = rules.RuleHit(rules.ISSUE_FORMAT_REVIEW, rules.RISK_LOW, "r-dup", [1, 0])
    h3 = rules.RuleHit(rules.ISSUE_EXACT_DUPLICATE, rules.RISK_HIGH, "x", [0, 1])
    out = _dedup_hits([h1, h2, h3])
    assert len(out) == 2  # h2 dropped (same type + same rows), h3 kept (diff type)


def test_analyze_issue_groups_columns_and_members():
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
        {"material": "M2", "batch": "31-5357", "sled": "2027-03-01", "po": "PO3"},
        {"material": "M2", "batch": "315357", "sled": "2027-03-01", "po": "PO4"},
    ])
    result = analyze(df)
    assert result.summary["total_issue_groups"] == 2
    assert list(result.flagged.columns) == GROUP_COLUMNS
    assert "Rule Risk Level" in result.flagged.columns
    # High risk sorts first
    assert result.flagged.iloc[0]["Rule Risk Level"] == "High"
    gid = result.flagged.iloc[0]["Issue Group ID"]
    assert gid == "IG-0001"
    assert set(result.members[gid]) == {0, 1}


def test_build_related_records_links_rows():
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
    ])
    result = analyze(df)
    related = build_related_records(result)
    assert "Issue Group ID" in related.columns
    assert len(related) == 2
    assert NORMALIZED_BATCH not in related.columns


# ---------------------------------------------------------------------------
# 18-19: AI agent context + schema
# ---------------------------------------------------------------------------
def test_ai_context_only_relevant_records():
    df = make_df([
        {"material": "A", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "A", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
        {"material": "B", "batch": "ZZZ9", "sled": "2027-05-01", "po": "PO3"},
        {"material": "C", "batch": "QQQ8", "sled": "2027-06-01", "po": "PO4"},
    ])
    result = analyze(df)
    gid = result.flagged.iloc[0]["Issue Group ID"]
    idx = result.members[gid]
    related = df.loc[idx]
    mat = related[COL_MATERIAL].iloc[0]
    material_records = df[df[COL_MATERIAL] == mat]
    nbs = list(related[NORMALIZED_BATCH].unique())
    normbatch_records = df[df[NORMALIZED_BATCH].isin(nbs)]

    grp_row = result.flagged.iloc[0].to_dict()
    ctx = ai_agent.build_ai_context(grp_row, related, material_records, normbatch_records)

    assert len(ctx["related_records"]) == 2
    for bucket in ("related_records", "material_records", "normalized_batch_records"):
        for rec in ctx[bucket]:
            assert rec["Material"] == "A"


def test_ai_schema_has_no_corrected_value_fields():
    fields = set(ai_agent.AIReviewResult.model_fields.keys())
    assert "ai_review_priority" in fields
    assert fields.isdisjoint(ai_agent.FORBIDDEN_FIELDS)
    for f in fields:
        assert "correct" not in f and "suggest" not in f and "replacement" not in f


def test_ai_fingerprint_stable_and_record_sensitive():
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
    ])
    fp1 = ai_agent.issue_fingerprint("X", df)
    fp2 = ai_agent.issue_fingerprint("X", df.copy())
    assert fp1 == fp2
    df2 = df.copy()
    df2.loc[0, COL_PO] = "PO-CHANGED"
    assert ai_agent.issue_fingerprint("X", df2) != fp1


def test_ai_config_helpers(monkeypatch):
    monkeypatch.delenv("ANTHROPIC_MODEL", raising=False)
    monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
    monkeypatch.delenv("BATCH_QUALITY_LLM_BASE_URL", raising=False)
    assert ai_agent.get_backend() == "anthropic"
    assert ai_agent.get_model() == ai_agent.DEFAULT_MODEL
    assert ai_agent.get_model("claude-sonnet-4-6") == "claude-sonnet-4-6"
    assert ai_agent.is_ai_configured() is False
    assert ai_agent.is_ai_configured("sk-test") is True


def test_openai_backend_switches_with_env(monkeypatch):
    monkeypatch.setenv("BATCH_QUALITY_LLM_BASE_URL", "https://router.huggingface.co/v1")
    monkeypatch.setenv("BATCH_QUALITY_LLM_MODEL", "meta-llama/Llama-3.1-8B-Instruct")
    assert ai_agent.get_backend() == "openai"
    assert ai_agent.get_model() == "meta-llama/Llama-3.1-8B-Instruct"
    assert ai_agent.is_ai_configured() is True  # base URL alone is enough (Ollama needs no key)


def test_extract_json_variants():
    assert ai_agent._extract_json('{"a": 1}') == '{"a": 1}'
    assert ai_agent._extract_json('```json\n{"a": 1}\n```') == '{"a": 1}'
    assert ai_agent._extract_json('Sure! {"a": 1} hope that helps') == '{"a": 1}'


def test_review_via_openai_parses_response(monkeypatch):
    import json as _json
    import requests
    monkeypatch.setenv("BATCH_QUALITY_LLM_BASE_URL", "https://router.huggingface.co/v1")
    monkeypatch.setenv("BATCH_QUALITY_LLM_API_KEY", "hf_test")
    monkeypatch.setenv("BATCH_QUALITY_LLM_MODEL", "meta-llama/Llama-3.1-8B-Instruct")
    body = {
        "ai_review_priority": "Low", "review_summary": "s", "pattern_identified": "p",
        "reason_for_review": "r", "recurring_pattern": "n", "records_involved": "2",
        "documents_to_verify": ["ASN"], "possible_root_causes": ["entry"],
        "recommended_review_steps": ["check"], "questions_for_supplier_or_receiver": ["q"],
        "review_note": "note",
    }

    class FakeResp:
        status_code = 200

        def raise_for_status(self):
            pass

        def json(self):
            return {"choices": [{"message": {"content": "```json\n" + _json.dumps(body) + "\n```"}}]}

    captured: dict = {}

    def fake_post(url, json=None, headers=None, timeout=None):
        captured.update(url=url, headers=headers, payload=json)
        return FakeResp()

    monkeypatch.setattr(requests, "post", fake_post)
    res = ai_agent.review_issue({"issue_group": {}})
    assert isinstance(res, ai_agent.AIReviewResult)
    assert res.ai_review_priority == "Low"
    assert captured["url"].endswith("/chat/completions")
    assert captured["headers"]["Authorization"] == "Bearer hf_test"
    assert captured["payload"]["model"] == "meta-llama/Llama-3.1-8B-Instruct"


def test_run_agent_caps_and_caches(monkeypatch):
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
        {"material": "M2", "batch": "31-5357", "sled": "2027-03-01", "po": "PO3"},
        {"material": "M2", "batch": "315357", "sled": "2027-03-01", "po": "PO4"},
    ])
    result = analyze(df)
    related = build_related_records(result)
    calls = {"n": 0}

    def fake_review(context, api_key=None, model=None):
        calls["n"] += 1
        return ai_agent.AIReviewResult(
            ai_review_priority="Medium", review_summary="s", pattern_identified="p",
            reason_for_review="r", recurring_pattern="n", records_involved="2",
            documents_to_verify=["ASN"], possible_root_causes=["entry"],
            recommended_review_steps=["check"], questions_for_supplier_or_receiver=["q"],
            review_note="note",
        )

    monkeypatch.setattr(ai_agent, "review_issue", fake_review)
    cache: dict = {}
    ai_map = ai_agent.run_agent(result, related, cache=cache, api_key="sk-test", max_issues=1)
    # Only the first (highest-risk) group reviewed; the rest skipped by the cap.
    assert calls["n"] == 1
    first_gid = result.flagged.iloc[0]["Issue Group ID"]
    assert isinstance(ai_map[first_gid], ai_agent.AIReviewResult)
    assert "skipped" in ai_map.values()
    # Re-run reuses the cache (no new call for the reviewed group).
    ai_agent.run_agent(result, related, cache=cache, api_key="sk-test", max_issues=1)
    assert calls["n"] == 1


def _fake_result(**over):
    base = dict(
        ai_review_priority="Medium", review_summary="s", pattern_identified="p",
        reason_for_review="r", recurring_pattern="n", records_involved="2",
        documents_to_verify=["ASN"], possible_root_causes=["entry"],
        recommended_review_steps=["check"], questions_for_supplier_or_receiver=["q"],
        review_note="note",
    )
    base.update(over)
    return ai_agent.AIReviewResult(**base)


def test_run_agent_reviews_all_groups_concurrently(monkeypatch):
    import threading
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
        {"material": "M2", "batch": "31-5357", "sled": "2027-03-01", "po": "PO3"},
        {"material": "M2", "batch": "315357", "sled": "2027-03-01", "po": "PO4"},
    ])
    result = analyze(df)
    related = build_related_records(result)
    lock = threading.Lock()
    calls = {"n": 0}

    def fake_review(context, api_key=None, model=None):
        with lock:
            calls["n"] += 1
        return _fake_result()

    monkeypatch.setattr(ai_agent, "review_issue", fake_review)
    ai_map = ai_agent.run_agent(result, related, cache={}, api_key="sk-test",
                                max_issues=None, concurrency=4)
    n_groups = result.summary["total_issue_groups"]
    assert calls["n"] == n_groups
    assert all(isinstance(v, ai_agent.AIReviewResult) for v in ai_map.values())
    assert len(ai_map) == n_groups


def test_run_agent_raises_on_total_failure(monkeypatch):
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
    ])
    result = analyze(df)
    related = build_related_records(result)

    def boom(context, api_key=None, model=None):
        raise RuntimeError("invalid x-api-key")

    monkeypatch.setattr(ai_agent, "review_issue", boom)
    with pytest.raises(RuntimeError):
        ai_agent.run_agent(result, related, cache={}, api_key="bad", concurrency=4)


# ---------------------------------------------------------------------------
# Results table merge (AI + human findings)
# ---------------------------------------------------------------------------
def test_build_results_table_marks_unavailable():
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
    ])
    result = analyze(df)
    table = build_results_table(result.flagged, {}, {})
    for col in ai_agent.AI_COLUMNS:
        assert col in table.columns
    for col in reviews.HUMAN_REVIEW_FIELDS:
        assert col in table.columns
    assert table.iloc[0]["AI Review Priority"] == "Not Available"


# ---------------------------------------------------------------------------
# 20: Human review capture
# ---------------------------------------------------------------------------
def test_save_and_get_review():
    store: dict = {}
    reviews.save_review(store, "IG-0001", {
        "Confirmed Issue": "Yes",
        "Root Cause": "Receiving data entry",
        "Reviewer Comment": "Looks like a typo.",
        "Ignored": "dropped",
    })
    saved = reviews.get_review(store, "IG-0001")
    assert saved["Confirmed Issue"] == "Yes"
    assert saved["Root Cause"] == "Receiving data entry"
    assert "Ignored" not in saved
    assert reviews.get_review(store, "IG-9999")["Confirmed Issue"] == ""


# ---------------------------------------------------------------------------
# 21: Three-sheet Excel export
# ---------------------------------------------------------------------------
def test_generate_excel_has_three_sheets():
    df = make_df([
        {"material": "M1", "batch": "ABC1", "sled": "2027-01-09", "po": "PO1"},
        {"material": "M1", "batch": "ABC1", "sled": "2027-02-09", "po": "PO2"},
    ])
    result = analyze(df)
    related = build_related_records(result)
    flagged_full = build_results_table(result.flagged, {}, {})
    xlsx = exporter.generate_excel(flagged_full, related, result.multi_batch)
    wb = load_workbook(io.BytesIO(xlsx))
    assert wb.sheetnames == ["Flagged Issues", "Related Records", "Multiple Batches"]
    # No AI Review sheet, and the Flagged Issues sheet carries the AI columns.
    assert "AI Review" not in wb.sheetnames
    header = [c.value for c in wb["Flagged Issues"][1]]
    assert "AI Review Priority" in header


# ---------------------------------------------------------------------------
# 22: App wiring + every page still loads (compiles)
# ---------------------------------------------------------------------------
def test_app_registers_batch_quality_page():
    content = Path("app.py").read_text(encoding="utf-8")
    assert "8_Batch_Quality_Analysis.py" in content


def test_all_pages_compile():
    for page in sorted(Path("pages").glob("*.py")):
        py_compile.compile(str(page), doraise=True)
    py_compile.compile("app.py", doraise=True)


def test_batch_quality_modules_import():
    for mod in ["loader", "normalization", "rules", "issue_groups",
                "ai_agent", "reviews", "exporter", "page"]:
        importlib.import_module(f"src.batch_quality.{mod}")
