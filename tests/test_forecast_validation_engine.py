"""Tests for the Forecast Validation engine.

Self-contained: every fixture workbook is built in memory.
"""
from __future__ import annotations

import io
from datetime import date, datetime

import openpyxl
import pandas as pd
import pytest

from src.forecast_validation_engine import (
    ForecastValidationError,
    build_data_quality,
    build_forecast_validation,
    build_main_table,
    build_monthly_summary,
    build_plant_summary,
    combine_sales_orders,
    default_history_and_forecast_months,
    generate_excel,
    load_forecast,
    load_open_orders,
    load_open_po,
    load_sales_orders,
    resolve_open_order_plants,
    same_period_window,
)

KEY_ACCOUNT_HEADERS = [
    "Sales Order", "Sales Order Item", "Key Account #", "Sales Order Type",
    "Creation Date", "Order Status", "Rejection Status", "Sold To Party",
    "Ship To Party", "Material", "Material Description",
    "Order Quantity (CS)", "Confirmed Quantity (CS)", "Picked Quantity (CS)",
    "Invoice Quantity (CS)", "Invoice #", "Outbound Delivery #",
]
PLANT_HEADERS = [
    "Sales Order", "Sales Order Item", "Sales Order Type", "Creation Date",
    "Order Status", "Rejection Status", "Sold To Party", "Ship To Party",
    "Material", "Material Description", "Order Quantity (CS)",
    "Invoice Quantity (CS)", "Invoice #", "Outbound Delivery #",
    "Plant", "Plant Name",
]


def _xlsx(headers, rows, sheet="SAPUI5 Export"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet
    ws.append(headers)
    for r in rows:
        ws.append(r)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _ka_row(**over):
    base = dict(zip(KEY_ACCOUNT_HEADERS, [
        "1001", 10, "SOBQU", "ZOR3", datetime(2025, 9, 5), "C", "A",
        "12000349", "20805975", "10026930", "TINKYA ELBOW BROWN 454G",
        10, 10, 10, 10, "90000001", "80000001",
    ]))
    base.update(over)
    return [base[h] for h in KEY_ACCOUNT_HEADERS]


def _plant_row(**over):
    base = dict(zip(PLANT_HEADERS, [
        "1001", 10, "ZOR3", datetime(2025, 9, 5), "C", "A",
        "12000349", "20805975", "10026930", "TINKYA ELBOW BROWN 454G",
        10, 10, "90000001", "80000001", "2910", "TOL Mississauga",
    ]))
    base.update(over)
    return [base[h] for h in PLANT_HEADERS]


# ---------------------------------------------------------------------------
# load_sales_orders
# ---------------------------------------------------------------------------
def test_load_sales_orders_key_account_variant():
    df = load_sales_orders(_xlsx(KEY_ACCOUNT_HEADERS, [_ka_row()]))
    assert list(df["Sales Order"]) == ["1001"]
    assert df.loc[0, "Plant"] == ""  # this variant has no Plant column
    assert df.loc[0, "Order Qty"] == 10
    assert df.loc[0, "Invoice Qty"] == 10


def test_load_sales_orders_plant_variant():
    df = load_sales_orders(_xlsx(PLANT_HEADERS, [_plant_row()]))
    assert df.loc[0, "Plant"] == "2910"
    assert df.loc[0, "Plant Name"] == "TOL Mississauga"


def test_load_sales_orders_missing_required_column_raises():
    bad_headers = [h for h in PLANT_HEADERS if h != "Material"]
    with pytest.raises(ForecastValidationError):
        load_sales_orders(_xlsx(bad_headers, [[v for h, v in zip(PLANT_HEADERS, _plant_row()) if h != "Material"]]))


def test_material_strips_numeric_conversion_artifact_only():
    df = load_sales_orders(_xlsx(PLANT_HEADERS, [_plant_row(Material="10026930.0")]))
    assert df.loc[0, "Material"] == "10026930"


def test_material_preserves_leading_zero():
    df = load_sales_orders(_xlsx(PLANT_HEADERS, [_plant_row(Material="00012345")]))
    assert df.loc[0, "Material"] == "00012345"


def test_blank_trailing_row_dropped():
    wb_rows = [_plant_row(), [None] * len(PLANT_HEADERS)]
    df = load_sales_orders(_xlsx(PLANT_HEADERS, wb_rows))
    assert len(df) == 1


# ---------------------------------------------------------------------------
# combine_sales_orders
# ---------------------------------------------------------------------------
def test_same_line_in_two_column_variants_not_double_counted():
    """The same Aug pull re-exported once with Key Account #, once with
    Plant — same Sales Order/Item/Invoice#/Outbound Delivery# — must
    coalesce into one row, not double Invoice Qty."""
    ka = load_sales_orders(_xlsx(KEY_ACCOUNT_HEADERS, [_ka_row()]))
    plant = load_sales_orders(_xlsx(PLANT_HEADERS, [_plant_row()]))
    combined = combine_sales_orders([ka, plant])
    assert len(combined) == 1
    row = combined.iloc[0]
    assert row["Order Qty"] == 10
    assert row["Invoice Qty"] == 10          # NOT 20
    assert row["Plant"] == "2910"            # picked up from the plant variant
    assert row["Invoice Status"] == "FULLY INVOICED"


def test_partial_invoices_within_one_file_are_summed():
    """Two distinct invoice lines against the same Sales Order + Item —
    genuinely separate shipments — must sum, not coalesce."""
    rows = [
        _plant_row(**{"Order Quantity (CS)": 10, "Invoice Quantity (CS)": 4, "Invoice #": "INV1"}),
        _plant_row(**{"Order Quantity (CS)": 10, "Invoice Quantity (CS)": 6, "Invoice #": "INV2"}),
    ]
    df = load_sales_orders(_xlsx(PLANT_HEADERS, rows))
    combined = combine_sales_orders([df])
    assert len(combined) == 1
    row = combined.iloc[0]
    assert row["Order Qty"] == 10
    assert row["Invoice Qty"] == 10
    assert row["Invoice Status"] == "FULLY INVOICED"


def test_uninvoiced_and_completion_pct():
    df = load_sales_orders(_xlsx(PLANT_HEADERS, [
        _plant_row(**{"Order Quantity (CS)": 10, "Invoice Quantity (CS)": 3}),
    ]))
    combined = combine_sales_orders([df])
    row = combined.iloc[0]
    assert row["Uninvoiced Qty"] == 7
    assert row["Invoice Completion %"] == pytest.approx(0.3)
    assert row["Invoice Status"] == "PARTIALLY INVOICED"


def test_not_invoiced_and_no_order_statuses():
    df = load_sales_orders(_xlsx(PLANT_HEADERS, [
        _plant_row(**{"Sales Order": "A", "Order Quantity (CS)": 5, "Invoice Quantity (CS)": 0}),
        _plant_row(**{"Sales Order": "B", "Order Quantity (CS)": 0, "Invoice Quantity (CS)": 0}),
    ]))
    combined = combine_sales_orders([df]).set_index("Sales Order")
    assert combined.loc["A", "Invoice Status"] == "NOT INVOICED"
    assert combined.loc["B", "Invoice Status"] == "NO ORDER"


def test_empty_input_returns_empty_frame():
    result = combine_sales_orders([])
    assert result.empty


# ---------------------------------------------------------------------------
# month window helpers
# ---------------------------------------------------------------------------
def test_default_history_and_forecast_months():
    hist, fc = default_history_and_forecast_months(date(2026, 8, 27))
    assert hist[0] == (2025, 9)
    assert hist[-1] == (2026, 8)
    assert len(hist) == 12
    assert fc[0] == (2026, 9)
    assert fc[-1] == (2027, 8)
    assert len(fc) == 12


def test_default_history_and_forecast_months_mid_fiscal_year():
    """Jan 2026 is still inside the Sept 2025-Aug 2026 fiscal year -- the
    window is anchored to that fixed fiscal year, not a 12-months-back
    rolling window from today."""
    hist, fc = default_history_and_forecast_months(date(2026, 1, 15))
    assert hist[0] == (2025, 9)
    assert hist[-1] == (2026, 8)
    assert fc[0] == (2026, 9)
    assert fc[-1] == (2027, 8)


def test_default_history_and_forecast_months_same_fiscal_year_regardless_of_month():
    """The bug this guards against: data ending in December must produce
    the SAME Sept-Aug window as data ending in August (both fall in the
    Sept 2025-Aug 2026 fiscal year) -- a rolling window would instead shift
    the whole 12-month span, silently mislabeling every column against the
    template's fixed Sept...Aug headers."""
    hist_dec, fc_dec = default_history_and_forecast_months(date(2025, 12, 31))
    hist_aug, fc_aug = default_history_and_forecast_months(date(2026, 8, 27))
    assert hist_dec == hist_aug == [
        (2025, 9), (2025, 10), (2025, 11), (2025, 12), (2026, 1), (2026, 2),
        (2026, 3), (2026, 4), (2026, 5), (2026, 6), (2026, 7), (2026, 8),
    ]
    assert fc_dec == fc_aug


def test_same_period_window_no_platform_crash():
    # %-d is not portable on Windows strftime — this must not raise.
    cur_start, cur_end, prior_start, prior_end = same_period_window(date(2026, 8, 26))
    assert cur_start == date(2026, 8, 1)
    assert cur_end == date(2026, 8, 26)
    assert prior_start == date(2025, 8, 1)
    assert prior_end == date(2025, 8, 26)


# ---------------------------------------------------------------------------
# load_forecast
# ---------------------------------------------------------------------------
def _forecast_wide_xlsx(with_history_section=False):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if with_history_section:
        ws.append([None, None, None, None, None, "History", None, None, None, None, None, None,
                   "Forecast", None, None, None])
        ws.append(["Plant", "Mat #", "Desc", "Brand name", "Buyer name",
                   "Sept", "Oct ", "Nov", "Dec", "Jan", "Feb", "Mar",
                   "Sept", "Oct ", "Nov", "Dec"])
        ws.append(["2910", "10026930", "TINKYA ELBOW BROWN 454G", "BrandX", "BuyerY",
                    111, 222, 333, 444, 555, 666, 777, 1000, 2000, 3000, 4000])
    else:
        ws.append(["Plant", "Mat #", "Desc", "Brand name", "Buyer name",
                   "Sept", "Oct ", "Nov", "Dec"])
        ws.append(["2910", "10026930", "TINKYA ELBOW BROWN 454G", "BrandX", "BuyerY",
                    1000, 2000, 3000, 4000])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def test_load_forecast_wide_plain():
    fc_months = [(2026, 9), (2026, 10), (2026, 11), (2026, 12)]
    df = load_forecast(_forecast_wide_xlsx(with_history_section=False), fc_months)
    assert len(df) == 4
    row = df[(df["Forecast Year"] == 2026) & (df["Forecast Month"] == 9)].iloc[0]
    assert row["Forecast Qty"] == 1000
    assert row["Buyer Name"] == "BuyerY"
    assert row["Brand Name"] == "BrandX"


def test_load_forecast_wide_with_history_section_uses_forecast_columns_only():
    """A re-uploaded template has BOTH a history and a forecast month block
    with identical labels ('Sept', 'Oct '...) — must read the Forecast
    section's numbers (1000/2000/3000/4000), not the History section's
    (111/222/333/444)."""
    fc_months = [(2026, 9), (2026, 10), (2026, 11), (2026, 12)]
    df = load_forecast(_forecast_wide_xlsx(with_history_section=True), fc_months)
    row = df[(df["Forecast Year"] == 2026) & (df["Forecast Month"] == 9)].iloc[0]
    assert row["Forecast Qty"] == 1000


def test_load_forecast_mrp_shape_aggregates_by_month():
    """This export's own headers don't describe their content: 'Material
    Group' actually holds the brand/supplier grouping text, 'Description'
    holds the buyer's name, and 'Material Number' holds the material's text
    description -- confirmed against a real reference workbook built from
    this same source. 'Brand Manager' is a red herring and must NOT be used."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Material", "Material Number", "Plnt", "Material Group", "Description",
               "Brand Manager", "Del/finish", "Planned qty"])
    ws.append(["10026930", "TINKYA ELBOW BROWN 454G", "2910", "ROBERTSONS", "Bita Farahani",
               "SOMEONE ELSE", datetime(2026, 9, 3), 100])
    ws.append(["10026930", "TINKYA ELBOW BROWN 454G", "2910", "ROBERTSONS", "Bita Farahani",
               "SOMEONE ELSE", datetime(2026, 9, 20), 50])
    ws.append(["10026930", "TINKYA ELBOW BROWN 454G", "2910", "ROBERTSONS", "Bita Farahani",
               "SOMEONE ELSE", datetime(2026, 10, 1), 30])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    df = load_forecast(buf, [(2026, 9), (2026, 10)])
    sept = df[(df["Forecast Year"] == 2026) & (df["Forecast Month"] == 9)].iloc[0]
    assert sept["Forecast Qty"] == 150
    assert sept["Brand Name"] == "ROBERTSONS"
    assert sept["Buyer Name"] == "Bita Farahani"
    assert sept["Material Description"] == "TINKYA ELBOW BROWN 454G"
    oct_ = df[(df["Forecast Year"] == 2026) & (df["Forecast Month"] == 10)].iloc[0]
    assert oct_["Forecast Qty"] == 30


def test_load_forecast_mrp_without_del_finish_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Material", "Plnt", "Planned qty"])
    ws.append(["10026930", "2910", 100])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ForecastValidationError):
        load_forecast(buf, [(2026, 9)])


def test_load_forecast_unrecognized_shape_raises():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Foo", "Bar"])
    ws.append([1, 2])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    with pytest.raises(ForecastValidationError):
        load_forecast(buf, [(2026, 9)])


# ---------------------------------------------------------------------------
# build_main_table
# ---------------------------------------------------------------------------
def _fact_two_years():
    rows_2025 = [
        _plant_row(**{"Sales Order": "S1", "Creation Date": datetime(2025, 9, 10),
                       "Order Quantity (CS)": 100, "Invoice Quantity (CS)": 100}),
    ]
    rows_2026 = [
        _plant_row(**{"Sales Order": "S2", "Creation Date": datetime(2026, 8, 2),
                       "Order Quantity (CS)": 84, "Invoice Quantity (CS)": 84}),
    ]
    return combine_sales_orders([
        load_sales_orders(_xlsx(PLANT_HEADERS, rows_2025)),
        load_sales_orders(_xlsx(PLANT_HEADERS, rows_2026)),
    ])


def test_build_main_table_shape_and_missing_months_left_blank():
    fact = _fact_two_years()
    hist, fc = default_history_and_forecast_months(date(2026, 8, 27))
    empty_forecast = pd.DataFrame(columns=["Plant", "Material", "Forecast Year", "Forecast Month", "Forecast Qty"])
    table = build_main_table(fact, empty_forecast, hist, fc)

    assert len(table) == 1
    row = table.iloc[0]
    assert row["Plant"] == 2910
    assert row["Mat #"] == "10026930"
    assert row["H|Sept"] == 100          # Sept 2025 invoiced
    assert row["H|Aug"] == 84            # Aug 2026 invoiced (partial, but present)
    assert pd.isna(row["H|Jan"])         # Jan 2026 never uploaded -> blank, not zero
    assert all(pd.isna(row[f"F|{l}"]) for l in ["Sept", "Oct ", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "June ", "July", "Aug"])


def test_build_main_table_includes_forecast_only_material():
    """A material with a forecast entry but zero sales history should still
    appear (a brand-new item)."""
    fact = _fact_two_years()
    hist, fc = default_history_and_forecast_months(date(2026, 8, 27))
    forecast = pd.DataFrame([{
        "Plant": "2910", "Material": "99999999", "Forecast Year": 2026,
        "Forecast Month": 9, "Forecast Qty": 500, "Buyer Name": "", "Brand Name": "",
        "Material Description": "BRAND NEW ITEM",
    }])
    table = build_main_table(fact, forecast, hist, fc)
    new_row = table[table["Mat #"] == "99999999"].iloc[0]
    assert new_row["F|Sept"] == 500
    assert new_row["Desc"] == "BRAND NEW ITEM"


# ---------------------------------------------------------------------------
# build_forecast_validation
# ---------------------------------------------------------------------------
def _validation_fixture(forecast_qty=None):
    rows = [
        # prior-year same period + same month baseline
        _plant_row(**{"Sales Order": "P1", "Creation Date": datetime(2025, 8, 10),
                       "Order Quantity (CS)": 100, "Invoice Quantity (CS)": 100}),
        _plant_row(**{"Sales Order": "P2", "Creation Date": datetime(2025, 9, 15),
                       "Order Quantity (CS)": 200, "Invoice Quantity (CS)": 200}),
        # current-year same period, showing growth
        _plant_row(**{"Sales Order": "C1", "Creation Date": datetime(2026, 8, 10),
                       "Order Quantity (CS)": 150, "Invoice Quantity (CS)": 120}),
    ]
    fact = combine_sales_orders([load_sales_orders(_xlsx(PLANT_HEADERS, rows))])
    if forecast_qty is None:
        forecast = pd.DataFrame(columns=["Plant", "Material", "Forecast Year", "Forecast Month", "Forecast Qty"])
    else:
        forecast = pd.DataFrame([{
            "Plant": "2910", "Material": "10026930", "Forecast Year": 2026,
            "Forecast Month": 9, "Forecast Qty": forecast_qty, "Buyer Name": "", "Brand Name": "",
        }])
    return fact, forecast


def test_forecast_validation_paste_forecast_when_blank():
    fact, forecast = _validation_fixture(forecast_qty=None)
    _, fc_months = default_history_and_forecast_months(date(2026, 8, 27))
    result = build_forecast_validation(fact, forecast, fc_months, pd.Timestamp(2026, 8, 27))
    sept = result[result["Forecast Month"] == "Sept 2026"].iloc[0]
    assert sept["Assessment"] == "PASTE FORECAST"


def test_forecast_validation_reasonable():
    # growth factor = 150/100 = 1.5; baseline = 200 * 1.5 = 300; forecast 300 -> exactly reasonable
    fact, forecast = _validation_fixture(forecast_qty=300)
    _, fc_months = default_history_and_forecast_months(date(2026, 8, 27))
    result = build_forecast_validation(fact, forecast, fc_months, pd.Timestamp(2026, 8, 27))
    sept = result[result["Forecast Month"] == "Sept 2026"].iloc[0]
    assert sept["Assessment"] == "REASONABLE"
    assert sept["Historical Baseline"] == pytest.approx(300)


def test_forecast_validation_high():
    fact, forecast = _validation_fixture(forecast_qty=1000)  # way above 300 baseline
    _, fc_months = default_history_and_forecast_months(date(2026, 8, 27))
    result = build_forecast_validation(fact, forecast, fc_months, pd.Timestamp(2026, 8, 27))
    sept = result[result["Forecast Month"] == "Sept 2026"].iloc[0]
    assert sept["Assessment"] == "HIGH"


def test_forecast_validation_low():
    fact, forecast = _validation_fixture(forecast_qty=50)  # way below 300 baseline
    _, fc_months = default_history_and_forecast_months(date(2026, 8, 27))
    result = build_forecast_validation(fact, forecast, fc_months, pd.Timestamp(2026, 8, 27))
    sept = result[result["Forecast Month"] == "Sept 2026"].iloc[0]
    assert sept["Assessment"] == "LOW"


def test_forecast_validation_new_no_history():
    rows = [
        _plant_row(**{"Sales Order": "C1", "Material": "77777777", "Creation Date": datetime(2026, 8, 10),
                       "Order Quantity (CS)": 10, "Invoice Quantity (CS)": 10}),
    ]
    fact = combine_sales_orders([load_sales_orders(_xlsx(PLANT_HEADERS, rows))])
    forecast = pd.DataFrame([{
        "Plant": "2910", "Material": "77777777", "Forecast Year": 2026,
        "Forecast Month": 9, "Forecast Qty": 20, "Buyer Name": "", "Brand Name": "",
    }])
    _, fc_months = default_history_and_forecast_months(date(2026, 8, 27))
    result = build_forecast_validation(fact, forecast, fc_months, pd.Timestamp(2026, 8, 27))
    sept = result[result["Forecast Month"] == "Sept 2026"].iloc[0]
    assert sept["Assessment"] == "NEW / NO HISTORY"


# ---------------------------------------------------------------------------
# build_plant_summary / build_monthly_summary / build_data_quality
# ---------------------------------------------------------------------------
def test_plant_summary_growth_signal():
    fact, _ = _validation_fixture()
    result = build_plant_summary(fact, pd.Timestamp(2026, 8, 27))
    row = result[result["Plant"] == "2910"].iloc[0]
    assert row["Demand Signal"] == "GROWING"  # 150 vs 100 = +50%


def test_monthly_summary_flags_current_month_partial():
    fact, _ = _validation_fixture()
    result = build_monthly_summary(fact, pd.Timestamp(2026, 8, 27))
    aug_2026 = result[(result["Period"] == "Aug 2026")].iloc[0]
    assert "PARTIAL" in aug_2026["Period Type"]
    aug_2025 = result[(result["Period"] == "Aug 2025")].iloc[0]
    assert aug_2025["Period Type"] == "FULL MONTH"


def test_monthly_summary_does_not_flag_partial_when_data_as_of_is_last_day_of_month():
    """Data uploaded through Dec 31 covers a full December -- it must not
    read as PARTIAL just because December is "the month data_as_of falls
    in"."""
    rows = [
        _plant_row(**{"Sales Order": "D1", "Creation Date": datetime(2025, 12, 15),
                       "Order Quantity (CS)": 10, "Invoice Quantity (CS)": 10}),
    ]
    fact = combine_sales_orders([load_sales_orders(_xlsx(PLANT_HEADERS, rows))])
    result = build_monthly_summary(fact, pd.Timestamp(2025, 12, 31))
    dec_2025 = result[result["Period"] == "Dec 2025"].iloc[0]
    assert dec_2025["Period Type"] == "FULL MONTH"


def test_data_quality_flags_invoice_over_order_and_negative_qty():
    rows = [
        _plant_row(**{"Sales Order": "X1", "Order Quantity (CS)": 5, "Invoice Quantity (CS)": 8}),
        # Invoice (-5) is not greater than Order (-1) -- keeps this row from
        # also tripping the Invoice > Order check, so each check is isolated.
        _plant_row(**{"Sales Order": "X2", "Order Quantity (CS)": -1, "Invoice Quantity (CS)": -5}),
    ]
    fact = combine_sales_orders([load_sales_orders(_xlsx(PLANT_HEADERS, rows))])
    hist, _ = default_history_and_forecast_months(date(2026, 8, 27))
    dq = build_data_quality(fact, pd.DataFrame(), hist).set_index("Check")["Result"]
    assert dq["Rows with Invoice Qty > Order Qty"] == "1"
    assert dq["Rows with negative Order Qty"] == "1"


# ---------------------------------------------------------------------------
# generate_excel
# ---------------------------------------------------------------------------
def test_generate_excel_main_sheet_matches_template_layout():
    fact = _fact_two_years()
    hist, fc = default_history_and_forecast_months(date(2026, 8, 27))
    empty_forecast = pd.DataFrame(columns=["Plant", "Material", "Forecast Year", "Forecast Month", "Forecast Qty"])
    table = build_main_table(fact, empty_forecast, hist, fc)
    validation = build_forecast_validation(fact, empty_forecast, fc, pd.Timestamp(2026, 8, 27))
    plant_summary = build_plant_summary(fact, pd.Timestamp(2026, 8, 27))
    monthly = build_monthly_summary(fact, pd.Timestamp(2026, 8, 27))
    dq = build_data_quality(fact, empty_forecast, hist)

    data = generate_excel(table, validation, plant_summary, pd.DataFrame(), monthly, dq,
                           hist, fc, pd.Timestamp(2026, 8, 27))
    wb = openpyxl.load_workbook(io.BytesIO(data))
    assert wb.sheetnames[0] == "Sheet1"
    ws = wb["Sheet1"]
    header = [ws.cell(2, c).value for c in range(1, 39)]
    assert header[:5] == ["Plant", "Mat #", "Desc", "Brand name", "Buyer name"]
    assert header[5:17] == ["Sept", "Oct ", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "June ", "July", "Aug (partial)"]
    assert header[17:29] == ["Sept", "Oct ", "Nov", "Dec", "Jan", "Feb", "Mar", "Apr", "May", "June ", "July", "Aug"]
    # build_main_table always carries Open Orders/Open PO/Stock on Hand
    # columns (blank when no such file was uploaded), so Sheet1 always
    # matches the extended "output_needed_with_stock_on_hand" layout.
    assert header[29:33] == ["Sept", "Oct", "Nov", "Dec"]
    assert header[33:37] == ["Sept", "Oct", "Nov", "Dec"]
    # Stock on Hand has no monthly breakdown, so its label lives in the
    # vertically-merged AL1:AL2 (row 1), not row 2 -- matching the source
    # template exactly.
    assert ws.cell(1, 38).value == "Stock on Hand"
    merged = {str(r) for r in ws.merged_cells.ranges}
    assert merged == {"F1:Q1", "R1:AC1", "AD1:AG1", "AH1:AK1", "AL1:AL2"}
    assert set(wb.sheetnames) == {"Sheet1", "Forecast Validation", "Plant Summary", "Item Detail", "Monthly Summary", "Data Quality"}


def test_generate_excel_does_not_mark_partial_when_data_as_of_is_last_day_of_month():
    fact = _fact_two_years()
    hist, fc = default_history_and_forecast_months(date(2025, 12, 31))
    empty_forecast = pd.DataFrame(columns=["Plant", "Material", "Forecast Year", "Forecast Month", "Forecast Qty"])
    table = build_main_table(fact, empty_forecast, hist, fc)
    dq = build_data_quality(fact, empty_forecast, hist)

    data = generate_excel(table, pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), dq,
                           hist, fc, pd.Timestamp(2025, 12, 31))
    ws = openpyxl.load_workbook(io.BytesIO(data))["Sheet1"]
    header = [ws.cell(2, c).value for c in range(1, 18)]
    assert "Dec" in header and "(partial)" not in " ".join(str(h) for h in header)


# ---------------------------------------------------------------------------
# Open Orders
# ---------------------------------------------------------------------------
OPEN_ORDERS_HEADERS = ["Sales Order", "Key Account Name", "Material", "Material Description",
                        "Sales Order Type", "Creation Date", "Requested Delivery Date",
                        "Order Quantity (CS)", "BDM Description", "Purchasing Group Name", "CDM Name"]


def _open_orders_xlsx(rows):
    return _xlsx(OPEN_ORDERS_HEADERS, rows, sheet="SAPUI5 Export")


def test_load_open_orders_parses_and_buckets_by_requested_month():
    rows = [
        ["1001", "Ontario Independents", "10026930", "TINKYA ELBOW BROWN 454G", "ZOR3",
         datetime(2026, 2, 23), datetime(2026, 10, 20), 5, "MICHELLE JEWELL NT", "Anabel Lopez", "Brandon Fuselli"],
    ]
    df = load_open_orders(_open_orders_xlsx(rows))
    assert len(df) == 1
    row = df.iloc[0]
    assert row["Sales Order"] == "1001"
    assert row["Material"] == "10026930"
    assert row["Open Order Qty"] == 5
    assert row["Buyer Name"] == "MICHELLE JEWELL NT"
    assert (row["Year"], row["Month"]) == (2026, 10)


def test_load_open_orders_missing_required_column_raises():
    bad_headers = [h for h in OPEN_ORDERS_HEADERS if h != "Requested Delivery Date"]
    with pytest.raises(ForecastValidationError):
        load_open_orders(_xlsx(bad_headers, [["1001", "A", "10026930", "D", "ZOR3", datetime(2026, 2, 23), 5, "B", "C", "D"]]))


def test_resolve_open_order_plants_joins_from_fact_and_leaves_unresolved_blank():
    fact_rows = [
        _plant_row(**{"Sales Order": "1001", "Material": "10026930", "Plant": "2910"}),
    ]
    fact = combine_sales_orders([load_sales_orders(_xlsx(PLANT_HEADERS, fact_rows))])

    open_orders = load_open_orders(_open_orders_xlsx([
        ["1001", "A", "10026930", "D", "ZOR3", datetime(2026, 2, 23), datetime(2026, 10, 20), 5, "B", "C", "D"],
        ["9999", "A", "99999999", "D", "ZOR3", datetime(2026, 2, 23), datetime(2026, 10, 20), 3, "B", "C", "D"],
    ]))
    resolved = resolve_open_order_plants(open_orders, fact)
    resolved = resolved.set_index("Sales Order")
    assert resolved.loc["1001", "Plant"] == "2910"
    assert resolved.loc["9999", "Plant"] == ""


# ---------------------------------------------------------------------------
# Open PO / Stock on Hand
# ---------------------------------------------------------------------------
OPEN_PO_HEADERS = ["PO Number", "Purchasing Group Description", "Document Type", "Vendor",
                    "Vendor Name", "Supplier Name", "Freight Supplier Name", "Inco Terms",
                    "Delivery Date", "Delivery Date Derived Field", "Plant", "Material",
                    "Material Description", "Stock on Hand", "PO Quantity", "Appt. Plant",
                    "BDM Description"]


def _open_po_xlsx(rows):
    return _xlsx(OPEN_PO_HEADERS, rows, sheet="SAPUI5 Export")


def test_load_open_po_parses_and_buckets_by_delivery_month():
    rows = [
        ["1000006577", "Jennifer Yang", "NB", "70000332", "WALKERS LTD", "WALKERS LTD",
         "Universal Logistics", "FOB", datetime(2026, 9, 1), "AB", "2930", "10078019",
         "JW ADVENT CALENDAR", 0, 816.0, None, "SAMANTHA RODRIGUES GB"],
    ]
    df = load_open_po(_open_po_xlsx(rows))
    assert len(df) == 1
    row = df.iloc[0]
    assert row["Plant"] == "2930"
    assert row["Material"] == "10078019"
    assert row["Open PO Qty"] == 816
    assert row["Stock on Hand"] == 0
    assert (row["Year"], row["Month"]) == (2026, 9)


def test_load_open_po_missing_required_column_raises():
    bad_headers = [h for h in OPEN_PO_HEADERS if h != "Delivery Date"]
    with pytest.raises(ForecastValidationError):
        load_open_po(_xlsx(bad_headers, [["1", "A", "NB", "V", "VN", "SN", "FS", "FOB",
                                           "2930", "10078019", "D", 0, 816.0, None, "B"]]))


# ---------------------------------------------------------------------------
# build_main_table with Open Orders / Open PO / Stock on Hand
# ---------------------------------------------------------------------------
def test_build_main_table_includes_open_orders_po_and_stock():
    """Stock on Hand comes only from the Open PO export's own column --
    there is no separate dedicated stock/inventory source."""
    fact = _fact_two_years()
    hist, fc = default_history_and_forecast_months(date(2026, 8, 27))
    empty_forecast = pd.DataFrame(columns=["Plant", "Material", "Forecast Year", "Forecast Month", "Forecast Qty"])

    open_orders = pd.DataFrame([{
        "Plant": "2910", "Material": "10026930", "Year": 2026, "Month": 9, "Open Order Qty": 42,
    }])
    open_po = pd.DataFrame([{
        "Plant": "2910", "Material": "10026930", "Year": 2026, "Month": 9,
        "Open PO Qty": 100, "Stock on Hand": 7,
    }])

    table = build_main_table(fact, empty_forecast, hist, fc, open_orders=open_orders, open_po=open_po)
    row = table[table["Mat #"] == "10026930"].iloc[0]
    assert row["OO|Sept"] == 42
    assert row["PO|Sept"] == 100
    assert row["Stock on Hand"] == 7


def test_build_main_table_includes_material_with_only_an_open_po():
    """A material that has never sold but has an incoming PO should still
    appear -- it's real, actionable information even with zero history."""
    fact = _fact_two_years()
    hist, fc = default_history_and_forecast_months(date(2026, 8, 27))
    empty_forecast = pd.DataFrame(columns=["Plant", "Material", "Forecast Year", "Forecast Month", "Forecast Qty"])
    open_po = pd.DataFrame([{
        "Plant": "2910", "Material": "55555555", "Year": 2026, "Month": 9,
        "Open PO Qty": 60, "Stock on Hand": None,
    }])
    table = build_main_table(fact, empty_forecast, hist, fc, open_po=open_po)
    assert (table["Mat #"] == "55555555").any()
