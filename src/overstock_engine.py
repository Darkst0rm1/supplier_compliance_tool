"""Processing engine for the weekly Overstock Report.

Two SAPUI5 exports go in:

* **Materials** — inventory snapshot (one row per Material / Plant / Storage
  Location / Batch) with the three stock buckets and the Shelf Life Expiration
  Date (SLED).
* **Last Sell / BDM Material Master** — one row per ``Product Number`` (the
  master may repeat a product across vendors) carrying the Brand Manager and
  the product's ``Last Sell Day`` offset.

The report flags excess stock that is approaching its last sellable date, split
into one sheet per warehouse region (Mississauga / Calgary / Surrey). The exact
finished workbook supplied by the business is the golden specification; the
rules below were reverse-engineered against it (259 / 260 rows reproduced — the
single difference is a hand-edited row in the golden whose Plant/Storage/stock
combination does not exist in the source Materials file).

Selection rules (all must hold):

1. Total stock = Unrestricted + Quality Inspection + Blocked > 0.
2. Plant belongs to the sheet's region (see ``REGION_PLANTS``).
3. Storage Location is the main warehouse ("1000") OR the row is Customer
   Consignment (which carries a blank Storage Location).
4. Material number does NOT start with "40" (those are display / shipper /
   label / sticker / sample packaging materials, never sellable stock).
5. The Material matches a master ``Product Number`` (so it has a Last Sell Day).
6. Shelf Life Expiration Date present and on/after the SLED floor
   (report date + ``SLED_FLOOR_OFFSET_DAYS``).
7. Last sell by date (= SLED - Last Sell Day) on/before the cutoff
   (report date + ``LAST_SELL_CUTOFF_OFFSET_DAYS``).
8. NOT the RANA retail brand handled by Sandra (Brand Manager "Sandra
   Gaganiaras" AND description starting "RANA"); the foodservice "RANA FS"
   line handled by another BDM stays in.
9. NOT the Sweet Street ("SSD …") brand.

Rows are sorted by Shelf Life Expiration Date ascending within each sheet.

Storage bins (optional third input)
-----------------------------------
An **EWM stock export** may be supplied per plant (2910 / 2920 / 2930). It
carries one row per Product / Batch / Storage Bin, and lets the report name the
bin each overstocked batch is sitting in. This reproduces the manual workflow
the business ran in Excel: build a ``Materialbatch`` key (Material & Batch),
then ``VLOOKUP`` it against the EWM sheet to pull ``Storage Bin``.

Faithful details of that VLOOKUP, reverse-engineered from the finished
"including BIN" workbook (271 of 276 rows reproduced — the 5 exceptions are
batches absent from the EWM extracts supplied alongside it, i.e. a different
snapshot, not a rule difference):

* The key is ``Material`` & ``Batch`` concatenated with no separator.
* EWM repeats a Product/Batch across many bins. ``VLOOKUP`` returns the
  **first** match in file order, so the lookup keeps first and file order is
  load-bearing.
* Only plants 2910 / 2920 / 2930 have an EWM extract. Region plants **2925 and
  2935 never get a bin** — there is no export for them.
* The bins are a lookup only. They never add, drop, or reorder a row.

An empty Bin and an ``#N/A`` Bin mean different things, and the distinction is
the reader's only way to tell them apart:

* **blank** — the plant's EWM export was searched and this batch wasn't in it.
* **#N/A** — there was nothing to search. Always the case for 2925 / 2935, and
  also for 2910 / 2920 / 2930 when that plant's export wasn't uploaded.

The finished workbook carried the ``Materialbatch`` key twice (columns G and K)
— residue of the manual VLOOKUP. Only the first is kept here; the second was
redundant once the lookup moved into code.

Bins are optional: with no EWM file the report is built exactly as before,
with the Bin column present but empty.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class OverstockError(Exception):
    """Raised when an uploaded file isn't a usable Overstock source export."""


# ---------------------------------------------------------------------------
# Source columns (exact export headers)
# ---------------------------------------------------------------------------
MAT_MATERIAL      = "Material"
MAT_DESCRIPTION   = "Material Description"
MAT_PLANT         = "Plant"
MAT_PLANT_NAME    = "Plant Name"
MAT_STORAGE_LOC   = "Storage Location"
MAT_STORAGE_DESC  = "Description of Storage Location"
MAT_BATCH         = "Batch"
MAT_SLED          = "Shelf Life Expiration Date"
MAT_SPECIAL_STOCK = "Special Stock Type Description"
MAT_UNRESTRICTED  = "Unrestricted Stock"
MAT_QUALITY       = "Stock in Quality Inspection"
MAT_BLOCKED       = "Blocked Stock"

STOCK_COLS = [MAT_UNRESTRICTED, MAT_QUALITY, MAT_BLOCKED]
MAT_TEXT_COLS = [MAT_MATERIAL, MAT_PLANT, MAT_STORAGE_LOC, MAT_BATCH]

MASTER_PRODUCT   = "Product Number"
MASTER_BDM_NAME  = "Brand Manager Name"
MASTER_LAST_SELL = "Last Sell Day"

# EWM stock export (one row per Product / Batch / Storage Bin).
EWM_PRODUCT = "Product"
EWM_BATCH   = "Batch"
EWM_BIN     = "Storage Bin"
EWM_OWNER   = "Party Entitled to Dispose"   # "BP2910" — lets a file name its plant

# ---------------------------------------------------------------------------
# Output columns
# ---------------------------------------------------------------------------
OUT_BDM           = "BDM"
OUT_LAST_SELL_DAY = "Last sell by day"
OUT_LAST_SELL_DT  = "Last sell by date"
OUT_MATERIALBATCH = "Materialbatch"
OUT_BIN           = "Bin"

# Base layout shared by every sheet (in finished-workbook order).
BASE_COLUMNS = [
    MAT_MATERIAL,
    MAT_DESCRIPTION,
    MAT_PLANT,
    MAT_PLANT_NAME,
    OUT_BDM,
    MAT_STORAGE_LOC,
    OUT_MATERIALBATCH,
    OUT_BIN,
    MAT_STORAGE_DESC,
    MAT_BATCH,
    MAT_SLED,
    OUT_LAST_SELL_DAY,
    OUT_LAST_SELL_DT,
    MAT_SPECIAL_STOCK,
    MAT_UNRESTRICTED,
    MAT_QUALITY,
    MAT_BLOCKED,
]

# Per-sheet trailing blank column (kept empty for the team to annotate).
SHEET_TRAILING_COL = {
    "Mississauga": "Notes",
    "Calgary": None,
    "Surrey": "COMMENTS",
}

# Region -> plant codes. Order of the sheets is preserved on output.
REGION_PLANTS: dict[str, list[str]] = {
    "Mississauga": ["2910"],
    "Calgary": ["2920", "2925"],
    "Surrey": ["2930", "2935"],
}

# ---------------------------------------------------------------------------
# Business-rule constants (auditable; change here, not in the UI)
# ---------------------------------------------------------------------------
# A Material whose number starts with any of these prefixes is packaging /
# display / promo material, never sellable stock — excluded entirely.
EXCLUDED_MATERIAL_PREFIXES = ("40",)

# Storage Location of the main warehouse.
MAIN_WAREHOUSE_STORAGE_LOC = "1000"

# Brand exclusions (matched case-insensitively after stripping).
SWEET_STREET_DESC_PREFIX = "SSD"          # Sweet Street desserts brand
RANA_DESC_PREFIX = "RANA"                 # Giovanni Rana
RANA_EXCLUDED_BDM = "SANDRA GAGANIARAS GB"  # her RANA retail line is dropped

# Date window, relative to the report run date.
SLED_FLOOR_OFFSET_DAYS = 6        # include SLED on/after report_date + 6
LAST_SELL_CUTOFF_OFFSET_DAYS = 7  # include last-sell-by on/before report_date + 7

# Plants with an EWM stock export. Region plants absent here (2925, 2935) have
# no bin data at all.
EWM_PLANTS: list[str] = ["2910", "2920", "2930"]

# Shown when there is no EWM export to look the row up in, so a reader can tell
# "checked, this batch has no bin" (blank) from "no bin data exists for this
# plant" (#N/A). openpyxl binds this string to a real Excel error value, which
# is what the golden's VLOOKUP produced.
NO_LOOKUP_MARKER = "#N/A"


# ---------------------------------------------------------------------------
# Importers
# ---------------------------------------------------------------------------
def _read_excel_str(file_obj: Any) -> pd.DataFrame:
    file_obj.seek(0)
    df = pd.read_excel(file_obj, dtype=str, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _require(df: pd.DataFrame, cols: list[str], what: str) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise OverstockError(
            f"This doesn't look like the {what} export — missing column(s): "
            + ", ".join(missing)
        )


def load_materials(file_obj: Any) -> pd.DataFrame:
    """Read the Materials inventory export. Stock buckets become numeric, the
    SLED becomes a datetime, and id columns are kept as clean text strings."""
    df = _read_excel_str(file_obj)
    _require(df, [MAT_MATERIAL, MAT_PLANT, MAT_SLED] + STOCK_COLS, "Materials")

    for c in STOCK_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df[MAT_SLED] = pd.to_datetime(df[MAT_SLED], errors="coerce")

    for c in MAT_TEXT_COLS:
        if c in df.columns:
            # Strip only ASCII whitespace so the finished workbook's preserved
            # batch padding (trailing non-breaking spaces) survives verbatim.
            df[c] = (
                df[c].astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.strip(" \t\r\n")
                .replace({"nan": "", "None": "", "NaT": ""})
            )
    return df


def load_master(file_obj: Any) -> pd.DataFrame:
    """Read the Last Sell / BDM master and reduce it to one row per Product
    Number (the master repeats products across vendors). Last Sell Day becomes
    numeric."""
    df = _read_excel_str(file_obj)
    _require(df, [MASTER_PRODUCT, MASTER_BDM_NAME, MASTER_LAST_SELL],
             "Last Sell / BDM Material Master")

    df[MASTER_PRODUCT] = (
        df[MASTER_PRODUCT].astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )
    df[MASTER_LAST_SELL] = pd.to_numeric(df[MASTER_LAST_SELL], errors="coerce")
    df = df.drop_duplicates(subset=MASTER_PRODUCT, keep="first")
    return df[[MASTER_PRODUCT, MASTER_BDM_NAME, MASTER_LAST_SELL]].reset_index(drop=True)


def _clean_text(s: pd.Series) -> pd.Series:
    """Blank-safe text: NA becomes "", float artifacts are dropped, and padding
    is trimmed.

    ``.str.strip()`` rather than a ``\\s+`` regex on purpose. Batches carry
    **non-breaking** space padding — ``load_materials`` preserves it
    deliberately — and pandas runs ``.str.replace(regex=True)`` on PyArrow's
    RE2 engine, whose ``\\s`` is ASCII-only and leaves U+00A0 in place.
    ``str.strip`` is Unicode-aware, so it takes the NBSP out. Likewise
    ``fillna`` rather than ``astype(str)``: the latter renders NA as the string
    "nan" on object columns but leaves it as NA on the string dtype, so the
    old "nan" -> "" mapping silently missed it.
    """
    return (
        s.fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .replace({"nan": "", "None": "", "NaT": ""})
    )


def _bin_key(material: pd.Series, batch: pd.Series) -> pd.Series:
    """Material & Batch, trimmed on both sides — the join key. Batches arrive
    padded inconsistently between the two exports, so the key strips while the
    displayed Materialbatch keeps whatever padding it came with."""
    return _clean_text(material) + _clean_text(batch)


def load_ewm_bins(file_obj: Any, expected_plant: str | None = None) -> pd.Series:
    """Read one plant's EWM stock export and reduce it to a
    ``Material+Batch -> Storage Bin`` lookup.

    EWM lists a Product/Batch once per bin it occupies; the business's VLOOKUP
    takes the first hit, so this keeps the first row in file order. Pass
    ``expected_plant`` to reject a file uploaded into the wrong plant's box —
    each export names its own plant in ``Party Entitled to Dispose`` ("BP2910").
    """
    df = _read_excel_str(file_obj)
    _require(df, [EWM_PRODUCT, EWM_BATCH, EWM_BIN],
             f"EWM stock ({expected_plant})" if expected_plant else "EWM stock")

    if expected_plant and EWM_OWNER in df.columns:
        owners = (
            df[EWM_OWNER].dropna().astype(str).str.strip()
            .str.upper().str.removeprefix("BP")
        )
        found = sorted({o for o in owners if o})
        if found and expected_plant not in found:
            raise OverstockError(
                f"This file is the plant {', '.join(found)} EWM export, but it "
                f"was uploaded in the {expected_plant} box. Swap the files and "
                "try again."
            )

    keys = _bin_key(df[EWM_PRODUCT], df[EWM_BATCH])
    bins = _clean_text(df[EWM_BIN])

    lut = pd.DataFrame({"key": keys, "bin": bins})
    lut = lut[(lut["key"] != "") & (lut["bin"] != "")]
    lut = lut.drop_duplicates(subset="key", keep="first")   # VLOOKUP first match
    return lut.set_index("key")["bin"]


# ---------------------------------------------------------------------------
# Core build
# ---------------------------------------------------------------------------
def _coerce_report_date(report_date: date | datetime | None) -> date:
    if report_date is None:
        return date.today()
    if isinstance(report_date, datetime):
        return report_date.date()
    return report_date


def default_window(report_date: date | datetime | None = None) -> tuple[date, date]:
    """The default (SLED floor, last-sell cutoff) for a given run date. Used to
    pre-fill the page; the user can override either date freely."""
    rpt = _coerce_report_date(report_date)
    return (rpt + timedelta(days=SLED_FLOOR_OFFSET_DAYS),
            rpt + timedelta(days=LAST_SELL_CUTOFF_OFFSET_DAYS))


def build_overstock(
    materials: pd.DataFrame,
    master: pd.DataFrame,
    sled_floor: date | datetime | None = None,
    last_sell_cutoff: date | datetime | None = None,
    report_date: date | datetime | None = None,
    ewm_bins: dict[str, pd.Series] | None = None,
) -> dict[str, pd.DataFrame]:
    """Apply every selection rule and return ``{sheet_name: DataFrame}`` with
    one DataFrame per region in finished-workbook column order.

    The date window is open: pass ``sled_floor`` (include SLED on/after) and
    ``last_sell_cutoff`` (include last-sell-by on/before) explicitly. Either one
    left as ``None`` falls back to the default offset from ``report_date``
    (today if also ``None``).

    ``ewm_bins`` maps a plant code to that plant's ``load_ewm_bins`` lookup. It
    only fills the Bin column — row selection is identical with or without it.
    Plants with no entry get ``#N/A``; plants with one get the bin, or blank
    where the batch isn't in their export."""
    default_floor, default_cutoff = default_window(report_date)
    sled_floor = pd.Timestamp(sled_floor if sled_floor is not None else default_floor)
    last_sell_cutoff = pd.Timestamp(
        last_sell_cutoff if last_sell_cutoff is not None else default_cutoff
    )

    m = materials.copy()

    # Master lookup: Brand Manager Name -> BDM, Last Sell Day -> Last sell by day
    lut = master.rename(columns={
        MASTER_BDM_NAME: OUT_BDM,
        MASTER_LAST_SELL: OUT_LAST_SELL_DAY,
    })
    m = m.merge(lut, how="left", left_on=MAT_MATERIAL, right_on=MASTER_PRODUCT)

    # Derived dates / totals
    m[OUT_LAST_SELL_DT] = m[MAT_SLED] - pd.to_timedelta(m[OUT_LAST_SELL_DAY], unit="D")
    total_stock = m[STOCK_COLS].sum(axis=1)

    desc = m[MAT_DESCRIPTION].astype(str).str.strip().str.upper()
    bdm = m[OUT_BDM].astype(str).str.strip().str.upper()
    sloc = m[MAT_STORAGE_LOC].astype(str).str.strip()
    special = m[MAT_SPECIAL_STOCK].astype(str).str.upper()

    is_packaging = m[MAT_MATERIAL].astype(str).str.startswith(EXCLUDED_MATERIAL_PREFIXES)
    is_sweet_street = desc.str.startswith(SWEET_STREET_DESC_PREFIX)
    is_rana_sandra = (bdm == RANA_EXCLUDED_BDM) & desc.str.startswith(RANA_DESC_PREFIX)
    storage_ok = (sloc == MAIN_WAREHOUSE_STORAGE_LOC) | special.str.contains("CONSIGNMENT")

    keep = (
        (total_stock > 0)
        & m[MAT_SLED].notna()
        & m[OUT_LAST_SELL_DAY].notna()
        & storage_ok
        & ~is_packaging
        & ~is_sweet_street
        & ~is_rana_sandra
        & (m[MAT_SLED] >= sled_floor)
        & (m[OUT_LAST_SELL_DT] <= last_sell_cutoff)
    )
    selected = m[keep].copy()

    ewm_bins = ewm_bins or {}

    sheets: dict[str, pd.DataFrame] = {}
    for sheet_name, plants in REGION_PLANTS.items():
        sub = selected[selected[MAT_PLANT].isin(plants)].copy()
        sub = sub.sort_values(
            [MAT_SLED, MAT_MATERIAL, MAT_BATCH], kind="mergesort"
        ).reset_index(drop=True)

        # Displayed key keeps the batch's own padding; the join key strips it.
        # fillna before astype — on the string dtype astype(str) leaves NA as
        # NA, which would turn the whole key blank for a batch-less row.
        sub[OUT_MATERIALBATCH] = (
            sub[MAT_MATERIAL].fillna("").astype(str)
            + sub[MAT_BATCH].fillna("").astype(str)
        )

        # A plant with no lookup gets #N/A — nothing was checked. A plant with a
        # lookup gets the bin, or blank when the batch simply isn't in it.
        join_key = _bin_key(sub[MAT_MATERIAL], sub[MAT_BATCH])
        sub[OUT_BIN] = pd.Series(
            [NO_LOOKUP_MARKER] * len(sub), index=sub.index, dtype=object
        )
        for plant in plants:
            lut = ewm_bins.get(plant)
            if lut is None or lut.empty:
                continue
            at_plant = sub[MAT_PLANT] == plant
            sub.loc[at_plant, OUT_BIN] = join_key[at_plant].map(lut)

        out = pd.DataFrame({col: sub.get(col) for col in BASE_COLUMNS})
        # Special Stock / BDM blanks render as empty cells, not the string "nan".
        out[MAT_SPECIAL_STOCK] = out[MAT_SPECIAL_STOCK].where(
            out[MAT_SPECIAL_STOCK].notna(), None
        )
        trailing = SHEET_TRAILING_COL.get(sheet_name)
        if trailing:
            out[trailing] = None
        sheets[sheet_name] = out

    return sheets


# ---------------------------------------------------------------------------
# Excel export (matches the finished workbook's formatting)
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="FFF7F7F7")
_HEADER_FONT = Font(name="Arial", size=11, bold=True)
_BODY_FONT = Font(name="Arial", size=11)
_DATE_FMT = "mm/dd/yyyy"
_STOCK_FMT = r'0\ "CS"'

# Column widths copied from the finished workbook (Mississauga profile), keyed
# by output header so they apply regardless of a sheet's trailing column.
_COL_WIDTHS = {
    MAT_MATERIAL: 13.0,
    MAT_DESCRIPTION: 40.82,
    MAT_PLANT: 7.0,
    MAT_PLANT_NAME: 16.0,
    OUT_BDM: 24.54,
    MAT_STORAGE_LOC: 20.0,
    OUT_MATERIALBATCH: 13.0,
    OUT_BIN: 20.0,
    MAT_STORAGE_DESC: 36.27,
    MAT_BATCH: 12.73,
    MAT_SLED: 27.45,
    OUT_LAST_SELL_DAY: 18.0,
    OUT_LAST_SELL_DT: 19.09,
    MAT_SPECIAL_STOCK: 34.09,
    MAT_UNRESTRICTED: 20.0,
    MAT_QUALITY: 14.27,
    MAT_BLOCKED: 18.0,
    "Notes": 25.27,
    "COMMENTS": 25.27,
}

# Written as text so long numeric-looking keys aren't mangled into floats.
# Bin is left General, matching the golden.
_TEXT_OUT_COLS = set(MAT_TEXT_COLS) | {OUT_MATERIALBATCH}


def _write_sheet(ws, df: pd.DataFrame) -> None:
    headers = list(df.columns)

    for c_idx, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    sled_idx = headers.index(MAT_SLED) + 1 if MAT_SLED in headers else None
    lsd_idx = headers.index(OUT_LAST_SELL_DT) + 1 if OUT_LAST_SELL_DT in headers else None
    stock_idx = {headers.index(c) + 1 for c in STOCK_COLS if c in headers}
    text_idx = {headers.index(c) + 1 for c in _TEXT_OUT_COLS if c in headers}

    for r_off, (_, row) in enumerate(df.iterrows()):
        r_idx = r_off + 2
        for c_idx, col in enumerate(headers, 1):
            val = row[col]
            if pd.isna(val):
                val = None
            if c_idx in text_idx and val is not None:
                val = str(val)
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = _BODY_FONT
            if c_idx in (sled_idx, lsd_idx) and val is not None:
                cell.number_format = _DATE_FMT
            elif c_idx in stock_idx:
                cell.number_format = _STOCK_FMT
            elif c_idx in text_idx:
                cell.number_format = "@"

    # Filters across the data columns (through Blocked Stock — matches golden).
    last_filter_col = headers.index(MAT_BLOCKED) + 1 if MAT_BLOCKED in headers else len(headers)
    last_row = max(len(df) + 1, 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(last_filter_col)}{last_row}"

    for c_idx, col in enumerate(headers, 1):
        width = _COL_WIDTHS.get(col)
        if width:
            ws.column_dimensions[get_column_letter(c_idx)].width = width


def generate_excel(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Render the region DataFrames into a formatted workbook (bytes)."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in REGION_PLANTS:                 # preserve region order
        ws = wb.create_sheet(sheet_name)
        _write_sheet(ws, sheets[sheet_name])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
