"""Processing engine for SAP/BW Delivery Fill Rate exports."""
from __future__ import annotations

import io
import re
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Column keyword mapping
# Order matters — total_short_amount must come before short_amount so
# Total_Short_Amt doesn't accidentally map to short_amount.
# ---------------------------------------------------------------------------
COLUMN_KEYWORDS: dict[str, list[str]] = {
    "outbound_delivery": ["outbound delivery", "obd", "outbnd del", "ob delivery"],
    "plant":             ["plant", "werk", "shipping point"],
    "product":           ["product", "material", "matnr", "mat.", "sku", "article"],
    "sales_order":       ["sales order", "sales document", "s/o", "order no", "so no"],
    "requested_delivery_date": [
        "requested delivery date", "req. del", "rdd", "delivery date",
        "confirmed delivery", "req del",
    ],
    "order_quantity": [
        "order quantity", "order qty", "ordered qty", "order quan", "open qty", "target qty",
    ],
    "delivered_quantity": [
        "delivered quantity", "delivered qty", "gi qty", "goods issue qty",
        "del. qty", "shipped qty", "actual qty", "deliv qty",
        "dvl qty",          # Dvl_Qty (BU) from SAP/BW export
    ],
    "short_quantity": [
        "short quantity", "short qty", "shortage qty",
        "short qty",        # Dvl_ShortQty → normalises to "dvl short qty"
    ],
    "wh_fill_rate": [
        "wh fill",          # WH_FillRate → normalises to "wh fill rate"
        "warehouse fill", "fill rate wh", "wh fr",
    ],
    "customer_fill_rate": [
        "customer fill",    # Customer_FillRate → normalises to "customer fill rate"
        "cust fill", "cfr", "customer fr", "cust. fill",
    ],
    "net_value": ["net value", "value (lc)", "net val", "value lc", "total value"],
    # total_short_amount MUST be listed before short_amount so "Total_Short_Amt"
    # maps here (via "total short") rather than to short_amount (via "short amt").
    "total_short_amount": ["total short"],
    "short_amount": [
        "short amount", "shortage amount", "short amt", "short value", "shortage value",
        "dvl short",    # Dvl_Short_Amt → normalises to "dvl short amt"
    ],
}


# ---------------------------------------------------------------------------
# Column name normalisation (used for both scoring and matching)
# Handles: underscores → spaces, camelCase splitting, parentheses removal
# Examples:
#   WH_FillRate       → wh fill rate
#   Customer_FillRate → customer fill rate
#   Dvl_ShortQty      → dvl short qty
#   Dvl_Qty (BU)      → dvl qty bu
#   Total_Short_Amt   → total short amt
# ---------------------------------------------------------------------------

def _normalize_col(col_name: str) -> str:
    s = str(col_name).strip()
    s = re.sub(r'([a-z])([A-Z])', r'\1 \2', s)   # split camelCase before lowercasing
    s = s.lower()
    s = s.replace("_", " ").replace("(", " ").replace(")", " ").replace(".", " ")
    s = re.sub(r'\s+', ' ', s).strip()
    return s


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _score_row(row: pd.Series) -> int:
    """Count how many COLUMN_KEYWORDS appear in this row (after normalisation)."""
    score = 0
    for val in row:
        if not isinstance(val, str):
            continue
        v = _normalize_col(val)
        for aliases in COLUMN_KEYWORDS.values():
            if any(a in v for a in aliases):
                score += 1
                break
    return score


def detect_header_row(file_obj: Any) -> int:
    """Return the 0-based row index of the real column header."""
    raw = pd.read_excel(file_obj, header=None, nrows=40, engine="openpyxl")
    scores = [(i, _score_row(row)) for i, row in raw.iterrows()]
    best_idx, best_score = max(scores, key=lambda x: x[1])
    return int(best_idx) if best_score > 0 else 0


# ---------------------------------------------------------------------------
# Column normalisation
# ---------------------------------------------------------------------------

def _match_column(col_name: str) -> str | None:
    """Return the canonical key for a raw column name, or None."""
    v = _normalize_col(col_name)
    for canonical, aliases in COLUMN_KEYWORDS.items():
        if any(a in v for a in aliases):
            return canonical
    return None


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Rename columns to canonical snake_case names; keep unmatched as-is."""
    rename_map: dict[str, str] = {}
    seen: set[str] = set()

    # Product Group arrives as a code+name pair, both literally headed
    # "Product Group" (pandas dedups the second to "Product Group.1").
    # BEx exports key first then text, so first = code, second = name.
    # Handle it explicitly because "product group" contains the substring
    # "product" and would otherwise collide with the `product` keyword.
    pg_cols = [c for c in df.columns if _normalize_col(c).startswith("product group")]
    if pg_cols:
        rename_map[pg_cols[0]] = "product_group_code"
        seen.add("product_group_code")
        if len(pg_cols) > 1:
            rename_map[pg_cols[1]] = "product_group_name"
            seen.add("product_group_name")

    for col in df.columns:
        if col in rename_map:
            continue
        key = _match_column(str(col))
        if key and key not in seen:
            rename_map[col] = key
            seen.add(key)
    return df.rename(columns=rename_map)


# ---------------------------------------------------------------------------
# Type conversion helpers
# ---------------------------------------------------------------------------

def _to_numeric(series: pd.Series) -> pd.Series:
    """Strip currency/percent symbols and coerce to float."""
    if series.dtype == object:
        series = (
            series.astype(str)
            .str.replace(r"[$,%\s]", "", regex=True)
            .str.replace(",", "", regex=False)
        )
    return pd.to_numeric(series, errors="coerce")


def _to_date(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=False)


def _to_pct(series: pd.Series) -> pd.Series:
    """Normalise fill rates to 0–100 scale.
    SAP/BW exports fill rates as 0–1 decimals (e.g. 1.0 = 100%, 0.184 = 18.4%).
    Values already on 0–100 scale (max > 1.5) are left unchanged.
    """
    s = _to_numeric(series)
    non_null = s.dropna()
    if non_null.empty:
        return s
    if non_null.max() > 1.5:   # already 0–100
        return s
    return s * 100              # 0–1 → 0–100


# ---------------------------------------------------------------------------
# Main loader
# ---------------------------------------------------------------------------

def load_fill_rate(
    file_obj: Any, threshold: float
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Returns (df_clean, df_raw).
    df_raw   = data read from the real header row, minimal processing.
    df_clean = fully cleaned + calculated dataframe.
    """
    # detect real header row
    file_obj.seek(0)
    header_row = detect_header_row(file_obj)

    # read from real header
    file_obj.seek(0)
    df_raw = pd.read_excel(
        file_obj,
        header=header_row,
        engine="openpyxl",
        dtype=str,
    )

    # drop entirely empty rows / columns
    df_raw = df_raw.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)

    # strip whitespace
    for col in df_raw.columns:
        if df_raw[col].dtype == object:
            df_raw[col] = df_raw[col].str.strip()

    # replace bare "nan" strings
    df_raw.replace({"nan": np.nan, "NaN": np.nan, "": np.nan}, inplace=True)

    # normalise column names
    df = normalize_columns(df_raw.copy())

    # fill-down SAP merged-cell pattern
    for col in ["plant", "product", "sales_order", "outbound_delivery",
                "product_group_code", "product_group_name"]:
        if col in df.columns:
            df[col] = df[col].ffill()

    # type conversions
    for col in ["order_quantity", "delivered_quantity", "short_quantity",
                "net_value", "short_amount", "total_short_amount"]:
        if col in df.columns:
            df[col] = _to_numeric(df[col])

    for col in ["wh_fill_rate", "customer_fill_rate"]:
        if col in df.columns:
            df[col] = _to_pct(df[col])

    if "requested_delivery_date" in df.columns:
        df["requested_delivery_date"] = _to_date(df["requested_delivery_date"])

    # drop rows with no order quantity or delivery ID
    key_cols = [c for c in ["order_quantity", "outbound_delivery"] if c in df.columns]
    if key_cols:
        df = df.dropna(subset=key_cols[:1]).reset_index(drop=True)

    # derive short_quantity if not present
    if "short_quantity" not in df.columns:
        if "order_quantity" in df.columns and "delivered_quantity" in df.columns:
            df["short_quantity"] = (
                df["order_quantity"] - df["delivered_quantity"]
            ).clip(lower=0)

    # ensure short_amount exists (may be zero if no delivery shortage)
    if "short_amount" not in df.columns:
        df["short_amount"] = np.nan

    # add calculated columns
    df = _add_calculated_columns(df, threshold)

    return df, df_raw


def _add_calculated_columns(df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    short_qty = df.get("short_quantity", pd.Series(0, index=df.index)).fillna(0)
    wh_rate   = df.get("wh_fill_rate",   pd.Series(100.0, index=df.index)).fillna(100)
    cust_rate = df.get("customer_fill_rate", pd.Series(100.0, index=df.index)).fillna(100)

    # Use Total_Short_Amt for priority if available; fall back to Dvl_Short_Amt
    if "total_short_amount" in df.columns:
        priority_amt = df["total_short_amount"].fillna(0)
    else:
        priority_amt = df.get("short_amount", pd.Series(0, index=df.index)).fillna(0)

    df["Shortage Status"]      = np.where(short_qty > 0, "Shorted", "Fully Delivered")
    df["WH Fill Status"]       = np.where(wh_rate   < 100, "WH Fill Rate Issue", "Good")
    df["Customer Fill Status"] = np.where(cust_rate  < 100, "Customer Impact", "Good")

    conditions = [
        priority_amt >= threshold,
        (short_qty > 0) | (priority_amt > 0),
    ]
    choices = ["High Priority", "Medium Priority"]
    df["Priority"] = np.select(conditions, choices, default="Low Priority")

    return df


def reclassify_priority(df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    """Recompute only the Priority column for a new dollar threshold.

    Lets the High-Priority threshold be a post-load setting (driven by a saved
    Variance Profile) without re-reading the file. Returns the same dataframe.
    """
    short_qty = df.get("short_quantity", pd.Series(0, index=df.index)).fillna(0)
    if "total_short_amount" in df.columns:
        priority_amt = df["total_short_amount"].fillna(0)
    else:
        priority_amt = df.get("short_amount", pd.Series(0, index=df.index)).fillna(0)

    conditions = [
        priority_amt >= threshold,
        (short_qty > 0) | (priority_amt > 0),
    ]
    df["Priority"] = np.select(conditions, ["High Priority", "Medium Priority"], default="Low Priority")
    return df


# ---------------------------------------------------------------------------
# KPI builder
# ---------------------------------------------------------------------------

def build_kpis(df: pd.DataFrame) -> dict[str, Any]:
    def _sum(col: str) -> float:
        return float(df[col].sum()) if col in df.columns else 0.0

    def _mean(col: str) -> float:
        return float(df[col].mean()) if col in df.columns else 0.0

    def _nunique(col: str) -> int:
        return int(df[col].nunique()) if col in df.columns else 0

    shorted = (
        (df.get("Shortage Status", pd.Series("Fully Delivered", index=df.index)) == "Shorted") |
        (df.get("Customer Fill Status", pd.Series("Good", index=df.index)) == "Customer Impact")
    )

    # Prefer Total_Short_Amt for the headline short amount KPI
    short_amt_col = "total_short_amount" if "total_short_amount" in df.columns else "short_amount"

    return {
        "total_order_qty":             _sum("order_quantity"),
        "total_delivered_qty":         _sum("delivered_quantity"),
        "total_short_qty":             _sum("short_quantity"),
        "total_short_amount":          _sum(short_amt_col),
        "overall_wh_fill_rate":        _mean("wh_fill_rate"),
        "overall_customer_fill_rate":  _mean("customer_fill_rate"),
        "num_deliveries":              _nunique("outbound_delivery"),
        "num_shorted_lines":           int(shorted.sum()),
        "num_products_impacted":       int(
            df.loc[shorted, "product"].nunique() if "product" in df.columns else 0
        ),
    }


# ---------------------------------------------------------------------------
# Report builders
# ---------------------------------------------------------------------------

def build_shortage_report(df: pd.DataFrame) -> pd.DataFrame:
    short_qty  = df.get("short_quantity",      pd.Series(0,     index=df.index)).fillna(0)
    short_amt  = df.get("short_amount",        pd.Series(0,     index=df.index)).fillna(0)
    total_amt  = df.get("total_short_amount",  pd.Series(0,     index=df.index)).fillna(0)
    wh_rate    = df.get("wh_fill_rate",        pd.Series(100.0, index=df.index)).fillna(100)
    cust_rate  = df.get("customer_fill_rate",  pd.Series(100.0, index=df.index)).fillna(100)

    mask = (short_qty > 0) | (short_amt > 0) | (total_amt > 0) | (wh_rate < 100) | (cust_rate < 100)
    return df[mask].reset_index(drop=True)


def _group_summary(df: pd.DataFrame, group_col: str) -> pd.DataFrame | None:
    if group_col not in df.columns:
        return None

    # Use Total_Short_Amt if available
    short_amt_src = "total_short_amount" if "total_short_amount" in df.columns else "short_amount"

    agg: dict[str, Any] = {}
    if "order_quantity"    in df.columns: agg["Order Qty"]             = ("order_quantity",    "sum")
    if "delivered_quantity" in df.columns: agg["Delivered Qty"]         = ("delivered_quantity", "sum")
    if "short_quantity"    in df.columns: agg["Short Qty"]             = ("short_quantity",    "sum")
    if short_amt_src       in df.columns: agg["Short Amount ($)"]      = (short_amt_src,       "sum")
    if "wh_fill_rate"      in df.columns: agg["Avg WH Fill Rate (%)"]  = ("wh_fill_rate",      "mean")
    if "customer_fill_rate" in df.columns: agg["Avg Cust Fill Rate (%)"] = ("customer_fill_rate", "mean")

    if not agg:
        return None

    result = df.groupby(group_col).agg(**agg).reset_index()
    result.rename(columns={group_col: group_col.replace("_", " ").title()}, inplace=True)
    return result


def build_product_summary(df: pd.DataFrame) -> pd.DataFrame | None:
    return _group_summary(df, "product")


def build_plant_summary(df: pd.DataFrame) -> pd.DataFrame | None:
    return _group_summary(df, "plant")


def build_product_group_summary(df: pd.DataFrame) -> pd.DataFrame | None:
    """Roll up by Product Group. Groups by the readable name when present,
    otherwise the code; the first column is always labelled 'Product Group'."""
    group_col = (
        "product_group_name"
        if "product_group_name" in df.columns
        else "product_group_code"
    )
    result = _group_summary(df, group_col)
    if result is not None and not result.empty:
        result.rename(columns={result.columns[0]: "Product Group"}, inplace=True)
    return result


def build_top10(df: pd.DataFrame) -> dict[str, pd.DataFrame | None]:
    short_amt_src = "total_short_amount" if "total_short_amount" in df.columns else "short_amount"

    pg_col = "product_group_name" if "product_group_name" in df.columns else "product_group_code"

    def _top(group_col: str, value_col: str, label: str, group_label: str | None = None) -> pd.DataFrame | None:
        if group_col not in df.columns or value_col not in df.columns:
            return None
        result = (
            df.groupby(group_col)[value_col]
            .sum()
            .reset_index()
            .sort_values(value_col, ascending=False)
            .head(10)
        )
        result.columns = [group_label or group_col.replace("_", " ").title(), label]
        return result

    return {
        "products_by_short_amount":       _top("product",          short_amt_src,   "Short Amount ($)"),
        "products_by_short_qty":          _top("product",          "short_quantity", "Short Qty"),
        "product_groups_by_short_amount": _top(pg_col,             short_amt_src,   "Short Amount ($)", "Product Group"),
        "plants_by_short_amount":         _top("plant",            short_amt_src,   "Short Amount ($)"),
        "deliveries_by_short_amount":     _top("outbound_delivery", short_amt_src,  "Short Amount ($)"),
    }


# ---------------------------------------------------------------------------
# Excel report generator
# ---------------------------------------------------------------------------

_HEADER_FILL    = PatternFill("solid", fgColor="1F4E79")
_KPI_FILL       = PatternFill("solid", fgColor="2E75B6")
_ALT_FILL       = PatternFill("solid", fgColor="D6E4F0")
_WHITE_FILL     = PatternFill("solid", fgColor="FFFFFF")

_HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT     = Font(bold=True, color="1F4E79", size=14)
_KPI_LABEL_FONT = Font(bold=True, color="FFFFFF", size=10)
_KPI_VALUE_FONT = Font(bold=True, color="FFFFFF", size=13)

_THIN        = Side(style="thin", color="BFBFBF")
_THIN_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left",   vertical="center")


def _autofit(ws, min_w: int = 8, max_w: int = 45) -> None:
    for col_cells in ws.columns:
        length = min_w
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length


def _style_header_row(ws, row: int, num_cols: int) -> None:
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border    = _THIN_BORDER


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    """Write dataframe; return next empty row number."""
    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col)
    _style_header_row(ws, start_row, len(df.columns))
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    for r_idx, row_data in enumerate(
        dataframe_to_rows(df, index=False, header=False), start_row + 1
    ):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = _THIN_BORDER
            cell.alignment = _LEFT
            if r_idx % 2 == 0:
                cell.fill = _ALT_FILL

    return start_row + len(df) + 2


def _apply_number_formats(ws, df: pd.DataFrame, start_row: int) -> None:
    for c_idx, col in enumerate(df.columns, 1):
        col_lower = col.lower()
        if any(k in col_lower for k in ["amount", "value", "$"]):
            fmt = '$#,##0.00'
        elif any(k in col_lower for k in ["rate", "pct", "%"]):
            fmt = '0.00"%"'
        elif any(k in col_lower for k in ["qty", "quantity", "deliveries", "lines", "products"]):
            fmt = '#,##0'
        elif "date" in col_lower:
            fmt = 'YYYY-MM-DD'
        else:
            continue
        for r_idx in range(start_row + 1, start_row + len(df) + 1):
            ws.cell(row=r_idx, column=c_idx).number_format = fmt


def _apply_shortage_cf(ws, df: pd.DataFrame, data_start_row: int) -> None:
    if "Shortage Status" not in df.columns:
        return
    status_col_idx = list(df.columns).index("Shortage Status") + 1
    col_letter = get_column_letter(status_col_idx)
    last_row   = data_start_row + len(df)
    last_col   = get_column_letter(len(df.columns))
    data_range = f"A{data_start_row}:{last_col}{last_row}"
    ws.conditional_formatting.add(
        data_range,
        FormulaRule(
            formula=[f'${col_letter}{data_start_row}="Shorted"'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
        ),
    )


def _instructions_sheet(ws) -> None:
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Delivery Fill Rate Dashboard — Excel Report", True,  16, "1F4E79"),
        ("", False, 11, "000000"),
        ("Generated automatically from a SAP/BW Delivery Fill Rate export.", False, 11, "000000"),
        ("", False, 11, "000000"),
        ("Sheets:", True, 12, "1F4E79"),
        ("  • Instructions       — This page", False, 11, "000000"),
        ("  • Executive Summary  — KPI overview for management", False, 11, "000000"),
        ("  • Clean Data         — Cleaned and enriched dataset", False, 11, "000000"),
        ("  • Shortage Report    — Problem rows only", False, 11, "000000"),
        ("  • Product Summary    — Aggregated by product", False, 11, "000000"),
        ("  • Plant Summary      — Aggregated by plant", False, 11, "000000"),
        ("  • Product Grp Summary — Aggregated by product group", False, 11, "000000"),
        ("  • Top 10 Issues      — Top offenders by short amount / qty", False, 11, "000000"),
        ("  • Raw Export Preview — First 1000 rows of the original file", False, 11, "000000"),
        ("", False, 11, "000000"),
        ("Notes:", True, 12, "1F4E79"),
        ("  • Fill rates are shown as 0–100 (e.g. 95.5 = 95.5%).", False, 11, "000000"),
        ("  • Priority uses Total_Short_Amt where available, otherwise Dvl_Short_Amt.", False, 11, "000000"),
        ("  • Only actual data rows are written — no blank rows or formulas.", False, 11, "000000"),
    ]
    for r_idx, (text, bold, size, color) in enumerate(lines, 1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=color)


def _exec_summary_sheet(ws, kpis: dict[str, Any], threshold: float) -> None:
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    title_cell = ws.cell(row=1, column=1, value="Delivery Fill Rate — Executive Summary")
    title_cell.font      = _TITLE_FONT
    title_cell.alignment = _CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    ws.cell(row=2, column=1, value=f"High Priority Threshold: ${threshold:,.2f}")
    ws.cell(row=2, column=1).font      = Font(italic=True, color="595959")
    ws.cell(row=2, column=1).alignment = _CENTER

    kpi_rows = [
        ("Total Order Quantity",        f"{kpis['total_order_qty']:,.0f}"),
        ("Total Delivered Quantity",     f"{kpis['total_delivered_qty']:,.0f}"),
        ("Total Short Quantity",         f"{kpis['total_short_qty']:,.0f}"),
        ("Total Short Amount",           f"${kpis['total_short_amount']:,.2f}"),
        ("Overall WH Fill Rate",         f"{kpis['overall_wh_fill_rate']:.1f}%"),
        ("Overall Customer Fill Rate",   f"{kpis['overall_customer_fill_rate']:.1f}%"),
        ("Outbound Deliveries",          f"{kpis['num_deliveries']:,}"),
        ("Shorted / Impacted Lines",     f"{kpis['num_shorted_lines']:,}"),
        ("Products Impacted",            f"{kpis['num_products_impacted']:,}"),
    ]
    row = 4
    for label, value in kpi_rows:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.fill = vc.fill = _KPI_FILL
        lc.font      = _KPI_LABEL_FONT;  lc.alignment = _LEFT;   lc.border = _THIN_BORDER
        vc.font      = _KPI_VALUE_FONT;  vc.alignment = _CENTER; vc.border = _THIN_BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22


def generate_excel_report(
    df_clean: pd.DataFrame,
    df_raw: pd.DataFrame,
    kpis: dict[str, Any],
    shortage_df: pd.DataFrame,
    product_df: pd.DataFrame | None,
    plant_df: pd.DataFrame | None,
    top10: dict[str, pd.DataFrame | None],
    threshold: float,
    product_group_df: pd.DataFrame | None = None,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _instructions_sheet(wb.create_sheet("Instructions"))
    _exec_summary_sheet(wb.create_sheet("Executive Summary"), kpis, threshold)

    ws_clean = wb.create_sheet("Clean Data")
    _write_df(ws_clean, df_clean, start_row=1)
    _apply_number_formats(ws_clean, df_clean, start_row=1)
    _apply_shortage_cf(ws_clean, df_clean, data_start_row=2)
    _autofit(ws_clean)

    ws_short = wb.create_sheet("Shortage Report")
    if not shortage_df.empty:
        _write_df(ws_short, shortage_df, start_row=1)
        _apply_number_formats(ws_short, shortage_df, start_row=1)
        _autofit(ws_short)
    else:
        ws_short.cell(row=1, column=1, value="No shortages found.")

    ws_prod = wb.create_sheet("Product Summary")
    if product_df is not None and not product_df.empty:
        _write_df(ws_prod, product_df, start_row=1)
        _apply_number_formats(ws_prod, product_df, start_row=1)
        _autofit(ws_prod)
    else:
        ws_prod.cell(row=1, column=1, value="Product column not found in data.")

    ws_plant = wb.create_sheet("Plant Summary")
    if plant_df is not None and not plant_df.empty:
        _write_df(ws_plant, plant_df, start_row=1)
        _apply_number_formats(ws_plant, plant_df, start_row=1)
        _autofit(ws_plant)
    else:
        ws_plant.cell(row=1, column=1, value="Plant column not found in data.")

    ws_pg = wb.create_sheet("Product Group Summary")
    if product_group_df is not None and not product_group_df.empty:
        _write_df(ws_pg, product_group_df, start_row=1)
        _apply_number_formats(ws_pg, product_group_df, start_row=1)
        _autofit(ws_pg)
    else:
        ws_pg.cell(row=1, column=1, value="Product Group column not found in data.")

    ws_top = wb.create_sheet("Top 10 Issues")
    top_row = 1
    for section_title, section_df in [
        ("Top 10 Products by Short Amount",        top10.get("products_by_short_amount")),
        ("Top 10 Products by Short Qty",           top10.get("products_by_short_qty")),
        ("Top 10 Product Groups by Short Amount",  top10.get("product_groups_by_short_amount")),
        ("Top 10 Plants by Short Amount",          top10.get("plants_by_short_amount")),
        ("Top 10 Deliveries by Short Amount",      top10.get("deliveries_by_short_amount")),
    ]:
        ws_top.cell(row=top_row, column=1, value=section_title).font = Font(bold=True, size=12, color="1F4E79")
        top_row += 1
        if section_df is not None and not section_df.empty:
            top_row = _write_df(ws_top, section_df, start_row=top_row)
        else:
            ws_top.cell(row=top_row, column=1, value="No data available.")
            top_row += 2
    _autofit(ws_top)

    ws_raw = wb.create_sheet("Raw Export Preview")
    _write_df(ws_raw, df_raw.head(1000), start_row=1)
    _autofit(ws_raw)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
