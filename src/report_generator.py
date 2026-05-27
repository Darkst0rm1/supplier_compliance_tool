"""Write the dataframes built by compliance_engine into a formatted .xlsx workbook."""
from __future__ import annotations

from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def generate_workbook(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Return the workbook as bytes ready for download or disk write."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl", datetime_format="yyyy-mm-dd") as writer:
        for name, df in sheets.items():
            sheet_name = name[:31]  # Excel sheet-name length limit

            if df is None or df.empty:
                placeholder = pd.DataFrame({"Note": [f"No data for: {name}"]})
                placeholder.to_excel(writer, sheet_name=sheet_name, index=False)
                _format_sheet(writer.sheets[sheet_name], placeholder)
            else:
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                _format_sheet(writer.sheets[sheet_name], df)

    return buf.getvalue()


def _format_sheet(ws, df: pd.DataFrame) -> None:
    """Apply header styling, freeze panes, auto-filter, and width heuristics."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"

    last_col = get_column_letter(ws.max_column)
    ws.auto_filter.ref = f"A1:{last_col}{ws.max_row}"

    # Column width = longest of header or first 200 values, clamped to [12, 50].
    for i, col in enumerate(df.columns, start=1):
        sample = df[col].head(200).tolist() if len(df) else []
        max_len = max([len(str(col))] + [len(str(v)) if v is not None else 0 for v in sample])
        ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 50)
