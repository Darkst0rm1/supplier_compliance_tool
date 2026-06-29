"""Processing engine for the Risky Inventory page.

The user downloads one report: the full 0–180 day Risky Inventory export. Each
row carries an ``MRP Last Sell Date``. This engine splits the rows into buckets
relative to a report run date (cutoff = run date + 90 days):

* ``0-90 Day``  — Last Sell Date on/before the cutoff,
* ``91-180 Day`` — Last Sell Date after the cutoff,
* ``No Last Sell Date`` — blank/missing date.

The output is a workbook whose ``Detail`` sheet holds every row with a ``Bucket``
column appended, and whose ``Summary`` sheet is a real, interactive Excel
PivotTable filterable by ``Bucket`` (and by Description p. group / Brand Manager
Desc / MRP Area). openpyxl cannot create a PivotTable, so the workbook is built
by filling a committed template (``templates/risky_inventory_template.xlsx``,
derived once from a golden export and scrubbed of supplier data) and pointing the
pivot cache at the new data with ``refreshOnLoad`` so Excel rebuilds it on open.

Source detail values are preserved verbatim: text codes stay text, dates stay
dates, negatives / zeros / blanks are untouched. The uploaded ``Sheet1`` must
match ``DETAIL_HEADERS`` exactly.

This module is intentionally self-contained — like every other engine in this
app it stands alone and does not import the other report logic.
"""
from __future__ import annotations

import copy
import io
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

TEMPLATE_PATH = Path(__file__).resolve().parent / "templates" / "risky_inventory_template.xlsx"
TEMPLATE_DETAIL_SHEET = "Detail"
TEMPLATE_SUMMARY_SHEET = "Summary"


class RiskyInventoryError(Exception):
    """Raised when an uploaded file isn't a usable Risky Inventory export."""


# ---------------------------------------------------------------------------
# Source layout (exact sheet name and headers)
# ---------------------------------------------------------------------------
DETAIL_SHEET = "Sheet1"

# The full 20-column detail header, in order. The uploaded files must match.
DETAIL_HEADERS = [
    "Material",
    "Material Description",
    "Material Group",
    "Material Group Desc.",
    "Purchasing Group",
    "Description p. group",
    "Brand Manager",
    "Brand Manager Desc",
    "MRP Area",
    "Batch",
    "SLED Offset in days",
    "Batch Expiry Date",
    "MRP Last Sell Date",
    "Quantity",
    "Base Unit of Measure",
    "Qty Val. UoM",
    "Total Stock",
    "Moving price",
    "Value",
    "Batch Comment",
]

MRP_LAST_SELL_COL = "MRP Last Sell Date"
BUCKET_COL = "Bucket"
BUCKET_0_90 = "0-90 Day"
BUCKET_91_180 = "91-180 Day"
BUCKET_NONE = "No Last Sell Date"
CUTOFF_DAYS = 90


def compute_cutoff(run_date: date) -> date:
    """The 0–90 / 91–180 dividing date: report run date + 90 days."""
    return run_date + timedelta(days=CUTOFF_DAYS)


def _as_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return None


def bucket_for(last_sell: Any, cutoff: date) -> str:
    """Bucket one MRP Last Sell Date value. Blank/non-date -> No Last Sell Date;
    on/before cutoff -> 0-90 Day; after cutoff -> 91-180 Day."""
    d = _as_date(last_sell)
    if d is None:
        return BUCKET_NONE
    return BUCKET_0_90 if d <= cutoff else BUCKET_91_180


def assign_buckets(detail: "DetailTable", cutoff: date) -> tuple["DetailTable", dict[str, int]]:
    """Return a copy of ``detail`` with a Bucket column appended to every row,
    plus per-bucket counts. Row order is preserved."""
    li = detail.index[MRP_LAST_SELL_COL]
    counts = {BUCKET_0_90: 0, BUCKET_91_180: 0, BUCKET_NONE: 0}
    new_rows = []
    for row in detail.rows:
        b = bucket_for(row[li], cutoff)
        counts[b] += 1
        new_rows.append(list(row) + [b])
    bucketed = copy.copy(detail)
    bucketed.headers = detail.headers + [BUCKET_COL]
    bucketed.rows = new_rows
    return bucketed, counts


# ---------------------------------------------------------------------------
# Detail table — values plus the formatting captured from the source sheet
# ---------------------------------------------------------------------------
@dataclass
class DetailTable:
    """The Sheet1 detail rows together with the formatting we read off it."""

    headers: list[str]
    rows: list[list[Any]]
    number_formats: dict[int, str] = field(default_factory=dict)   # 1-based col -> fmt
    column_widths: dict[str, float] = field(default_factory=dict)  # letter -> width
    header_fonts: list[Any] = field(default_factory=list)          # per-column copy
    header_fills: list[Any] = field(default_factory=list)
    header_alignments: list[Any] = field(default_factory=list)

    def __len__(self) -> int:
        return len(self.rows)

    @property
    def index(self) -> dict[str, int]:
        return {h: i for i, h in enumerate(self.headers)}


def load_detail(file_obj: Any) -> DetailTable:
    """Read ``Sheet1`` from an uploaded workbook, preserving native cell values.

    Text codes stay strings, dates stay ``datetime``, blanks stay blank. The
    per-column number formats, column widths and header cell styling are
    captured so the output workbook reproduces the source look exactly.
    """
    try:
        wb = load_workbook(file_obj, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise RiskyInventoryError(f"Could not open the workbook: {exc}") from exc

    if DETAIL_SHEET not in wb.sheetnames:
        raise RiskyInventoryError(
            f"Worksheet '{DETAIL_SHEET}' not found — this does not look like a "
            "Risky Inventory export."
        )
    ws = wb[DETAIL_SHEET]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    # Trim trailing all-empty header columns.
    while headers and headers[-1] in (None, ""):
        headers.pop()
    if [h for h in headers] != DETAIL_HEADERS:
        raise RiskyInventoryError(
            "Sheet1 column headers do not match the expected Risky Inventory "
            "layout. Expected:\n  " + " | ".join(DETAIL_HEADERS)
        )

    ncols = len(headers)
    rows: list[list[Any]] = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ncols + 1)]
        if all(v is None for v in values):
            continue  # skip fully blank rows
        rows.append(values)

    # Capture per-column number formats from the first data row (column-uniform
    # in these exports). Fall back to the header cell if there are no rows.
    number_formats: dict[int, str] = {}
    fmt_row = 2 if ws.max_row >= 2 else 1
    for c in range(1, ncols + 1):
        number_formats[c] = ws.cell(fmt_row, c).number_format

    column_widths = {
        letter: dim.width
        for letter, dim in ws.column_dimensions.items()
        if dim.width is not None
    }

    header_fonts, header_fills, header_aligns = [], [], []
    for c in range(1, ncols + 1):
        cell = ws.cell(1, c)
        header_fonts.append(copy.copy(cell.font))
        header_fills.append(copy.copy(cell.fill))
        header_aligns.append(copy.copy(cell.alignment))

    return DetailTable(
        headers=headers,
        rows=rows,
        number_formats=number_formats,
        column_widths=column_widths,
        header_fonts=header_fonts,
        header_fills=header_fills,
        header_alignments=header_aligns,
    )


# ---------------------------------------------------------------------------
# Workbook output — fill the committed PivotTable template
# ---------------------------------------------------------------------------
def generate_excel(bucketed: DetailTable) -> bytes:
    """Fill the committed PivotTable template with the bucketed detail and return
    the workbook bytes. The Summary pivot is repointed at the new data and set to
    refresh on open, so Excel rebuilds it from the Detail sheet."""
    wb = load_workbook(TEMPLATE_PATH)
    ws = wb[TEMPLATE_DETAIL_SHEET]
    if ws.max_row > 1:
        ws.delete_rows(2, ws.max_row - 1)

    for r_off, row in enumerate(bucketed.rows):
        r = r_off + 2
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            fmt = bucketed.number_formats.get(c)
            if fmt:
                cell.number_format = fmt

    n = len(bucketed.rows)
    last_col = get_column_letter(len(bucketed.headers))   # 'U'
    piv = wb[TEMPLATE_SUMMARY_SHEET]._pivots[0]
    piv.cache.cacheSource.worksheetSource.ref = f"A1:{last_col}{n + 1}"
    piv.cache.recordCount = n
    piv.cache.refreshOnLoad = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
