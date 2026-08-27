"""Forecast Validation — checks whether the current sales forecast makes sense
against actual demand.

Inputs are raw SAP Sales Order exports (one workbook per pull; real exports
come in two column layouts — one carries "Key Account #", the other carries
"Plant" / "Plant Name" instead, because SAP's UI5 export lets the field layout
be changed between pulls) and, optionally, a current forecast workbook. The
forecast workbook can be either the same "Plant / Mat # / monthly columns"
shape as the required main-sheet output, or a raw SAP MRP export (Material,
Plnt, Period/Del-finish, Planned qty).

Order Quantity is the demand signal; Invoice Quantity is the fulfillment
signal shown alongside it. They are never conflated — an open 2026 order
that hasn't invoiced yet is real demand, not zero demand.
"""
from __future__ import annotations

import calendar
import io
import re
from dataclasses import dataclass
from datetime import date

import numpy as np
import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Font

REQUIRED_PLANTS = ["2910", "2920", "2930", "2925", "2935"]

MONTH_ABBR = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug",
              "Sep", "Oct", "Nov", "Dec"]

# Template header labels (row 2 of the required main-sheet layout), in order.
# The 12 history + 12 forecast slots are filled positionally with whatever
# 12 consecutive (year, month) pairs the page computes for "now" — the labels
# themselves carry no year, matching the source template exactly.
TEMPLATE_MONTH_LABELS = ["Sept", "Oct ", "Nov", "Dec", "Jan", "Feb", "Mar",
                          "Apr", "May", "June ", "July", "Aug"]
TEMPLATE_ID_COLS = ["Plant", "Mat #", "Desc", "Brand name", "Buyer name"]

# The extended template ("output_needed_with_stock_on_hand") adds Open
# Orders and Open PO for just the first 4 forecast months (near-term supply/
# demand visibility, not a full 12-month projection) plus one Stock on Hand
# column. Note this section's own labels have NO trailing space ("Oct", not
# "Oct ") — that inconsistency is in the source template itself and is
# reproduced here deliberately, not a typo.
OPEN_MONTH_LABELS = ["Sept", "Oct", "Nov", "Dec"]


class ForecastValidationError(Exception):
    """Raised when an uploaded file is missing columns this engine needs."""


# ---------------------------------------------------------------------------
# Sales order import
# ---------------------------------------------------------------------------
SO_COLUMN_ALIASES = {
    "Sales Order Item": "Item",
    "Creation Date": "Order Date",
    "Order Quantity (CS)": "Order Qty",
    "Invoice Quantity (CS)": "Invoice Qty",
}

SO_REQUIRED_COLUMNS = ["Sales Order", "Item", "Order Date", "Material", "Order Qty", "Invoice Qty"]

# Carried through when present; absence never blocks import (Plant is the
# common one to be missing — some export pulls use Key Account # instead).
SO_OPTIONAL_COLUMNS = [
    "Material Description", "Plant", "Plant Name", "Order Status",
    "Rejection Status", "Sold To Party", "Invoice #", "Outbound Delivery #",
]

_TRAILING_DOT_ZERO = re.compile(r"\.0$")


def _clean_id(series: pd.Series) -> pd.Series:
    """Clean an identifier column (Material, Plant, Sales Order, Invoice #,
    Outbound Delivery #) that must compare equal across export pulls that
    stored it differently — one pull's numeric-formatted cell round-trips
    through pandas as "2910.0", another pull's text-formatted cell as
    "2910". Only strip that numeric-conversion artifact (trailing ".0"),
    never reformat or zero-pad — a real leading zero in the source must
    survive untouched. Getting this wrong on Plant specifically silently
    drops the majority of plant-tagged rows from every plant-level output,
    since "2910.0" never matches the required-plants list "2910"."""
    s = series.astype(str).str.strip()
    s = s.str.replace(_TRAILING_DOT_ZERO, "", regex=True)
    return s.replace({"nan": "", "None": ""})


# Backward-compatible alias — Material cleaning is the same operation.
_clean_material = _clean_id


def load_sales_orders(file) -> pd.DataFrame:
    """Read one raw SAP Sales Order export into canonical columns.

    Returns line-level rows (not yet deduplicated across files) — use
    :func:`combine_sales_orders` to merge multiple uploads into one fact
    table.
    """
    df = pd.read_excel(file, sheet_name=0, dtype={"Material": str})
    df = df.rename(columns={k: v for k, v in SO_COLUMN_ALIASES.items() if k in df.columns})

    missing = [c for c in SO_REQUIRED_COLUMNS if c not in df.columns]
    if missing:
        raise ForecastValidationError(
            f"Sales order export is missing required column(s): {', '.join(missing)}. "
            f"Found columns: {', '.join(df.columns)}"
        )

    out = pd.DataFrame()
    out["Sales Order"] = _clean_id(df["Sales Order"])
    out["Item"] = pd.to_numeric(df["Item"], errors="coerce").fillna(0).astype(int)
    out["Order Date"] = pd.to_datetime(df["Order Date"], errors="coerce")
    out["Material"] = _clean_id(df["Material"])
    out["Order Qty"] = pd.to_numeric(df["Order Qty"], errors="coerce").fillna(0)
    out["Invoice Qty"] = pd.to_numeric(df["Invoice Qty"], errors="coerce").fillna(0)

    id_like = ("Plant", "Invoice #", "Outbound Delivery #")
    for col in SO_OPTIONAL_COLUMNS:
        if col not in df.columns:
            out[col] = ""
        elif col in id_like:
            out[col] = _clean_id(df[col])
        else:
            out[col] = df[col].astype(str).str.strip().replace({"nan": ""})

    # Blank rows (a trailing Excel artifact) carry no Sales Order — drop them.
    out = out[out["Sales Order"] != ""].reset_index(drop=True)
    return out


def _first_nonblank(s: pd.Series):
    for v in s:
        if pd.notna(v) and str(v).strip() != "":
            return v
    return ""


def combine_sales_orders(dfs: list[pd.DataFrame]) -> pd.DataFrame:
    """Merge multiple sales-order exports into one fact table, one row per
    Sales Order + Item.

    Two duplication patterns have to be handled differently:

    1. The SAME line re-exported in a different column layout (e.g. one pull
       has Key Account #, another of the same data has Plant instead) —
       these must be coalesced, not summed, or Invoice Qty doubles.
    2. Genuinely repeated lines WITHIN one export, one row per partial
       invoice against the same Sales Order + Item — these must be summed,
       since each row is a distinct shipment.

    Resolved by first coalescing on the natural line key (Sales Order, Item,
    Invoice #, Outbound Delivery #) — same invoice line reported twice
    collapses to one — then summing Invoice Qty across whatever distinct
    invoice lines remain per Sales Order + Item.
    """
    if not dfs:
        return pd.DataFrame(columns=SO_REQUIRED_COLUMNS + SO_OPTIONAL_COLUMNS)

    df = pd.concat(dfs, ignore_index=True)
    df["_line_key"] = (
        df["Sales Order"].astype(str) + "|" + df["Item"].astype(str) + "|"
        + df["Invoice #"].astype(str) + "|" + df["Outbound Delivery #"].astype(str)
    )

    # "first non-blank per group" needs to be a NaN-skipping value, not a
    # per-group Python callable — pandas' built-in "first" aggregator already
    # returns the first non-NaN value in original row order (exactly what
    # _first_nonblank did) but runs vectorized in Cython. With real exports
    # (hundreds of thousands of rows collapsing through two groupbys) the
    # custom-callable version took minutes; this takes under a second.
    text_cols = ["Material Description", "Plant", "Plant Name", "Order Status",
                 "Rejection Status", "Sold To Party"]
    work = df.copy()
    for c in text_cols:
        work[c] = work[c].replace("", np.nan)

    agg = {
        "Sales Order": "first", "Item": "first", "Order Date": "first",
        "Material": "first", "Order Qty": "max", "Invoice Qty": "max",
        **{c: "first" for c in text_cols},
    }
    line_level = work.groupby("_line_key", as_index=False).agg(agg)

    result = line_level.groupby(["Sales Order", "Item"], as_index=False).agg({
        "Order Date": "first", "Material": "first", "Order Qty": "max",
        "Invoice Qty": "sum",
        **{c: "first" for c in text_cols},
    })
    for c in text_cols:
        result[c] = result[c].fillna("")

    result["Year"] = result["Order Date"].dt.year
    result["Month"] = result["Order Date"].dt.month
    result["Uninvoiced Qty"] = (result["Order Qty"] - result["Invoice Qty"]).clip(lower=0)
    with np.errstate(divide="ignore", invalid="ignore"):
        pct = result["Invoice Qty"] / result["Order Qty"]
    result["Invoice Completion %"] = pct.where(result["Order Qty"] > 0)

    def _status(row):
        if row["Order Qty"] <= 0:
            return "NO ORDER"
        if row["Invoice Qty"] >= row["Order Qty"]:
            return "FULLY INVOICED"
        if row["Invoice Qty"] > 0:
            return "PARTIALLY INVOICED"
        return "NOT INVOICED"

    result["Invoice Status"] = result.apply(_status, axis=1)
    return result


# ---------------------------------------------------------------------------
# Forecast import
# ---------------------------------------------------------------------------
_MONTH_NAME_TO_NUM = {name.lower(): i + 1 for i, name in enumerate(
    ["january", "february", "march", "april", "may", "june", "july",
     "august", "september", "october", "november", "december"]
)}
_MONTH_HEADER_RE = re.compile(
    r"^\s*(jan|feb|mar|apr|may|jun|jul|aug|sep|sept|oct|nov|dec)[a-z]*"
    r"[\s\-/']*'?(\d{2,4})?\s*$", re.IGNORECASE
)


def _parse_month_header(text: str) -> tuple[int, int] | None:
    """Parse a header like 'Sep 2026', 'September-26', 'Sept' into (year,
    month) — year is None if not present in the header text itself."""
    m = _MONTH_HEADER_RE.match(str(text))
    if not m:
        return None
    mon_txt, yr_txt = m.group(1).lower(), m.group(2)
    mon_txt = "sep" if mon_txt == "sept" else mon_txt
    month = {"jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
             "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12}[mon_txt]
    if not yr_txt:
        return (None, month)
    year = int(yr_txt)
    if year < 100:
        year += 2000
    return (year, month)


def load_forecast(file, forecast_months: list[tuple[int, int]]) -> pd.DataFrame:
    """Read a current-forecast workbook into long form: Plant, Material,
    Forecast Year, Forecast Month, Forecast Qty, Buyer Name, Brand Name.

    Detects one of two shapes:
      * Wide — Plant/Mat# + one column per calendar month (the same layout
        as the required output). Column year is inferred by matching each
        month name against ``forecast_months`` in order.
      * SAP MRP export — Material, Plnt, Del/finish (or Period), Planned qty
        — summed by calendar month.

    Returns an empty frame (not an error) if a forecast file isn't supplied;
    the page treats every forecast cell as blank ("PASTE FORECAST") in that
    case rather than fabricating anything.
    """
    # Check the SAP MRP shape first — it also has Material/Plnt columns, so
    # if this check ran second it would never fire (the wide-table check
    # below only needs Plant + Material, which the MRP shape has too).
    df_named = pd.read_excel(file, sheet_name=0)
    if "Planned qty" in df_named.columns and ("Plnt" in df_named.columns or "Plant" in df_named.columns):
        return _load_forecast_mrp(df_named)

    df = pd.read_excel(file, sheet_name=0, header=None)

    header_row_idx = None
    for i in range(min(5, len(df))):
        row_raw = df.iloc[i].tolist()
        row_vals = [str(v).strip().lower() for v in row_raw]
        has_plant = any(v in ("plant", "plnt") for v in row_vals)
        has_mat = any(v in ("mat #", "material", "material number") for v in row_vals)
        has_month = any(_parse_month_header(v) is not None for v in row_raw)
        if has_plant and has_mat and has_month:
            header_row_idx = i
            break

    if header_row_idx is not None:
        section_row = df.iloc[header_row_idx - 1] if header_row_idx > 0 else None
        return _load_forecast_wide(df, header_row_idx, forecast_months, section_row)

    raise ForecastValidationError(
        "Could not recognize the forecast file's layout — expected either a "
        "Plant/Mat#/monthly-columns table or a SAP MRP export with a "
        "'Planned qty' column."
    )


def _load_forecast_wide(raw: pd.DataFrame, header_row_idx: int,
                         forecast_months: list[tuple[int, int]],
                         section_row: pd.Series | None = None) -> pd.DataFrame:
    # Worked entirely by column POSITION, never by header name: the history
    # and forecast sections of a re-uploaded template repeat the exact same
    # labels ("Sept", "Oct "...), so a name-keyed lookup would collide (two
    # columns both called "Sept") and pandas would hand back a Series where
    # a scalar was expected.
    headers = [str(v).strip() for v in raw.iloc[header_row_idx].tolist()]
    body = raw.iloc[header_row_idx + 1:].reset_index(drop=True)

    def _find(candidates):
        for i, h in enumerate(headers):
            if h.lower() in candidates:
                return i
        return None

    plant_idx = _find({"plant", "plnt"})
    mat_idx = _find({"mat #", "material", "material number"})
    desc_idx = _find({"desc", "description"})
    brand_idx = next((i for i, h in enumerate(headers) if "brand" in h.lower()), None)
    buyer_idx = next((i for i, h in enumerate(headers) if "buyer" in h.lower()), None)
    if plant_idx is None or mat_idx is None:
        raise ForecastValidationError("Forecast file is missing a Plant or Mat # column.")

    # A merged section-label row above the headers (e.g. "History ... " /
    # "Forecast ...") only carries text in the anchor cell of each merge —
    # every other spanned cell reads NaN — so forward-fill it to know which
    # section every column belongs to.
    section_labels = None
    if section_row is not None:
        labels = section_row.ffill()
        section_labels = [str(v).strip().lower() if pd.notna(v) else "" for v in labels.tolist()]

    # Header text alone ("Sept") carries no year, so month columns are
    # matched to forecast_months positionally in appearance order. If a
    # section-label row is present, only columns under a "forecast" section
    # are eligible — this is what lets a re-uploaded template (history +
    # forecast side by side) resolve to the right 12 columns instead of the
    # first 12 (which would silently be the history ones).
    month_idxs = []
    id_idxs = {plant_idx, mat_idx, desc_idx, brand_idx, buyer_idx}
    for idx, c in enumerate(headers):
        if idx in id_idxs:
            continue
        if _parse_month_header(c) is None:
            continue
        if section_labels is not None and idx < len(section_labels) and "hist" in section_labels[idx]:
            continue
        month_idxs.append(idx)

    rows = []
    for _, r in body.iterrows():
        vals = r.tolist()
        plant = str(vals[plant_idx]).strip()
        material = _clean_material(pd.Series([vals[mat_idx]])).iloc[0]
        if not plant or plant.lower() == "nan" or not material:
            continue
        for pos, col_idx in enumerate(month_idxs):
            if pos >= len(forecast_months):
                break
            qty = pd.to_numeric(vals[col_idx], errors="coerce")
            if pd.isna(qty):
                continue
            year, month = forecast_months[pos]
            rows.append({
                "Plant": plant, "Material": material,
                "Forecast Year": year, "Forecast Month": month,
                "Forecast Qty": float(qty),
                "Buyer Name": str(vals[buyer_idx]).strip() if buyer_idx is not None else "",
                "Brand Name": str(vals[brand_idx]).strip() if brand_idx is not None else "",
                "Material Description": str(vals[desc_idx]).strip() if desc_idx is not None else "",
            })
    return pd.DataFrame(rows)


def _load_forecast_mrp(df: pd.DataFrame) -> pd.DataFrame:
    plnt_col = "Plnt" if "Plnt" in df.columns else "Plant"
    date_col = "Del/finish" if "Del/finish" in df.columns else None
    if date_col is None:
        raise ForecastValidationError(
            "SAP MRP forecast export needs a 'Del/finish' date column to "
            "determine which calendar month each planned quantity belongs to."
        )

    out = pd.DataFrame()
    out["Plant"] = _clean_id(df[plnt_col])
    out["Material"] = _clean_id(df["Material"])
    out["_date"] = pd.to_datetime(df[date_col], errors="coerce")
    out["Forecast Qty"] = pd.to_numeric(df["Planned qty"], errors="coerce").fillna(0)
    # This export's column headers don't describe their own content: "Material
    # Group" actually holds the brand/supplier grouping text (e.g. "ROBERTSONS"),
    # "Description" holds the buyer's name (e.g. "Bita Farahani"), and
    # "Material Number" holds the material's text description — confirmed by
    # cross-checking against a reference workbook built from this same
    # source. "Brand Manager" and "BDM" are NOT used here; they don't match.
    out["Brand Name"] = df["Material Group"].astype(str).str.strip() if "Material Group" in df.columns else ""
    out["Buyer Name"] = df["Description"].astype(str).str.strip() if "Description" in df.columns else ""
    out["Material Description"] = df["Material Number"].astype(str).str.strip() if "Material Number" in df.columns else ""
    out = out.dropna(subset=["_date"])
    out["Forecast Year"] = out["_date"].dt.year
    out["Forecast Month"] = out["_date"].dt.month

    grouped = out.groupby(["Plant", "Material", "Forecast Year", "Forecast Month"], as_index=False).agg({
        "Forecast Qty": "sum", "Buyer Name": _first_nonblank,
        "Brand Name": _first_nonblank, "Material Description": _first_nonblank,
    })
    return grouped


# ---------------------------------------------------------------------------
# Open Orders / Open PO / Stock on Hand
# ---------------------------------------------------------------------------
OPEN_ORDERS_REQUIRED = ["Sales Order", "Material", "Requested Delivery Date", "Order Quantity (CS)"]
OPEN_PO_REQUIRED = ["Plant", "Material", "Delivery Date", "PO Quantity"]


def load_open_orders(file) -> pd.DataFrame:
    """Read the Open Orders export — sales orders already booked (Creation
    Date) but requested for delivery in a future month. This export has no
    Plant column of its own; use :func:`resolve_open_order_plants` to attach
    one by matching back to the combined sales-order fact table."""
    df = pd.read_excel(file, sheet_name=0)
    missing = [c for c in OPEN_ORDERS_REQUIRED if c not in df.columns]
    if missing:
        raise ForecastValidationError(
            f"Open Orders export is missing required column(s): {', '.join(missing)}. "
            f"Found columns: {', '.join(df.columns)}"
        )

    out = pd.DataFrame()
    out["Sales Order"] = _clean_id(df["Sales Order"])
    out["Material"] = _clean_id(df["Material"])
    out["Requested Delivery Date"] = pd.to_datetime(df["Requested Delivery Date"], errors="coerce")
    out["Open Order Qty"] = pd.to_numeric(df["Order Quantity (CS)"], errors="coerce").fillna(0)
    out["Buyer Name"] = df["BDM Description"].astype(str).str.strip() if "BDM Description" in df.columns else ""
    out = out.dropna(subset=["Requested Delivery Date"])
    out["Year"] = out["Requested Delivery Date"].dt.year
    out["Month"] = out["Requested Delivery Date"].dt.month
    return out


def resolve_open_order_plants(open_orders: pd.DataFrame, fact: pd.DataFrame) -> pd.DataFrame:
    """Attach Plant to each Open Orders row by matching Sales Order +
    Material back to the combined sales-order fact table (which does carry
    Plant, from the SAP exports that have it). A row whose Sales Order +
    Material isn't found there keeps a blank Plant and is excluded from
    plant-level Open Orders totals — never guessed from Material alone,
    since one material can sit at more than one plant."""
    if open_orders.empty:
        return open_orders.assign(Plant="")
    lookup = (
        fact[fact["Plant"] != ""]
        .drop_duplicates(subset=["Sales Order", "Material"])[["Sales Order", "Material", "Plant"]]
    )
    merged = open_orders.merge(lookup, on=["Sales Order", "Material"], how="left")
    merged["Plant"] = merged["Plant"].fillna("")
    return merged


def load_open_po(file) -> pd.DataFrame:
    """Read the Open PO export — inbound purchase orders not yet received,
    by Plant + Material + expected Delivery month. This file's own Stock on
    Hand column is also the Stock on Hand source for the main sheet — there
    is no separate dedicated stock/inventory export."""
    df = pd.read_excel(file, sheet_name=0)
    missing = [c for c in OPEN_PO_REQUIRED if c not in df.columns]
    if missing:
        raise ForecastValidationError(
            f"Open PO export is missing required column(s): {', '.join(missing)}. "
            f"Found columns: {', '.join(df.columns)}"
        )

    out = pd.DataFrame()
    out["Plant"] = _clean_id(df["Plant"])
    out["Material"] = _clean_id(df["Material"])
    out["Delivery Date"] = pd.to_datetime(df["Delivery Date"], errors="coerce")
    out["Open PO Qty"] = pd.to_numeric(df["PO Quantity"], errors="coerce").fillna(0)
    out["Stock on Hand"] = pd.to_numeric(df["Stock on Hand"], errors="coerce") if "Stock on Hand" in df.columns else np.nan
    out = out.dropna(subset=["Delivery Date"])
    out["Year"] = out["Delivery Date"].dt.year
    out["Month"] = out["Delivery Date"].dt.month
    return out


# ---------------------------------------------------------------------------
# Month-window helpers
# ---------------------------------------------------------------------------
def _add_months(year: int, month: int, n: int) -> tuple[int, int]:
    idx = (year * 12 + (month - 1)) + n
    return idx // 12, idx % 12 + 1


def default_history_and_forecast_months(today: date) -> tuple[list[tuple[int, int]], list[tuple[int, int]]]:
    """History = Sept-Aug of the fiscal year `today` falls in (partial once
    it reaches the current month); Forecast = the next fiscal year's Sept-Aug.

    Anchored to a fixed September start, NOT a rolling "12 months back from
    today" window -- the required template's row-2 labels are the fixed
    sequence Sept, Oct, Nov, Dec, Jan, ... Aug (this business's actual
    season, tied to Christmas import buying starting in the fall), and
    those labels are matched to hist_months/fc_months purely by position in
    generate_excel. A rolling window whose first month isn't September
    would silently mislabel every column whenever `today` isn't in August
    (e.g. history ending in December would read Jan-Dec under "Sept...Aug"
    headers) -- this happened in production and is exactly the bug this
    anchoring fixes."""
    fiscal_start_year = today.year if today.month >= 9 else today.year - 1
    hist_months = [_add_months(fiscal_start_year, 9, n) for n in range(12)]
    fc_months = [_add_months(fiscal_start_year, 9, n) for n in range(12, 24)]
    return hist_months, fc_months


def same_period_window(data_as_of: date) -> tuple[date, date, date, date]:
    """1st..data_as_of of the current-month window (e.g. Aug 1-26, 2026) vs
    the same day-of-month window a year earlier — the fair partial-month
    comparison, not full-month-vs-partial."""
    cur_start = date(data_as_of.year, data_as_of.month, 1)
    cur_end = data_as_of
    prior_start = date(data_as_of.year - 1, data_as_of.month, 1)
    prior_day = min(data_as_of.day, 28) if data_as_of.month == 2 else data_as_of.day
    prior_end = date(data_as_of.year - 1, data_as_of.month, prior_day)
    return cur_start, cur_end, prior_start, prior_end


def _fmt_day(d: date) -> str:
    """'%-d' (no leading zero) isn't portable to Windows strftime — spell it
    out instead."""
    return f"{d.strftime('%b')} {d.day}"


def _is_month_complete(d: date) -> bool:
    """True if ``d`` is the last calendar day of its month — data uploaded
    through, say, Dec 31 covers a FULL December, not a partial one, even
    though December is "the month data_as_of falls in"."""
    return calendar.monthrange(d.year, d.month)[1] == d.day


# ---------------------------------------------------------------------------
# Main (template-shaped) table
# ---------------------------------------------------------------------------
@dataclass
class ForecastValidationResult:
    main_table: pd.DataFrame          # exact output-template layout
    validation: pd.DataFrame          # Plant+Material+Forecast Month assessment
    plant_summary: pd.DataFrame
    item_detail: pd.DataFrame
    monthly_summary: pd.DataFrame
    data_quality: pd.DataFrame
    data_as_of: pd.Timestamp | None


def _qty_pivot(df: pd.DataFrame, group_cols: list[str], qty_col: str) -> dict:
    """Sum ``qty_col`` per group and return it as a plain dict keyed by the
    group tuple, for O(1) lookups. Used instead of re-filtering the whole
    frame per (Plant, Material, Month) — with real exports (hundreds of
    thousands of rows, thousands of distinct materials) repeated boolean
    masking there was the difference between seconds and many minutes.

    A missing key means no row existed for that combination at all — left
    blank downstream — which is different from a key that sums to 0 (rows
    existed, e.g. an order with no invoice yet)."""
    if df.empty:
        return {}
    return df.groupby(group_cols)[qty_col].sum().to_dict()


def build_main_table(fact: pd.DataFrame, forecast: pd.DataFrame,
                      hist_months: list[tuple[int, int]],
                      fc_months: list[tuple[int, int]],
                      required_plants: list[str] = REQUIRED_PLANTS,
                      open_orders: pd.DataFrame | None = None,
                      open_po: pd.DataFrame | None = None) -> pd.DataFrame:
    """Build the main sheet in the exact required layout: Plant, Mat #, Desc,
    Brand name, Buyer name, then 12 history (actual Invoice Qty) columns,
    12 forecast columns, then — when supplied — Open Orders and Open PO for
    the first 4 forecast months and one Stock on Hand column, matching the
    extended "output_needed_with_stock_on_hand" template. Stock on Hand
    comes from the Open PO export's own Stock on Hand column — there is no
    separate stock/inventory source. One row per Plant + Material actually
    observed anywhere (sales history, forecast, open orders, or open PO) —
    a brand-new item with only an open PO still shows up."""
    plant_fact = fact[fact["Plant"].isin(required_plants) & (fact["Material"] != "")]
    open_orders = open_orders if open_orders is not None else pd.DataFrame(columns=["Plant", "Material", "Year", "Month", "Open Order Qty"])
    open_po = open_po if open_po is not None else pd.DataFrame(columns=["Plant", "Material", "Year", "Month", "Open PO Qty", "Stock on Hand"])

    keys = set(zip(plant_fact["Plant"], plant_fact["Material"]))
    if not forecast.empty:
        keys |= set(zip(forecast["Plant"], forecast["Material"]))
    if not open_orders.empty:
        keys |= set(zip(open_orders["Plant"], open_orders["Material"]))
    if not open_po.empty:
        keys |= set(zip(open_po["Plant"], open_po["Material"]))
    keys = {(p, m) for p, m in keys if p in required_plants and m}

    desc_lookup = (
        plant_fact.sort_values("Order Date")
        .groupby(["Plant", "Material"])["Material Description"]
        .agg(_first_nonblank)
    )
    if not forecast.empty and "Material Description" in forecast.columns:
        fc_desc = forecast.groupby(["Plant", "Material"])["Material Description"].agg(_first_nonblank)
    else:
        fc_desc = pd.Series(dtype=str)
    buyer_lookup = forecast.groupby(["Plant", "Material"])["Buyer Name"].agg(_first_nonblank) if not forecast.empty else pd.Series(dtype=str)
    brand_lookup = forecast.groupby(["Plant", "Material"])["Brand Name"].agg(_first_nonblank) if not forecast.empty else pd.Series(dtype=str)

    hist_pivot = _qty_pivot(plant_fact, ["Plant", "Material", "Year", "Month"], "Invoice Qty")
    fc_pivot = _qty_pivot(forecast, ["Plant", "Material", "Forecast Year", "Forecast Month"], "Forecast Qty")
    oo_pivot = _qty_pivot(open_orders, ["Plant", "Material", "Year", "Month"], "Open Order Qty")
    po_pivot = _qty_pivot(open_po, ["Plant", "Material", "Year", "Month"], "Open PO Qty")

    # Stock on Hand comes from the Open PO export's own column — only covers
    # materials that appear there (an open PO), which matches what's
    # actually available; nothing is fabricated for the rest.
    if not open_po.empty and "Stock on Hand" in open_po.columns:
        stock_lookup = (
            open_po.dropna(subset=["Stock on Hand"])
            .groupby(["Plant", "Material"])["Stock on Hand"].first().to_dict()
        )
    else:
        stock_lookup = {}

    open_months = fc_months[:len(OPEN_MONTH_LABELS)]

    rows = []
    for plant, material in sorted(keys, key=lambda k: (required_plants.index(k[0]) if k[0] in required_plants else 99, k[1])):
        desc = desc_lookup.get((plant, material), "") or fc_desc.get((plant, material), "")
        row = {
            "Plant": int(plant) if str(plant).isdigit() else plant,
            "Mat #": material,
            "Desc": desc,
            "Brand name": brand_lookup.get((plant, material), ""),
            "Buyer name": buyer_lookup.get((plant, material), ""),
        }
        for label, (year, month) in zip(TEMPLATE_MONTH_LABELS, hist_months):
            row[f"H|{label}"] = hist_pivot.get((plant, material, year, month))
        for label, (year, month) in zip(TEMPLATE_MONTH_LABELS, fc_months):
            row[f"F|{label}"] = fc_pivot.get((plant, material, year, month))
        for label, (year, month) in zip(OPEN_MONTH_LABELS, open_months):
            row[f"OO|{label}"] = oo_pivot.get((plant, material, year, month))
        for label, (year, month) in zip(OPEN_MONTH_LABELS, open_months):
            row[f"PO|{label}"] = po_pivot.get((plant, material, year, month))
        row["Stock on Hand"] = stock_lookup.get((plant, material))
        rows.append(row)

    out = pd.DataFrame(rows)
    if out.empty:
        cols = (TEMPLATE_ID_COLS + [f"H|{l}" for l in TEMPLATE_MONTH_LABELS]
                + [f"F|{l}" for l in TEMPLATE_MONTH_LABELS]
                + [f"OO|{l}" for l in OPEN_MONTH_LABELS] + [f"PO|{l}" for l in OPEN_MONTH_LABELS]
                + ["Stock on Hand"])
        return pd.DataFrame(columns=cols)
    return out


# ---------------------------------------------------------------------------
# Forecast validation
# ---------------------------------------------------------------------------
def _assess(current: float | None, baseline: float | None) -> str:
    if current is None or pd.isna(current):
        return "PASTE FORECAST"
    if baseline is None or pd.isna(baseline) or baseline == 0:
        return "NEW / NO HISTORY" if current > 0 else "PASTE FORECAST"
    growth = current / baseline - 1
    if growth > 0.20:
        return "HIGH"
    if growth < -0.20:
        return "LOW"
    if abs(growth) <= 0.10:
        return "REASONABLE"
    return "REVIEW"


def build_forecast_validation(fact: pd.DataFrame, forecast: pd.DataFrame,
                               fc_months: list[tuple[int, int]],
                               data_as_of: pd.Timestamp,
                               required_plants: list[str] = REQUIRED_PLANTS) -> pd.DataFrame:
    """One row per Plant + Material + Forecast Month, comparing the current
    forecast to a same-month-last-year baseline scaled by the current
    same-period (Aug 1..data_as_of) order growth factor for that Plant +
    Material — falling back to Plant-level, then company-level growth when
    the item's own same-period volume is too thin to trust."""
    plant_fact = fact[fact["Plant"].isin(required_plants)]
    cur_start, cur_end, prior_start, prior_end = same_period_window(data_as_of.date())

    in_cur = (plant_fact["Order Date"] >= pd.Timestamp(cur_start)) & (plant_fact["Order Date"] <= pd.Timestamp(cur_end))
    in_prior = (plant_fact["Order Date"] >= pd.Timestamp(prior_start)) & (plant_fact["Order Date"] <= pd.Timestamp(prior_end))

    company_cur = plant_fact.loc[in_cur, "Order Qty"].sum()
    company_prior = plant_fact.loc[in_prior, "Order Qty"].sum()
    company_factor = (company_cur / company_prior) if company_prior > 0 else None

    plant_cur = plant_fact.loc[in_cur].groupby("Plant")["Order Qty"].sum()
    plant_prior = plant_fact.loc[in_prior].groupby("Plant")["Order Qty"].sum()
    plant_factor = {
        p: (plant_cur.get(p, 0) / plant_prior[p]) for p in plant_prior.index if plant_prior[p] > 0
    }

    # Item-level same-period order qty and prior-year monthly Order/Invoice
    # Qty, precomputed once as dicts — looping per (Plant, Material) and
    # re-filtering the whole fact table each time does not scale past a few
    # hundred thousand rows / a few thousand distinct materials.
    item_cur_map = plant_fact.loc[in_cur].groupby(["Plant", "Material"])["Order Qty"].sum().to_dict()
    item_prior_map = plant_fact.loc[in_prior].groupby(["Plant", "Material"])["Order Qty"].sum().to_dict()
    order_pivot = _qty_pivot(plant_fact, ["Plant", "Material", "Year", "Month"], "Order Qty")
    invoice_pivot = _qty_pivot(plant_fact, ["Plant", "Material", "Year", "Month"], "Invoice Qty")
    fc_pivot = _qty_pivot(forecast, ["Plant", "Material", "Forecast Year", "Forecast Month"], "Forecast Qty")
    desc_lookup = plant_fact.groupby(["Plant", "Material"])["Material Description"].agg(_first_nonblank)

    rows = []
    keys = set(zip(plant_fact["Plant"], plant_fact["Material"]))
    if not forecast.empty:
        keys |= set(zip(forecast["Plant"], forecast["Material"]))
    keys = {(p, m) for p, m in keys if p in required_plants and m}

    for plant, material in keys:
        item_cur = item_cur_map.get((plant, material), 0)
        item_prior = item_prior_map.get((plant, material), 0)
        if item_prior > 0:
            growth_factor, factor_source = item_cur / item_prior, "Material"
        elif plant_factor.get(plant):
            growth_factor, factor_source = plant_factor[plant], "Plant"
        elif company_factor:
            growth_factor, factor_source = company_factor, "Company"
        else:
            growth_factor, factor_source = None, "NO AUGUST BASE"

        desc = desc_lookup.get((plant, material), "")

        for label, (year, month) in zip(TEMPLATE_MONTH_LABELS, fc_months):
            prior_year, prior_month = year - 1, month
            hist_order = order_pivot.get((plant, material, prior_year, prior_month))
            hist_invoice = invoice_pivot.get((plant, material, prior_year, prior_month))

            current_forecast = fc_pivot.get((plant, material, year, month))

            # Priority: nothing to grade yet > no prior-year data at all >
            # have a raw prior-year figure but no growth factor to scale it >
            # normal math against the scaled baseline.
            if current_forecast is None:
                assessment = "PASTE FORECAST"
                baseline = hist_order * growth_factor if (hist_order is not None and growth_factor is not None) else hist_order
            elif hist_order is None:
                baseline, assessment = None, "NEW / NO HISTORY"
            elif growth_factor is None:
                baseline, assessment = hist_order, "NO AUGUST BASE"
            else:
                baseline = hist_order * growth_factor
                assessment = _assess(current_forecast, baseline)

            variance = (current_forecast - baseline) if (current_forecast is not None and baseline is not None) else None
            growth_pct = (current_forecast / baseline - 1) if (current_forecast is not None and baseline) else None

            rows.append({
                "Plant": plant, "Material": material, "Material Description": desc,
                "Forecast Month": f"{label.strip()} {year}",
                "Prior-Year Same-Month Order Qty": hist_order,
                "Prior-Year Same-Month Invoice Qty": hist_invoice,
                "Growth Factor Basis": factor_source,
                "Order Growth %": (growth_factor - 1) if growth_factor is not None else None,
                "Historical Baseline": baseline,
                "Current Forecast": current_forecast,
                "Forecast Variance": variance,
                "Forecast Growth %": growth_pct,
                "Assessment": assessment,
            })

    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Plant Summary
# ---------------------------------------------------------------------------
def build_plant_summary(fact: pd.DataFrame, data_as_of: pd.Timestamp,
                         required_plants: list[str] = REQUIRED_PLANTS) -> pd.DataFrame:
    cur_start, cur_end, prior_start, prior_end = same_period_window(data_as_of.date())
    rows = []
    for plant in required_plants:
        g = fact[fact["Plant"] == plant]
        cur = g[(g["Order Date"] >= pd.Timestamp(cur_start)) & (g["Order Date"] <= pd.Timestamp(cur_end))]
        prior = g[(g["Order Date"] >= pd.Timestamp(prior_start)) & (g["Order Date"] <= pd.Timestamp(prior_end))]

        plant_name = _first_nonblank(g["Plant Name"])
        cur_order, cur_inv = cur["Order Qty"].sum(), cur["Invoice Qty"].sum()
        prior_order, prior_inv = prior["Order Qty"].sum(), prior["Invoice Qty"].sum()
        order_growth = (cur_order / prior_order - 1) if prior_order > 0 else None
        invoice_growth = (cur_inv / prior_inv - 1) if prior_inv > 0 else None

        if order_growth is None:
            signal = "NO PRIOR-YEAR DATA"
        elif order_growth >= 0.10:
            signal = "GROWING"
        elif order_growth <= -0.10:
            signal = "DECLINING"
        else:
            signal = "STABLE"

        prior_label = f"{_fmt_day(prior_start)}-{prior_end.day}, {prior_end.year}"
        cur_label = f"{_fmt_day(cur_start)}-{cur_end.day}, {cur_end.year}"
        rows.append({
            "Plant": plant, "Plant Name": plant_name,
            f"{prior_label} Order Qty": prior_order,
            f"{prior_label} Invoice Qty": prior_inv,
            f"{cur_label} Order Qty": cur_order,
            f"{cur_label} Invoice Qty": cur_inv,
            "Order Growth %": order_growth,
            "Invoice Growth %": invoice_growth,
            "Demand Signal": signal,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Item Detail (long form, every month actually present in the data)
# ---------------------------------------------------------------------------
def build_item_detail(fact: pd.DataFrame, required_plants: list[str] = REQUIRED_PLANTS) -> pd.DataFrame:
    g = fact[fact["Plant"].isin(required_plants)]
    grouped = g.groupby(["Plant", "Material", "Year", "Month"], as_index=False).agg(
        **{
            "Material Description": ("Material Description", _first_nonblank),
            "Order Qty": ("Order Qty", "sum"),
            "Invoice Qty": ("Invoice Qty", "sum"),
        }
    )
    grouped["Uninvoiced Qty"] = (grouped["Order Qty"] - grouped["Invoice Qty"]).clip(lower=0)
    with np.errstate(divide="ignore", invalid="ignore"):
        pct = grouped["Invoice Qty"] / grouped["Order Qty"]
    grouped["Invoice Completion %"] = pct.where(grouped["Order Qty"] > 0)
    grouped["Period"] = grouped.apply(lambda r: f"{MONTH_ABBR[int(r['Month']) - 1]} {int(r['Year'])}", axis=1)
    return grouped.sort_values(["Plant", "Material", "Year", "Month"])[[
        "Plant", "Material", "Material Description", "Period", "Order Qty",
        "Invoice Qty", "Uninvoiced Qty", "Invoice Completion %",
    ]].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Monthly Summary
# ---------------------------------------------------------------------------
def build_monthly_summary(fact: pd.DataFrame, data_as_of: pd.Timestamp) -> pd.DataFrame:
    grouped = fact.groupby(["Year", "Month"], as_index=False).agg(
        **{"Order Qty": ("Order Qty", "sum"), "Invoice Qty": ("Invoice Qty", "sum")}
    )
    grouped["Uninvoiced Qty"] = (grouped["Order Qty"] - grouped["Invoice Qty"]).clip(lower=0)
    with np.errstate(divide="ignore", invalid="ignore"):
        pct = grouped["Invoice Qty"] / grouped["Order Qty"]
    grouped["Invoice Completion %"] = pct.where(grouped["Order Qty"] > 0)
    grouped["Period"] = grouped.apply(lambda r: f"{MONTH_ABBR[int(r['Month']) - 1]} {int(r['Year'])}", axis=1)
    is_current_month_partial = not _is_month_complete(data_as_of.date())
    grouped["Period Type"] = grouped.apply(
        lambda r: "PARTIAL MONTH (through {:%b %d, %Y})".format(data_as_of)
        if is_current_month_partial and (int(r["Year"]), int(r["Month"])) == (data_as_of.year, data_as_of.month)
        else "FULL MONTH",
        axis=1,
    )
    return grouped.sort_values(["Year", "Month"])[[
        "Period", "Order Qty", "Invoice Qty", "Uninvoiced Qty", "Invoice Completion %", "Period Type",
    ]].reset_index(drop=True)


# ---------------------------------------------------------------------------
# Data Quality
# ---------------------------------------------------------------------------
def build_data_quality(fact: pd.DataFrame, forecast: pd.DataFrame,
                        hist_months: list[tuple[int, int]],
                        required_plants: list[str] = REQUIRED_PLANTS,
                        open_orders: pd.DataFrame | None = None,
                        open_po: pd.DataFrame | None = None) -> pd.DataFrame:
    checks = []

    def add(label, value):
        checks.append({"Check": label, "Result": value})

    add("Total sales order lines (Sales Order + Item)", f"{len(fact):,}")
    add("Rows missing Plant (excluded from plant-level analysis)", f"{(fact['Plant'] == '').sum():,}")
    add("Rows with blank Material", f"{(fact['Material'] == '').sum():,}")
    add("Rows with negative Order Qty", f"{(fact['Order Qty'] < 0).sum():,}")
    add("Rows with negative Invoice Qty", f"{(fact['Invoice Qty'] < 0).sum():,}")
    add("Rows with Invoice Qty > Order Qty", f"{(fact['Invoice Qty'] > fact['Order Qty']).sum():,}")
    add("Date range in uploaded sales order data",
        f"{fact['Order Date'].min():%Y-%m-%d} to {fact['Order Date'].max():%Y-%m-%d}" if len(fact) else "n/a")

    have = {(int(y), int(m)) for y, m in zip(fact["Year"].dropna(), fact["Month"].dropna())}
    missing_months = [f"{MONTH_ABBR[m-1]} {y}" for y, m in hist_months if (y, m) not in have]
    add("History months with no uploaded data (left blank, not fabricated)",
        ", ".join(missing_months) if missing_months else "none")

    present_plants = set(fact.loc[fact["Plant"] != "", "Plant"].unique())
    missing_plants = [p for p in required_plants if p not in present_plants]
    add("Required plants with no data uploaded", ", ".join(missing_plants) if missing_plants else "none")

    add("Forecast rows loaded", f"{len(forecast):,}" if forecast is not None else "0")
    if forecast is not None and not forecast.empty:
        fc_plants = set(forecast["Plant"].unique()) - set(required_plants)
        if fc_plants:
            add("Forecast rows for plants outside the required 5", ", ".join(sorted(fc_plants)))

    if open_orders is not None and not open_orders.empty:
        add("Open Orders rows loaded", f"{len(open_orders):,}")
        unresolved = int((open_orders.get("Plant", "") == "").sum()) if "Plant" in open_orders.columns else len(open_orders)
        add("Open Orders rows with no matching Sales Order + Material in the "
            "uploaded sales-order exports (Plant unknown, excluded from "
            "plant-level Open Orders)", f"{unresolved:,}")
    else:
        add("Open Orders rows loaded", "0")

    if open_po is not None and not open_po.empty:
        add("Open PO rows loaded", f"{len(open_po):,}")
        po_plants = set(open_po["Plant"].unique()) - set(required_plants)
        if po_plants:
            add("Open PO rows for plants outside the required 5", ", ".join(sorted(po_plants)))
    else:
        add("Open PO rows loaded", "0")

    if open_po is not None and not open_po.empty and "Stock on Hand" in open_po.columns:
        covered = int(open_po["Stock on Hand"].notna().groupby([open_po["Plant"], open_po["Material"]]).any().sum())
        add("Stock on Hand rows loaded (from the Open PO export)",
            f"{covered:,} distinct Plant + Material — only covers materials with an open PO")
    else:
        add("Stock on Hand source", "none uploaded")

    return pd.DataFrame(checks)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------
def generate_excel(main_table: pd.DataFrame, validation: pd.DataFrame,
                    plant_summary: pd.DataFrame, item_detail: pd.DataFrame,
                    monthly_summary: pd.DataFrame, data_quality: pd.DataFrame,
                    hist_months: list[tuple[int, int]],
                    fc_months: list[tuple[int, int]],
                    data_as_of: pd.Timestamp | None) -> bytes:
    """Build the workbook. Sheet1 reproduces the required output-template
    layout exactly (two header rows, Plant/Mat#/Desc/Brand/Buyer + 12 history
    + 12 forecast columns); the rest are supporting analysis sheets."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    bold = Font(bold=True)
    center = Alignment(horizontal="center")

    hist_span = f"{hist_months[0][0]}" if hist_months[0][0] == hist_months[-1][0] else \
        f"{hist_months[0][0]} and {hist_months[-1][0]}"
    fc_span = f"{fc_months[0][0]}-{fc_months[-1][0]}"

    ws.cell(1, 6, f"History {hist_span} - cases sold (invoiced)").font = bold
    ws.cell(1, 18, f"Forecast {fc_span}").font = bold
    ws.merge_cells(start_row=1, start_column=6, end_row=1, end_column=17)
    ws.merge_cells(start_row=1, start_column=18, end_row=1, end_column=29)
    ws.cell(1, 6).alignment = center
    ws.cell(1, 18).alignment = center

    hist_labels = list(TEMPLATE_MONTH_LABELS)
    if data_as_of is not None and not _is_month_complete(data_as_of.date()):
        for i, (y, m) in enumerate(hist_months):
            if (y, m) == (data_as_of.year, data_as_of.month):
                hist_labels[i] = hist_labels[i].strip() + " (partial)"

    # Open Orders / Open PO / Stock on Hand — the extended template columns.
    # main_table always carries these (build_main_table fills them blank
    # when no Open Orders/Open PO/stock file was supplied), so the layout
    # here is unconditional and matches "output_needed_with_stock_on_hand".
    has_open_cols = any(c.startswith(("OO|", "PO|")) or c == "Stock on Hand" for c in main_table.columns)
    if has_open_cols:
        open_year = fc_months[0][0]
        ws.cell(1, 30, f"Open Orders {open_year}").font = bold
        ws.cell(1, 34, f"Open PO {open_year}").font = bold
        ws.cell(1, 38, "Stock on Hand").font = bold
        ws.merge_cells(start_row=1, start_column=30, end_row=1, end_column=33)
        ws.merge_cells(start_row=1, start_column=34, end_row=1, end_column=37)
        ws.merge_cells(start_row=1, start_column=38, end_row=2, end_column=38)
        ws.cell(1, 30).alignment = center
        ws.cell(1, 34).alignment = center
        ws.cell(1, 38).alignment = Alignment(horizontal="center", vertical="center")

    header = TEMPLATE_ID_COLS + hist_labels + list(TEMPLATE_MONTH_LABELS)
    if has_open_cols:
        header += list(OPEN_MONTH_LABELS) + list(OPEN_MONTH_LABELS)
    for c, label in enumerate(header, start=1):
        cell = ws.cell(2, c, label)
        cell.font = bold

    hist_cols = [f"H|{l}" for l in TEMPLATE_MONTH_LABELS]
    fc_cols = [f"F|{l}" for l in TEMPLATE_MONTH_LABELS]
    oo_cols = [f"OO|{l}" for l in OPEN_MONTH_LABELS]
    po_cols = [f"PO|{l}" for l in OPEN_MONTH_LABELS]
    for r, row_d in enumerate(main_table.to_dict("records"), start=3):
        values = [row_d.get(c) for c in TEMPLATE_ID_COLS] + \
                 [row_d.get(c) for c in hist_cols] + \
                 [row_d.get(c) for c in fc_cols]
        if has_open_cols:
            values += [row_d.get(c) for c in oo_cols] + [row_d.get(c) for c in po_cols] \
                      + [row_d.get("Stock on Hand")]
        for c, v in enumerate(values, start=1):
            is_blank = v is None or (isinstance(v, float) and pd.isna(v))
            ws.cell(r, c, None if is_blank else v)

    for sheet_name, df in [
        ("Forecast Validation", validation),
        ("Plant Summary", plant_summary),
        ("Item Detail", item_detail),
        ("Monthly Summary", monthly_summary),
        ("Data Quality", data_quality),
    ]:
        ws2 = wb.create_sheet(sheet_name)
        if df is None or df.empty:
            ws2.cell(1, 1, "No data.")
            continue
        for c, col in enumerate(df.columns, start=1):
            ws2.cell(1, c, col).font = bold
        for r, row in enumerate(df.itertuples(index=False), start=2):
            for c, v in enumerate(row, start=1):
                ws2.cell(r, c, (v if not (isinstance(v, float) and pd.isna(v)) else None))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
