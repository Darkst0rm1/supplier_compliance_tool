"""Processing engine for SAP/BW Sales Order Fill Rate exports."""
from __future__ import annotations

import io
import re
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------------------------------------------------------------------------
# Column keyword mapping
# IMPORTANT: unconfirmed_qty must come before unconfirmed_demand_amount so
# "Unconfirmed Demand Quantity" maps to qty, not to the dollar amount column.
# ---------------------------------------------------------------------------
COLUMN_KEYWORDS: dict[str, list[str]] = {
    "sales_order":    ["sales order", "sales document", "so number", "order number"],
    "key_account":    ["key account", "account name", "sold to", "customer name"],
    "requested_delivery_date": [
        "requested delivery date", "req. del", "rdd", "delivery date", "req del",
    ],
    "cdm_name":       ["cdm name", "cdm", "sales rep", "account manager"],
    "product":        ["product", "material", "matnr", "sku", "article"],
    "plant":          ["plant", "werk"],
    "order_qty":      ["order qty", "order quantity", "ordered qty", "open qty"],
    "confirmed_qty":  ["confd qty", "confirmed qty", "cumltv confd", "conf qty", "confirmed quantity"],
    # unconfirmed_qty MUST come before unconfirmed_demand_amount
    "unconfirmed_qty": [
        "unconfirmed demand quantity",   # full phrase matches before "unconfirmed demand"
        "unconfirmed demand qty",
        "unconf demand qty",
        "unconfirmed qty",
        "unconf qty",
    ],
    "fill_rate":      ["fill rate"],
    "net_value":      ["net value", "net val", "value lc", "total value"],
    "unconfirmed_demand_amount": [
        "unconfirmed demand",            # shorter — matches after qty column is already mapped
        "unconf demand",
        "demand amount",
    ],
}


# ---------------------------------------------------------------------------
# Column name normalisation
# Handles: camelCase splitting, underscores → spaces, parentheses removal
# Examples:
#   Cumltv Confd Qty(BU)        → cumltv confd qty bu
#   Unconfirmed Demand Quantity → unconfirmed demand quantity
#   Fill Rate                   → fill rate
# ---------------------------------------------------------------------------

def _normalize_col(col_name: str) -> str:
    s = str(col_name).strip()
    s = re.sub(r'([a-z])([A-Z])', r'\1 \2', s)
    s = s.lower()
    s = s.replace("_", " ").replace("(", " ").replace(")", " ").replace(".", " ")
    return re.sub(r'\s+', ' ', s).strip()


# ---------------------------------------------------------------------------
# Header detection
# ---------------------------------------------------------------------------

def _score_row(row: pd.Series) -> int:
    score = 0
    for val in row:
        if not isinstance(val, str):
            continue
        v = _normalize_col(val)
        for aliases in COLUMN_KEYWORDS.values():
            if any(a in v for a in aliases):
                score += 1
                break
    return score


def detect_header_row(file_obj: Any) -> int:
    raw = pd.read_excel(file_obj, header=None, nrows=50, engine="openpyxl")
    scores = [(i, _score_row(row)) for i, row in raw.iterrows()]
    best_idx, best_score = max(scores, key=lambda x: x[1])
    return int(best_idx) if best_score > 0 else 0


# ---------------------------------------------------------------------------
# Column normalisation
# ---------------------------------------------------------------------------

def _match_column(col_name: str) -> str | None:
    v = _normalize_col(col_name)
    for canonical, aliases in COLUMN_KEYWORDS.items():
        if any(a in v for a in aliases):
            return canonical
    return None


def normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map: dict[str, str] = {}
    seen: set[str] = set()
    for col in df.columns:
        key = _match_column(str(col))
        if key and key not in seen:
            rename_map[col] = key
            seen.add(key)
    return df.rename(columns=rename_map)


# ---------------------------------------------------------------------------
# Type conversion helpers
# ---------------------------------------------------------------------------

def _to_numeric(series: pd.Series) -> pd.Series:
    if series.dtype == object:
        series = (
            series.astype(str)
            .str.replace(r"[$,%\s]", "", regex=True)
            .str.replace(",", "", regex=False)
        )
    return pd.to_numeric(series, errors="coerce")


def _to_date(series: pd.Series) -> pd.Series:
    return pd.to_datetime(series, errors="coerce", dayfirst=False)


def _to_pct(series: pd.Series) -> pd.Series:
    """SAP exports fill rate as 0–1 decimal. Normalise to 0–100."""
    s = _to_numeric(series)
    non_null = s.dropna()
    if non_null.empty:
        return s
    if non_null.max() > 1.5:
        return s
    return s * 100


# ---------------------------------------------------------------------------
# Main loader
# ---------------------------------------------------------------------------

def load_sales_order(
    file_obj: Any, threshold: float
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Return (df_clean, df_raw)."""
    file_obj.seek(0)
    header_row = detect_header_row(file_obj)

    file_obj.seek(0)
    df_raw = pd.read_excel(file_obj, header=header_row, engine="openpyxl", dtype=str)

    df_raw = df_raw.dropna(how="all").dropna(axis=1, how="all").reset_index(drop=True)

    for col in df_raw.columns:
        if df_raw[col].dtype == object:
            df_raw[col] = df_raw[col].str.strip()

    df_raw.replace({"nan": np.nan, "NaN": np.nan, "": np.nan}, inplace=True)

    df = normalize_columns(df_raw.copy())

    # fill-down SAP merged-cell pattern for key grouping fields
    for col in ["sales_order", "key_account", "cdm_name", "plant"]:
        if col in df.columns:
            df[col] = df[col].ffill()

    # numeric conversions
    for col in ["order_qty", "confirmed_qty", "unconfirmed_qty", "net_value",
                "unconfirmed_demand_amount"]:
        if col in df.columns:
            df[col] = _to_numeric(df[col])

    # fill rate as percentage
    if "fill_rate" in df.columns:
        df["fill_rate"] = _to_pct(df["fill_rate"])

    # date
    if "requested_delivery_date" in df.columns:
        df["requested_delivery_date"] = _to_date(df["requested_delivery_date"])

    # derive unconfirmed_qty if missing
    if "unconfirmed_qty" not in df.columns:
        if "order_qty" in df.columns and "confirmed_qty" in df.columns:
            df["unconfirmed_qty"] = (df["order_qty"] - df["confirmed_qty"]).clip(lower=0)

    # drop rows with no order qty
    if "order_qty" in df.columns:
        df = df.dropna(subset=["order_qty"]).reset_index(drop=True)

    df = _add_calculated_columns(df, threshold)
    return df, df_raw


def _add_calculated_columns(df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    unconf_qty = df.get("unconfirmed_qty", pd.Series(0, index=df.index)).fillna(0)
    fill_rate  = df.get("fill_rate",       pd.Series(100.0, index=df.index)).fillna(100)
    unc_amt    = df.get("unconfirmed_demand_amount", pd.Series(0, index=df.index)).fillna(0)

    # Demand Status
    df["Demand Status"] = np.where(unconf_qty > 0, "Unconfirmed Demand", "Fully Confirmed")

    # Fill Rate Status
    fr_conditions = [fill_rate >= 100, fill_rate >= 95]
    fr_choices    = ["Fully Filled", "Minor Gap"]
    df["Fill Rate Status"] = np.select(fr_conditions, fr_choices, default="Fill Rate Issue")

    # Priority
    p_conditions = [unc_amt >= threshold, unconf_qty > 0]
    p_choices    = ["High Priority", "Medium Priority"]
    df["Priority"] = np.select(p_conditions, p_choices, default="Low Priority")

    # Action Required
    a_conditions = [unc_amt >= threshold, unconf_qty > 0]
    a_choices    = ["Review with supply/planning", "Monitor"]
    df["Action Required"] = np.select(a_conditions, a_choices, default="No action required")

    return df


def reclassify_priority(df: pd.DataFrame, threshold: float) -> pd.DataFrame:
    """Recompute Priority + Action Required for a new dollar threshold.

    Lets the High-Priority threshold be a post-load setting (driven by a saved
    Variance Profile) without re-reading the file. Returns the same dataframe.
    """
    unconf_qty = df.get("unconfirmed_qty", pd.Series(0, index=df.index)).fillna(0)
    unc_amt    = df.get("unconfirmed_demand_amount", pd.Series(0, index=df.index)).fillna(0)
    conditions = [unc_amt >= threshold, unconf_qty > 0]
    df["Priority"] = np.select(conditions, ["High Priority", "Medium Priority"], default="Low Priority")
    df["Action Required"] = np.select(
        conditions, ["Review with supply/planning", "Monitor"], default="No action required"
    )
    return df


# ---------------------------------------------------------------------------
# KPI builder
# ---------------------------------------------------------------------------

def build_kpis(df: pd.DataFrame) -> dict[str, Any]:
    def _sum(col: str) -> float:
        return float(df[col].sum()) if col in df.columns else 0.0

    def _nunique(col: str) -> int:
        return int(df[col].nunique()) if col in df.columns else 0

    total_order = _sum("order_qty")
    total_conf  = _sum("confirmed_qty")
    overall_fr  = (total_conf / total_order * 100) if total_order > 0 else 0.0

    unconf_mask = df.get("unconfirmed_qty", pd.Series(0, index=df.index)).fillna(0) > 0

    return {
        "total_order_qty":             total_order,
        "total_confirmed_qty":         total_conf,
        "total_unconfirmed_qty":       _sum("unconfirmed_qty"),
        "total_net_value":             _sum("net_value"),
        "total_unconfirmed_amount":    _sum("unconfirmed_demand_amount"),
        "overall_fill_rate":           overall_fr,
        "num_sales_orders":            _nunique("sales_order"),
        "num_key_accounts":            _nunique("key_account"),
        "num_products_impacted":       int(
            df.loc[unconf_mask, "product"].nunique() if "product" in df.columns else 0
        ),
        "num_unconfirmed_lines":       int(unconf_mask.sum()),
    }


# ---------------------------------------------------------------------------
# Report / summary builders
# ---------------------------------------------------------------------------

def build_unconfirmed_report(df: pd.DataFrame) -> pd.DataFrame:
    unconf_qty = df.get("unconfirmed_qty",           pd.Series(0,     index=df.index)).fillna(0)
    unc_amt    = df.get("unconfirmed_demand_amount",  pd.Series(0,     index=df.index)).fillna(0)
    fill_rate  = df.get("fill_rate",                  pd.Series(100.0, index=df.index)).fillna(100)
    mask = (unconf_qty > 0) | (unc_amt > 0) | (fill_rate < 100)
    return df[mask].reset_index(drop=True)


def _group_summary(
    df: pd.DataFrame,
    group_col: str,
    extra_counts: list[tuple[str, str]] | None = None,
) -> pd.DataFrame | None:
    if group_col not in df.columns:
        return None

    agg: dict[str, Any] = {}
    if "order_qty"    in df.columns: agg["Order Qty"]               = ("order_qty",    "sum")
    if "confirmed_qty" in df.columns: agg["Confirmed Qty"]           = ("confirmed_qty", "sum")
    if "unconfirmed_qty" in df.columns: agg["Unconfirmed Qty"]       = ("unconfirmed_qty", "sum")
    if "net_value"    in df.columns: agg["Net Value ($)"]           = ("net_value",    "sum")
    if "unconfirmed_demand_amount" in df.columns:
        agg["Unconfirmed Demand ($)"] = ("unconfirmed_demand_amount", "sum")

    if not agg:
        return None

    result = df.groupby(group_col).agg(**agg).reset_index()

    # Overall Fill Rate = confirmed / ordered
    if "Order Qty" in result.columns and "Confirmed Qty" in result.columns:
        result["Fill Rate (%)"] = (
            result["Confirmed Qty"] / result["Order Qty"].replace(0, np.nan) * 100
        ).round(2)

    # extra nunique counts (e.g. number of products, number of sales orders)
    if extra_counts:
        for count_col, label in extra_counts:
            if count_col in df.columns:
                counts = df.groupby(group_col)[count_col].nunique().reset_index()
                counts.columns = [group_col, label]
                result = result.merge(counts, on=group_col, how="left")

    result.rename(columns={group_col: group_col.replace("_", " ").title()}, inplace=True)
    return result


def build_account_summary(df: pd.DataFrame) -> pd.DataFrame | None:
    return _group_summary(
        df, "key_account",
        extra_counts=[("sales_order", "# Sales Orders"), ("product", "# Products")]
    )


def build_product_summary(df: pd.DataFrame) -> pd.DataFrame | None:
    return _group_summary(
        df, "product",
        extra_counts=[("sales_order", "# Sales Orders")]
    )


def build_plant_summary(df: pd.DataFrame) -> pd.DataFrame | None:
    return _group_summary(df, "plant")


def build_cdm_summary(df: pd.DataFrame) -> pd.DataFrame | None:
    return _group_summary(
        df, "cdm_name",
        extra_counts=[("key_account", "# Key Accounts"), ("sales_order", "# Sales Orders")]
    )


def build_top10(df: pd.DataFrame) -> dict[str, pd.DataFrame | None]:
    def _top(group_col: str, value_col: str, label: str) -> pd.DataFrame | None:
        if group_col not in df.columns or value_col not in df.columns:
            return None
        result = (
            df.groupby(group_col)[value_col]
            .sum()
            .reset_index()
            .sort_values(value_col, ascending=False)
            .head(10)
        )
        result.columns = [group_col.replace("_", " ").title(), label]
        return result

    return {
        "accounts_by_unc_amount":  _top("key_account",  "unconfirmed_demand_amount", "Unconfirmed Demand ($)"),
        "products_by_unc_amount":  _top("product",       "unconfirmed_demand_amount", "Unconfirmed Demand ($)"),
        "orders_by_unc_amount":    _top("sales_order",   "unconfirmed_demand_amount", "Unconfirmed Demand ($)"),
        "cdm_by_unc_amount":       _top("cdm_name",      "unconfirmed_demand_amount", "Unconfirmed Demand ($)"),
    }


# ---------------------------------------------------------------------------
# Excel report generator
# ---------------------------------------------------------------------------

_HEADER_FILL    = PatternFill("solid", fgColor="1A3C5E")
_KPI_FILL       = PatternFill("solid", fgColor="2E75B6")
_ALT_FILL       = PatternFill("solid", fgColor="D6E4F0")
_HEADER_FONT    = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT     = Font(bold=True, color="1A3C5E", size=14)
_KPI_LABEL_FONT = Font(bold=True, color="FFFFFF", size=10)
_KPI_VALUE_FONT = Font(bold=True, color="FFFFFF", size=13)
_THIN           = Side(style="thin", color="BFBFBF")
_THIN_BORDER    = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER         = Alignment(horizontal="center", vertical="center")
_LEFT           = Alignment(horizontal="left",   vertical="center")


def _autofit(ws, min_w: int = 8, max_w: int = 45) -> None:
    for col_cells in ws.columns:
        length = min_w
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = length


def _style_header_row(ws, row: int, num_cols: int) -> None:
    for c in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _THIN_BORDER


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    for c_idx, col in enumerate(df.columns, 1):
        ws.cell(row=start_row, column=c_idx, value=col)
    _style_header_row(ws, start_row, len(df.columns))
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    for r_idx, row_data in enumerate(
        dataframe_to_rows(df, index=False, header=False), start_row + 1
    ):
        for c_idx, val in enumerate(row_data, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border    = _THIN_BORDER
            cell.alignment = _LEFT
            if r_idx % 2 == 0:
                cell.fill = _ALT_FILL

    return start_row + len(df) + 2


def _apply_number_formats(ws, df: pd.DataFrame, start_row: int) -> None:
    for c_idx, col in enumerate(df.columns, 1):
        col_lower = col.lower()
        if any(k in col_lower for k in ["$", "value", "amount"]):
            fmt = '$#,##0.00'
        elif any(k in col_lower for k in ["rate", "pct", "%"]):
            fmt = '0.00"%"'
        elif any(k in col_lower for k in ["qty", "quantity", "orders", "accounts", "products", "lines"]):
            fmt = '#,##0'
        elif "date" in col_lower:
            fmt = 'YYYY-MM-DD'
        else:
            continue
        for r_idx in range(start_row + 1, start_row + len(df) + 1):
            ws.cell(row=r_idx, column=c_idx).number_format = fmt


def _apply_unconfirmed_cf(ws, df: pd.DataFrame, data_start_row: int) -> None:
    """Red fill on rows where Demand Status == Unconfirmed Demand."""
    if "Demand Status" not in df.columns:
        return
    col_idx    = list(df.columns).index("Demand Status") + 1
    col_letter = get_column_letter(col_idx)
    last_row   = data_start_row + len(df)
    last_col   = get_column_letter(len(df.columns))
    ws.conditional_formatting.add(
        f"A{data_start_row}:{last_col}{last_row}",
        FormulaRule(
            formula=[f'${col_letter}{data_start_row}="Unconfirmed Demand"'],
            fill=PatternFill("solid", fgColor="FFCCCC"),
        ),
    )


def _instructions_sheet(ws) -> None:
    ws.title = "Instructions"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 100
    lines = [
        ("Sales Order Fill Rate Dashboard — Excel Report", True,  16, "1A3C5E"),
        ("", False, 11, "000000"),
        ("Generated automatically from a SAP/BW Sales Order Fill Rate export.", False, 11, "000000"),
        ("", False, 11, "000000"),
        ("Sheets:", True, 12, "1A3C5E"),
        ("  • Instructions             — This page", False, 11, "000000"),
        ("  • Executive Summary        — KPI overview for management", False, 11, "000000"),
        ("  • Clean Data               — Cleaned and enriched dataset", False, 11, "000000"),
        ("  • Unconfirmed Demand Rpt   — Problem lines only", False, 11, "000000"),
        ("  • Key Account Summary      — Aggregated by key account", False, 11, "000000"),
        ("  • Product Summary          — Aggregated by product", False, 11, "000000"),
        ("  • Plant Summary            — Aggregated by plant", False, 11, "000000"),
        ("  • CDM Summary              — Aggregated by CDM / sales rep", False, 11, "000000"),
        ("  • Top 10 Issues            — Top offenders by unconfirmed demand", False, 11, "000000"),
        ("  • Raw Export Preview       — First 1000 rows of the original file", False, 11, "000000"),
        ("", False, 11, "000000"),
        ("Notes:", True, 12, "1A3C5E"),
        ("  • Fill Rate is shown as 0–100 (e.g. 95.5 = 95.5%).", False, 11, "000000"),
        ("  • Overall Fill Rate = Total Confirmed Qty / Total Order Qty.", False, 11, "000000"),
        ("  • Priority threshold is set at upload time.", False, 11, "000000"),
        ("  • Only actual data rows are written — no blank rows or formulas.", False, 11, "000000"),
    ]
    for r_idx, (text, bold, size, color) in enumerate(lines, 1):
        cell = ws.cell(row=r_idx, column=1, value=text)
        cell.font = Font(bold=bold, size=size, color=color)


def _exec_summary_sheet(ws, kpis: dict[str, Any], threshold: float) -> None:
    ws.title = "Executive Summary"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:D1")
    c = ws.cell(row=1, column=1, value="Sales Order Fill Rate — Executive Summary")
    c.font = _TITLE_FONT
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:D2")
    c2 = ws.cell(row=2, column=1, value=f"High Priority Threshold: ${threshold:,.2f}")
    c2.font      = Font(italic=True, color="595959")
    c2.alignment = _CENTER

    kpi_rows = [
        ("Total Order Quantity",          f"{kpis['total_order_qty']:,.0f}"),
        ("Total Confirmed Quantity",       f"{kpis['total_confirmed_qty']:,.0f}"),
        ("Total Unconfirmed Quantity",     f"{kpis['total_unconfirmed_qty']:,.0f}"),
        ("Total Net Value",               f"${kpis['total_net_value']:,.2f}"),
        ("Total Unconfirmed Demand ($)",  f"${kpis['total_unconfirmed_amount']:,.2f}"),
        ("Overall Fill Rate",             f"{kpis['overall_fill_rate']:.1f}%"),
        ("Sales Orders",                  f"{kpis['num_sales_orders']:,}"),
        ("Key Accounts",                  f"{kpis['num_key_accounts']:,}"),
        ("Products Impacted",             f"{kpis['num_products_impacted']:,}"),
        ("Lines with Unconfirmed Demand", f"{kpis['num_unconfirmed_lines']:,}"),
    ]
    row = 4
    for label, value in kpi_rows:
        lc = ws.cell(row=row, column=1, value=label)
        vc = ws.cell(row=row, column=2, value=value)
        lc.fill = vc.fill = _KPI_FILL
        lc.font = _KPI_LABEL_FONT; lc.alignment = _LEFT;   lc.border = _THIN_BORDER
        vc.font = _KPI_VALUE_FONT; vc.alignment = _CENTER; vc.border = _THIN_BORDER
        ws.row_dimensions[row].height = 22
        row += 1

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 24


def generate_excel_report(
    df_clean: pd.DataFrame,
    df_raw: pd.DataFrame,
    kpis: dict[str, Any],
    unconfirmed_df: pd.DataFrame,
    account_df: pd.DataFrame | None,
    product_df: pd.DataFrame | None,
    plant_df: pd.DataFrame | None,
    cdm_df: pd.DataFrame | None,
    top10: dict[str, pd.DataFrame | None],
    threshold: float,
) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)

    _instructions_sheet(wb.create_sheet("Instructions"))
    _exec_summary_sheet(wb.create_sheet("Executive Summary"), kpis, threshold)

    # Clean Data
    ws = wb.create_sheet("Clean Data")
    _write_df(ws, df_clean, start_row=1)
    _apply_number_formats(ws, df_clean, start_row=1)
    _apply_unconfirmed_cf(ws, df_clean, data_start_row=2)
    _autofit(ws)

    # Unconfirmed Demand Report
    ws = wb.create_sheet("Unconfirmed Demand Rpt")
    if not unconfirmed_df.empty:
        _write_df(ws, unconfirmed_df, start_row=1)
        _apply_number_formats(ws, unconfirmed_df, start_row=1)
        _autofit(ws)
    else:
        ws.cell(row=1, column=1, value="No unconfirmed demand found.")

    # Summary sheets
    for sheet_name, summary_df in [
        ("Key Account Summary", account_df),
        ("Product Summary",     product_df),
        ("Plant Summary",       plant_df),
        ("CDM Summary",         cdm_df),
    ]:
        ws = wb.create_sheet(sheet_name)
        if summary_df is not None and not summary_df.empty:
            _write_df(ws, summary_df, start_row=1)
            _apply_number_formats(ws, summary_df, start_row=1)
            _autofit(ws)
        else:
            ws.cell(row=1, column=1, value="No data available.")

    # Top 10 Issues
    ws = wb.create_sheet("Top 10 Issues")
    top_row = 1
    for title, t_df in [
        ("Top 10 Key Accounts by Unconfirmed Demand",  top10.get("accounts_by_unc_amount")),
        ("Top 10 Products by Unconfirmed Demand",      top10.get("products_by_unc_amount")),
        ("Top 10 Sales Orders by Unconfirmed Demand",  top10.get("orders_by_unc_amount")),
        ("Top 10 CDMs by Unconfirmed Demand",          top10.get("cdm_by_unc_amount")),
    ]:
        ws.cell(row=top_row, column=1, value=title).font = Font(bold=True, size=12, color="1A3C5E")
        top_row += 1
        if t_df is not None and not t_df.empty:
            top_row = _write_df(ws, t_df, start_row=top_row)
        else:
            ws.cell(row=top_row, column=1, value="No data available.")
            top_row += 2
    _autofit(ws)

    # Raw Export Preview (first 1000 rows only)
    ws = wb.create_sheet("Raw Export Preview")
    _write_df(ws, df_raw.head(1000), start_row=1)
    _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
