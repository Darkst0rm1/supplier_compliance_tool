"""Excel export for the Batch Quality Analysis dashboard.

One workbook, four sheets: Flagged Issues, Related Records, Multiple Batches,
AI Review. Bold headers, auto-filter, frozen top row, date formatting, sensible
column widths, and id columns preserved as text. No corrected batch / expiry
fields are ever exported.
"""
from __future__ import annotations

import io
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .loader import (
    COL_BATCH,
    COL_BATCH_SLED,
    COL_MATERIAL,
    COL_PLANT,
    COL_PO,
    COL_RECEIVED,
    COL_SUPPLIER,
    COL_VENDOR,
)

_HEADER_FILL = PatternFill("solid", fgColor="FFF2F2F2")
_HEADER_FONT = Font(name="Calibri", size=11, bold=True)
_DATE_FMT = "yyyy-mm-dd"

# Columns that should render and round-trip as text (ids), keyed by header name.
_TEXT_HEADERS = {COL_VENDOR, COL_SUPPLIER, COL_PO, COL_PLANT, COL_MATERIAL, COL_BATCH}
_DATE_HEADERS = {COL_BATCH_SLED, COL_RECEIVED, "Earliest Expiry Date", "Latest Expiry Date"}


def _write_sheet(ws, df: pd.DataFrame) -> None:
    if df is None or df.empty:
        ws.cell(row=1, column=1, value="No data.")
        return

    headers = list(df.columns)
    for c_idx, name in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=str(name))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    date_idx = {headers.index(h) + 1 for h in _DATE_HEADERS if h in headers}
    text_idx = {headers.index(h) + 1 for h in _TEXT_HEADERS if h in headers}

    for r_off, (_, row) in enumerate(df.iterrows()):
        r_idx = r_off + 2
        for c_idx, col in enumerate(headers, 1):
            val = row[col]
            if pd.isna(val):
                val = None
            if c_idx in text_idx and val is not None:
                val = str(val)
            elif isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if c_idx in date_idx and val is not None and not (c_idx in text_idx):
                cell.number_format = _DATE_FMT
            elif c_idx in text_idx:
                cell.number_format = "@"

    n_rows = len(df) + 1
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{n_rows}"
    ws.freeze_panes = "A2"

    for c_idx, col in enumerate(headers, 1):
        max_len = len(str(col))
        for val in df[col].head(200):
            max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len + 2, 10), 50)


def generate_excel(
    flagged: pd.DataFrame,
    related: pd.DataFrame,
    multi_batch: pd.DataFrame,
    ai_review: Optional[pd.DataFrame] = None,
) -> bytes:
    """Render the four-sheet Batch Quality Analysis workbook (bytes)."""
    wb = Workbook()
    wb.remove(wb.active)
    _write_sheet(wb.create_sheet("Flagged Issues"), flagged)
    _write_sheet(wb.create_sheet("Related Records"), related)
    _write_sheet(wb.create_sheet("Multiple Batches"), multi_batch)
    _write_sheet(wb.create_sheet("AI Review"), ai_review if ai_review is not None else pd.DataFrame())
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
