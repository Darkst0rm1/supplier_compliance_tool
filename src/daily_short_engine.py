"""Processing engine for the SAPUI5 Daily Short Report export.

The export is order-line level. Each line carries an ordered quantity and the
quantities that made it through each fulfilment stage:

    Order Quantity -> Confirmed -> Total Delivery -> Picked -> Invoice

The workbook embeds a TEMPLATE at the bottom (below the data) describing the
report. This engine implements that template:

* Summary totals + fill-rate ratios (Confirmed/Order, Delivery/Order, Invoice/Order)
* Three "shorted" analyses, each Order-based:
    - Unconfirmed quantities            = Order - Confirmed
    - Shorted at outbound delivery       = Order - Delivered
    - Shorted at invoicing               = Order - Invoiced
* Each analysis is a table with the template columns:
    Plant | Sales order # | Customer | Material | Material description |
    Ordered | Confirmed | Shorted | Reason
"""
from __future__ import annotations

import io
from typing import Any

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ---------------------------------------------------------------------------
# Source columns (exact export headers) + tolerant aliases
# ---------------------------------------------------------------------------
COL_SALES_ORDER   = "Sales Order"
COL_MATERIAL      = "Material"
COL_ORDER_QTY     = "Order Quantity"
COL_CONFIRMED_QTY = "Confirmed Quantity (CS)"
COL_DELIVERY_QTY  = "Total Delivery (Sales) Quantity (CS)"
COL_PICKED_QTY    = "Picked Quantity (CS)"
COL_INVOICE_QTY   = "Invoice Quantity (CS)"
COL_PLANT         = "Plant"
COL_CUSTOMER      = "Sold To Name"
COL_MAT_DESC      = "TOL Material Description"
COL_REASON        = "Reason for Rejection Desc."
COL_ODS           = "Outbound Delivery Status"
COL_ORDER_TYPE    = "Sales Order Type Desc."
COL_NET_AMT       = "Item Net Amount"
COL_NET_AMT_CONF  = "Item Net Amount (Confirmed)"
COL_VENDOR_NAME   = "Vendor Name"
COL_CDM           = "CDM Name"
COL_REQ_DATE      = "Requested Delivery Date"
COL_SHIP_DATE     = "Shipped Date"

_QTY_COLS = [COL_ORDER_QTY, COL_CONFIRMED_QTY, COL_DELIVERY_QTY, COL_PICKED_QTY, COL_INVOICE_QTY]

# Header aliases -> canonical, applied case-insensitively after stripping.
_ALIASES = {
    "sold to name": COL_CUSTOMER,
    "customer": COL_CUSTOMER,
    "tol material description": COL_MAT_DESC,
    "material description": COL_MAT_DESC,
    "total delivery (sales) quantity (cs)": COL_DELIVERY_QTY,
    "delivered quantity (cs)": COL_DELIVERY_QTY,
    "reason for rejection desc.": COL_REASON,
    "reason for rejection": COL_REASON,
}


# ---------------------------------------------------------------------------
# Short analyses — measured stage-to-stage (each step's gap):
#   unconfirmed : Ordered    - Confirmed   (ordered but not confirmed)
#   delivery    : Confirmed  - Delivered   (confirmed but no outbound delivery)
#   invoice     : Delivered  - Invoiced    (delivered but not invoiced)
# ---------------------------------------------------------------------------
SHORT_DEFS: list[dict[str, str]] = [
    {"key": "unconfirmed", "label": "Unconfirmed Quantities",
     "title": "Ordered but not confirmed",
     "base_col": COL_ORDER_QTY, "base_label": "Ordered",
     "fulfilled_col": COL_CONFIRMED_QTY, "fulfilled_label": "Confirmed"},
    {"key": "delivery", "label": "Confirmed, No Outbound Delivery",
     "title": "Confirmed but no outbound delivery created",
     "base_col": COL_CONFIRMED_QTY, "base_label": "Confirmed",
     "fulfilled_col": COL_DELIVERY_QTY, "fulfilled_label": "Delivered"},
    {"key": "invoice", "label": "Delivered, Not Invoiced",
     "title": "Outbound delivery created but not invoiced",
     "base_col": COL_DELIVERY_QTY, "base_label": "Delivered",
     "fulfilled_col": COL_INVOICE_QTY, "fulfilled_label": "Invoiced"},
]


def _def(key: str) -> dict[str, str]:
    return next(d for d in SHORT_DEFS if d["key"] == key)


def _short_col(key: str) -> str:
    return f"_short_{key}"


# ---------------------------------------------------------------------------
# Loader
# ---------------------------------------------------------------------------
def _normalize_header(name: Any) -> str:
    return str(name).strip()


def _apply_aliases(df: pd.DataFrame) -> pd.DataFrame:
    rename = {}
    for col in df.columns:
        key = str(col).strip().lower()
        if key in _ALIASES and _ALIASES[key] != col:
            rename[col] = _ALIASES[key]
    return df.rename(columns=rename) if rename else df


def load_daily_short(file_obj: Any) -> pd.DataFrame:
    """Read the export, drop the embedded template/footer rows, coerce types,
    and compute the three Order-based short columns. Returns the clean lines."""
    file_obj.seek(0)
    df = pd.read_excel(file_obj, engine="openpyxl")
    df.columns = [_normalize_header(c) for c in df.columns]
    df = _apply_aliases(df)

    missing = [c for c in [COL_SALES_ORDER, COL_ORDER_QTY] if c not in df.columns]
    if missing:
        raise DailyShortError(
            "This doesn't look like a Daily Short export — missing column(s): "
            + ", ".join(missing)
        )

    # Coerce quantity columns to numeric (footer/template rows carry text and
    # become NaN, which lets us drop them cleanly).
    for c in _QTY_COLS:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # Keep only real order lines: a Sales Order present and a numeric order qty.
    df = df[df[COL_SALES_ORDER].notna() & df[COL_ORDER_QTY].notna()].copy()
    df = df.reset_index(drop=True)

    # Normalise text id columns to clean strings.
    for c in [COL_SALES_ORDER, COL_MATERIAL, COL_PLANT]:
        if c in df.columns:
            df[c] = (
                df[c].astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.strip()
            )

    # Dates
    for c in [COL_REQ_DATE, COL_SHIP_DATE]:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    # Stage-to-stage short columns (clip at 0 — never count a gain as a short).
    for d in SHORT_DEFS:
        base = df[d["base_col"]].fillna(0) if d["base_col"] in df.columns else 0
        fulfilled = df[d["fulfilled_col"]].fillna(0) if d["fulfilled_col"] in df.columns else 0
        df[_short_col(d["key"])] = (base - fulfilled).clip(lower=0)

    return df


class DailyShortError(Exception):
    """Raised when the uploaded file isn't a usable Daily Short export."""


# ---------------------------------------------------------------------------
# Summary / KPIs
# ---------------------------------------------------------------------------
def build_kpis(df: pd.DataFrame) -> dict[str, Any]:
    def _sum(col: str) -> float:
        return float(df[col].sum()) if col in df.columns else 0.0

    ordered   = _sum(COL_ORDER_QTY)
    confirmed = _sum(COL_CONFIRMED_QTY)
    delivered = _sum(COL_DELIVERY_QTY)
    invoiced  = _sum(COL_INVOICE_QTY)

    def _rate(num: float, den: float) -> float:
        return (num / den * 100) if den else 0.0

    def _short(key: str) -> float:
        c = _short_col(key)
        return float(df[c].sum()) if c in df.columns else 0.0

    def _lines(key: str) -> int:
        c = _short_col(key)
        return int((df[c] > 0).sum()) if c in df.columns else 0

    return {
        "lines": int(len(df)),
        "ordered": ordered,
        "confirmed": confirmed,
        "delivered": delivered,
        "invoiced": invoiced,
        # Confirmation: Confirmed / Ordered
        "confirm_rate": _rate(confirmed, ordered),
        # Outbound delivery: vs Ordered, and vs Confirmed
        "delivery_vs_order": _rate(delivered, ordered),
        "delivery_vs_confirmed": _rate(delivered, confirmed),
        # Invoicing: vs Ordered (total invoice), and vs Delivered
        "invoice_vs_order": _rate(invoiced, ordered),
        "invoice_vs_delivered": _rate(invoiced, delivered),
        "short_unconfirmed": _short("unconfirmed"),
        "short_delivery": _short("delivery"),
        "short_invoice": _short("invoice"),
        "lines_unconfirmed": _lines("unconfirmed"),
        "lines_delivery": _lines("delivery"),
        "lines_invoice": _lines("invoice"),
    }


# ---------------------------------------------------------------------------
# Short tables (template columns)
# ---------------------------------------------------------------------------
def template_columns(key: str) -> list[str]:
    """Column order for a stage's table: identity cols, Ordered for context,
    the stage's base + fulfilled quantities, then Shorted and Reason."""
    d = _def(key)
    cols = ["Plant", "Sales order #", "Customer", "Material", "Material description", "Ordered"]
    if d["base_label"] != "Ordered":
        cols.append(d["base_label"])
    cols += [d["fulfilled_label"], "Shorted", "Reason"]
    return cols


def build_short_table(df: pd.DataFrame, key: str, top_n: int | None = None) -> pd.DataFrame:
    """Table for one stage: only lines short at that stage (base > fulfilled),
    sorted by largest short. ``top_n=None`` returns all."""
    d = _def(key)
    scol = _short_col(key)
    if scol not in df.columns:
        return pd.DataFrame(columns=template_columns(key))

    sub = df[df[scol] > 0].copy().sort_values(scol, ascending=False)
    if top_n is not None:
        sub = sub.head(top_n)

    data: dict[str, Any] = {
        "Plant":               sub.get(COL_PLANT),
        "Sales order #":       sub.get(COL_SALES_ORDER),
        "Customer":            sub.get(COL_CUSTOMER),
        "Material":            sub.get(COL_MATERIAL),
        "Material description": sub.get(COL_MAT_DESC),
        "Ordered":             sub.get(COL_ORDER_QTY),
    }
    if d["base_label"] != "Ordered":
        data[d["base_label"]] = sub.get(d["base_col"])
    data[d["fulfilled_label"]] = sub.get(d["fulfilled_col"])
    data["Shorted"] = sub[scol]
    data["Reason"] = sub.get(COL_REASON)
    return pd.DataFrame(data).reset_index(drop=True)


def build_group_summary(df: pd.DataFrame, group_col: str, key: str) -> pd.DataFrame | None:
    """Roll a short analysis up by a grouping column (Customer, Vendor, …)."""
    scol = _short_col(key)
    if group_col not in df.columns or scol not in df.columns:
        return None
    g = (
        df.groupby(group_col)
        .agg(Ordered=(COL_ORDER_QTY, "sum"), Shorted=(scol, "sum"),
             Lines=(scol, lambda s: int((s > 0).sum())))
        .reset_index()
    )
    g = g[g["Shorted"] > 0].sort_values("Shorted", ascending=False)
    g.rename(columns={group_col: group_col}, inplace=True)
    return g.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel report
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_KPI_FILL    = PatternFill("solid", fgColor="2E75B6")
_ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT  = Font(bold=True, color="1F4E79", size=14)
_KPI_LABEL   = Font(bold=True, color="FFFFFF", size=10)
_KPI_VALUE   = Font(bold=True, color="FFFFFF", size=12)
_THIN        = Side(style="thin", color="BFBFBF")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_CENTER      = Alignment(horizontal="center", vertical="center")
_LEFT        = Alignment(horizontal="left", vertical="center")


def _autofit(ws, min_w=8, max_w=48):
    for col_cells in ws.columns:
        length = min_w
        letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value is not None:
                length = max(length, min(max_w, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = length


def _write_df(ws, df: pd.DataFrame, start_row: int = 1) -> int:
    for c_idx, col in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c_idx, value=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _CENTER
        cell.border = _BORDER
    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.border = _BORDER
            cell.alignment = _LEFT
            if r_idx % 2 == 0:
                cell.fill = _ALT_FILL
    return start_row + len(df) + 2


def _summary_sheet(ws, kpis: dict[str, Any]):
    ws.title = "Summary"
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:D1")
    t = ws.cell(row=1, column=1, value="Daily Short Report — Summary")
    t.font = _TITLE_FONT
    t.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    def _pct(v: float) -> str:
        return f"{v:.2f}%"

    # Stage table — percentages sit right next to each quantity line.
    stage_df = pd.DataFrame([
        {"Stage": "Ordered",   "Quantity": round(kpis["ordered"]),
         "% of Ordered": "100.00%", "% of Prior Stage": "—"},
        {"Stage": "Confirmed", "Quantity": round(kpis["confirmed"]),
         "% of Ordered": _pct(kpis["confirm_rate"]),
         "% of Prior Stage": f"{_pct(kpis['confirm_rate'])} of Ordered"},
        {"Stage": "Delivered", "Quantity": round(kpis["delivered"]),
         "% of Ordered": _pct(kpis["delivery_vs_order"]),
         "% of Prior Stage": f"{_pct(kpis['delivery_vs_confirmed'])} of Confirmed"},
        {"Stage": "Invoiced",  "Quantity": round(kpis["invoiced"]),
         "% of Ordered": _pct(kpis["invoice_vs_order"]),
         "% of Prior Stage": f"{_pct(kpis['invoice_vs_delivered'])} of Delivered"},
    ])
    next_row = _write_df(ws, stage_df, start_row=3)
    for r in range(4, 4 + len(stage_df)):       # format the Quantity column
        ws.cell(row=r, column=2).number_format = "#,##0"

    # Gap / line-count KPIs below the stage table.
    kpi_rows = [
        ("Order Lines",                   f"{kpis['lines']:,}"),
        ("Ordered − Confirmed",           f"{kpis['short_unconfirmed']:,.0f} qty · {kpis['lines_unconfirmed']:,} lines"),
        ("Confirmed − Delivered",         f"{kpis['short_delivery']:,.0f} qty · {kpis['lines_delivery']:,} lines"),
        ("Delivered − Invoiced",          f"{kpis['short_invoice']:,.0f} qty · {kpis['lines_invoice']:,} lines"),
    ]
    r = next_row
    for label, value in kpi_rows:
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        lc = ws.cell(row=r, column=1, value=label)
        vc = ws.cell(row=r, column=2, value=value)
        lc.fill = vc.fill = _KPI_FILL
        lc.font = _KPI_LABEL; lc.alignment = _LEFT; lc.border = _BORDER
        vc.font = _KPI_VALUE; vc.alignment = _LEFT; vc.border = _BORDER
        ws.row_dimensions[r].height = 20
        r += 1

    _autofit(ws)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 22


def generate_excel_report(df: pd.DataFrame, kpis: dict[str, Any], top_n: int | None = None) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    _summary_sheet(wb.create_sheet("Summary"), kpis)

    sheet_names = {
        "unconfirmed": "Top Unconfirmed",
        "delivery": "Top Shorted at Delivery",
        "invoice": "Top Shorted at Invoicing",
    }
    for d in SHORT_DEFS:
        ws = wb.create_sheet(sheet_names[d["key"]])
        table = build_short_table(df, d["key"], top_n=top_n)
        title = d["title"] + (f" — Top {top_n}" if top_n else " — all")
        ws.cell(row=1, column=1, value=title).font = _TITLE_FONT
        if table.empty:
            ws.cell(row=2, column=1, value="No shorts found.")
        else:
            _write_df(ws, table, start_row=2)
        _autofit(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
