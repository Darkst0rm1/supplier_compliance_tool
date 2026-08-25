"""Processing engine for the EWM Dispose list.

Starts from the **EWM stock export** — the warehouse's own bin-level view of
what is sitting where — keeps its full column set, and adds only the Brand
Manager. One sheet per plant.

Inputs:

* **EWM dispose exports**, one per plant (2910 / 2920 / 2930), file names like
  ``Mo - EWM 2910 dispose.xlsx``. Each carries one row per Storage Bin /
  Product / Batch with 54 columns of warehouse detail.
* **Last Sell / BDM Material Master** — used only to look up the Brand Manager.
* **Materials-based dispose export**, optional, for plants 2925 and 2935 —
  which have no EWM system and so can never produce an EWM export. Same
  15-column shape ``donate_dispose_engine`` outputs; BDM is already populated
  in it, so it needs no master lookup here. One file can carry both plants,
  split by its ``Plant`` column into two sheets.

The output is the EWM export itself with a single column added: ``BDM``, placed
straight after ``Batch``, plus one sheet per materials-based plant supplied.
Sheets are named by plant code, matching the supplied finished workbook
``Dispose list 0722 EWM (1).xlsx``.

Selection rule
--------------
One rule, and it is the same packaging exclusion the Overstock and Donate /
Dispose engines already apply:

* The ``Product`` number does NOT start with "40" (display / shipper / label /
  sticker / sample packaging — never sellable stock).

Everything else in the export is kept, in its original row order.

Why only one rule — read this before adding another
---------------------------------------------------
The finished workbook was cleaned by hand, differently on each sheet, so it
does **not** describe a single consistent rule set. Measured against it:

* **2910** applied the "40" exclusion (15 rows) but *kept* a
  ``9020 / SHIPPING / F2`` row sitting on an outbound delivery (Document
  Category ``PDO``).
* **2920** applied the "40" exclusion (10 rows) *and* removed three
  ``9020 / SHIPPING / F2 / PDO`` rows — indistinguishable from the one 2910
  kept.
* **2930** applied no exclusion at all: it still contains a "40" product
  (``40042627``, SLED 2023).

So reproducing that workbook exactly would mean encoding three mutually
inconsistent behaviours. This engine applies the one derivable rule uniformly.
Measured against the workbook it differs on 5 rows — 3 shipping-allocated rows
in 2920 that it keeps, 1 "40" product in 2930 that it drops, and 1 row in the
2910 sheet absent from the EWM export supplied with it — and on 99.75% of the
cells of every shared row it is identical. The remaining cell differences are
data vintage: 26 BDM values the workbook took from an older master (those brand
managers still exist but are assigned to other products now), and 2 rows whose
quantities are proportionally higher in the workbook than in the export.

If the shipping-allocated rows should really be dropped, that is a genuine new
rule — add it here deliberately rather than inferring it from one sheet.

Materials-based dispose date filter (2925 / 2935 only)
--------------------------------------------------------
The materials-based export for 2925/2935 is a **raw** SAP Materials inventory
extract (same shape ``overstock_engine``/``donate_dispose_engine`` read) — it
carries Shelf Life Expiration Date but not a pre-computed BDM, Last Sell Day,
or Last sell date. Those get filled in from the Last Sell / BDM master already
uploaded for the EWM plants' BDM lookup (``load_master``, ``Last Sell Day``
column), the same way ``donate_dispose_engine`` computes
``Last sell date = SLED − Last Sell Day``. If the file already carries any of
those three values (the older, pre-filtered shape), the existing value wins —
nothing here overwrites data that's already there.

The user picks one date on the page. A row is kept only if **both** Shelf
Life Expiration Date and Last sell date fall on/before that date +
``SLED_CUTOFF_OFFSET_DAYS`` — the same grace window
``donate_dispose_engine`` uses, applied here rather than left as a second
user-tweakable field, per this app's rule that anything affecting what gets
flagged is a hardcoded, auditable business rule, not a setting (see
``HANDOVER.md`` §4). A row missing either date (no master supplied, product
not in the master, or no SLED on the row) cannot be evaluated and is dropped.
Rows with no cutoff supplied at all (``materials_cutoff=None``) are not
date-filtered — only the packaging exclusion applies.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


class DisposeEwmError(Exception):
    """Raised when an uploaded file isn't a usable EWM Dispose source export."""


# ---------------------------------------------------------------------------
# Source columns (exact export headers)
# ---------------------------------------------------------------------------
EWM_PRODUCT = "Product"
EWM_BATCH   = "Batch"
EWM_BIN     = "Storage Bin"
EWM_OWNER   = "Party Entitled to Dispose"   # "BP2910" — lets a file name its plant

MASTER_PRODUCT       = "Product Number"
MASTER_BDM_NAME       = "Brand Manager Name"
MASTER_LAST_SELL_DAY  = "Last Sell Day"

# Source columns for the Materials-based dispose export — used for plants 2925
# and 2935, which have no EWM system of their own and so can never produce an
# EWM dispose export. Same 15-column shape the Donate/Dispose engine outputs
# (see ``donate_dispose_engine.OUTPUT_COLUMNS``), because that is where this
# extract comes from: someone runs that report filtered to 2925/2935 and it is
# handed to this tool as-is, BDM already looked up.
MATDISP_MATERIAL        = "Material"
MATDISP_PLANT            = "Plant"
MATDISP_UNRESTRICTED     = "Unrestricted Stock"
MATDISP_QUALITY          = "Stock in Quality Inspection"
MATDISP_BLOCKED          = "Blocked Stock"
MATDISP_SLED             = "Shelf Life Expiration Date"
MATDISP_LAST_SELL_DAY    = "Last sell day"
MATDISP_LAST_SELL_DATE   = "Last sell date"
MATDISP_STOCK_COLS = [MATDISP_UNRESTRICTED, MATDISP_QUALITY, MATDISP_BLOCKED]
MATDISP_DATE_COLS = [MATDISP_SLED, MATDISP_LAST_SELL_DATE]

# ---------------------------------------------------------------------------
# Output
# ---------------------------------------------------------------------------
OUT_BDM = "BDM"

# BDM goes straight after Batch. The finished workbook put it there on 2910 and
# 2930 but one column earlier on 2920; that inconsistency is manual, so the
# majority position is applied to every sheet.
BDM_INSERT_AFTER = EWM_BATCH

# Plants with an EWM dispose export, and the sheet order of the output.
PLANTS: list[str] = ["2910", "2920", "2930"]

# Plants sourced from the Materials-based export instead (no EWM system).
MATERIALS_PLANTS: list[str] = ["2925", "2935"]

# ---------------------------------------------------------------------------
# Business-rule constants (auditable; change here, not in the UI)
# ---------------------------------------------------------------------------
# A Product whose number starts with any of these is packaging / display /
# promo material, never sellable stock — excluded entirely.
EXCLUDED_PRODUCT_PREFIXES = ("40",)

# Materials-based dispose date filter (2925/2935 only): a row is kept when
# both Shelf Life Expiration Date and Last sell date fall on/before the
# user-picked date plus this many days. Matches donate_dispose_engine's
# SLED_CUTOFF_OFFSET_DAYS grace window.
SLED_CUTOFF_OFFSET_DAYS = 4

# ---------------------------------------------------------------------------
# Column typing. Everything not listed stays text, so ids keep their leading
# zeros (Consolidation Group "0005000242" must not become a number).
# ---------------------------------------------------------------------------
DATE_COLS = [
    "Shelf Life Expiration Date",
    "Goods Receipt Date",
    "Latest Delivery Date",
    "Last sell date",
]
TIME_COLS = ["Goods Receipt Time"]
INT_COLS = ["Quantity", "Packed Qty (AUoM)"]
DECIMAL_COLS = [
    "Overall Valuation Quantity",
    "Loading Weight",
    "Loading Volume",
    "Capacity Consumption",
]


# ---------------------------------------------------------------------------
# Importers
# ---------------------------------------------------------------------------
def _read_excel_str(file_obj: Any) -> pd.DataFrame:
    file_obj.seek(0)
    df = pd.read_excel(file_obj, dtype=str, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _require(df: pd.DataFrame, cols: list[str], what: str) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise DisposeEwmError(
            f"This doesn't look like the {what} export — missing column(s): "
            + ", ".join(missing)
        )


def _clean_text(s: pd.Series) -> pd.Series:
    r"""Blank-safe text: NA becomes "", float artifacts are dropped, padding is
    trimmed.

    ``.str.strip()`` rather than a ``\s+`` regex, and ``fillna`` rather than
    ``astype(str)``: on pandas' string dtype the regex runs on PyArrow's RE2
    engine (whose ``\s`` is ASCII-only and misses non-breaking spaces) and
    ``astype(str)`` leaves NA as NA instead of the string "nan".
    """
    return (
        s.fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .replace({"nan": "", "None": "", "NaT": ""})
    )


def load_ewm(file_obj: Any, expected_plant: str | None = None) -> pd.DataFrame:
    """Read one plant's EWM dispose export, keeping every column.

    Pass ``expected_plant`` to reject a file uploaded into the wrong plant's box
    — each export names its own plant in ``Party Entitled to Dispose``
    ("BP2910").
    """
    df = _read_excel_str(file_obj)
    _require(df, [EWM_PRODUCT, EWM_BATCH, EWM_BIN],
             f"EWM dispose ({expected_plant})" if expected_plant else "EWM dispose")

    if expected_plant and EWM_OWNER in df.columns:
        owners = (
            df[EWM_OWNER].dropna().astype(str).str.strip()
            .str.upper().str.removeprefix("BP")
        )
        found = sorted({o for o in owners if o})
        if found and expected_plant not in found:
            raise DisposeEwmError(
                f"This file is the plant {', '.join(found)} EWM dispose export, "
                f"but it was uploaded in the {expected_plant} box. Swap the "
                "files and try again."
            )

    for c in DATE_COLS:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")
    for c in INT_COLS + DECIMAL_COLS:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def load_materials_dispose(file_obj: Any) -> pd.DataFrame:
    """Read the Materials-based dispose export covering plants 2925 and 2935.

    Usually a **raw** SAP Materials inventory export — no BDM, Last sell day,
    or Last sell date columns at all; ``build_dispose_ewm`` fills those in from
    the Last Sell / BDM master. If the file already carries any of those three
    (the older, pre-filtered shape produced by a Donate/Dispose-style run),
    its values are kept as-is. One file can carry rows for both plants;
    ``build_dispose_ewm`` splits it by ``Plant``.
    """
    df = _read_excel_str(file_obj)
    _require(df, [MATDISP_MATERIAL, MATDISP_PLANT] + MATDISP_STOCK_COLS,
             "Materials dispose (2925/2935)")

    for c in (OUT_BDM, MATDISP_LAST_SELL_DAY, MATDISP_LAST_SELL_DATE):
        if c not in df.columns:
            df[c] = pd.NA

    for c in MATDISP_DATE_COLS:
        df[c] = pd.to_datetime(df[c], errors="coerce")
    df[MATDISP_LAST_SELL_DAY] = pd.to_numeric(df[MATDISP_LAST_SELL_DAY], errors="coerce")
    for c in MATDISP_STOCK_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce")
    return df


def load_master(file_obj: Any) -> pd.DataFrame:
    """Read the Last Sell / BDM master and reduce it to one row per product,
    columns ``bdm`` and ``last_sell_day``.

    The master repeats a product across vendors; values are the same on each,
    so the first row wins. ``last_sell_day`` is optional — if the master has
    no ``Last Sell Day`` column the result is all-NaN, and the materials-based
    2925/2935 rows that need it to compute Last sell date can't be dated."""
    df = _read_excel_str(file_obj)
    _require(df, [MASTER_PRODUCT, MASTER_BDM_NAME],
             "Last Sell / BDM Material Master")

    last_sell_day = (
        pd.to_numeric(df[MASTER_LAST_SELL_DAY], errors="coerce")
        if MASTER_LAST_SELL_DAY in df.columns
        else pd.Series(pd.NA, index=df.index)
    )

    lut = pd.DataFrame({
        "product": _clean_text(df[MASTER_PRODUCT]),
        "bdm": _clean_text(df[MASTER_BDM_NAME]),
        "last_sell_day": last_sell_day.to_numpy(),
    })
    lut = lut[lut["product"] != ""]
    lut = lut.drop_duplicates(subset="product", keep="first")
    return lut.set_index("product")[["bdm", "last_sell_day"]]


# ---------------------------------------------------------------------------
# Core build
# ---------------------------------------------------------------------------
def build_dispose_ewm(
    ewm_by_plant: dict[str, pd.DataFrame],
    bdm_lut: pd.DataFrame | None = None,
    materials_dispose: pd.DataFrame | None = None,
    materials_cutoff: date | datetime | None = None,
) -> dict[str, pd.DataFrame]:
    """Apply the packaging exclusion and add the BDM column.

    ``ewm_by_plant`` maps a plant code to that plant's ``load_ewm`` frame; only
    the plants supplied get a sheet. ``bdm_lut`` is ``load_master``'s frame
    (columns ``bdm``, ``last_sell_day``, indexed by product); without it the
    BDM column is present but empty.

    ``materials_dispose`` is ``load_materials_dispose``'s frame for plants 2925
    and 2935 (no EWM system); it is split by ``Plant`` into one sheet each,
    always both, even if a plant has no rows after exclusion. Missing BDM /
    Last sell day / Last sell date values on those rows are filled from
    ``bdm_lut``; existing values are kept as-is. If ``materials_cutoff`` is
    given, a materials-dispose row is kept only when both Shelf Life
    Expiration Date and Last sell date fall on/before
    ``materials_cutoff + SLED_CUTOFF_OFFSET_DAYS`` — rows missing either date
    are dropped. Leave ``materials_cutoff`` as ``None`` to skip date filtering
    entirely (only the packaging exclusion applies).

    Row order is the export's own — the warehouse's ordering is meaningful and
    nothing here re-sorts it."""
    if not ewm_by_plant and materials_dispose is None:
        raise DisposeEwmError(
            "Upload at least one EWM dispose export to build the list."
        )

    sheets: dict[str, pd.DataFrame] = {}
    for plant in PLANTS:
        df = ewm_by_plant.get(plant)
        if df is None:
            continue
        out = df.copy()

        is_packaging = _clean_text(out[EWM_PRODUCT]).str.startswith(
            EXCLUDED_PRODUCT_PREFIXES
        )
        out = out[~is_packaging].reset_index(drop=True)

        bdm = _clean_text(out[EWM_PRODUCT]).map(bdm_lut["bdm"]) if bdm_lut is not None \
            else pd.Series([None] * len(out), index=out.index, dtype=object)
        out.insert(out.columns.get_loc(BDM_INSERT_AFTER) + 1, OUT_BDM, bdm)

        sheets[plant] = out

    if materials_dispose is not None:
        md = materials_dispose.copy()
        is_packaging = _clean_text(md[MATDISP_MATERIAL]).str.startswith(
            EXCLUDED_PRODUCT_PREFIXES
        )
        md = md[~is_packaging].reset_index(drop=True)

        if bdm_lut is not None:
            products = _clean_text(md[MATDISP_MATERIAL])

            blank_bdm = _clean_text(md[OUT_BDM]) == ""
            md[OUT_BDM] = md[OUT_BDM].where(~blank_bdm, products.map(bdm_lut["bdm"]))

            md[MATDISP_LAST_SELL_DAY] = md[MATDISP_LAST_SELL_DAY].where(
                md[MATDISP_LAST_SELL_DAY].notna(),
                products.map(bdm_lut["last_sell_day"]),
            )

        needs_last_sell_date = (
            md[MATDISP_LAST_SELL_DATE].isna()
            & md[MATDISP_SLED].notna()
            & md[MATDISP_LAST_SELL_DAY].notna()
        )
        computed_last_sell_date = md[MATDISP_SLED] - pd.to_timedelta(
            md[MATDISP_LAST_SELL_DAY], unit="D"
        )
        md[MATDISP_LAST_SELL_DATE] = md[MATDISP_LAST_SELL_DATE].where(
            ~needs_last_sell_date, computed_last_sell_date
        )

        if materials_cutoff is not None:
            cutoff_ts = pd.Timestamp(materials_cutoff) + pd.Timedelta(
                days=SLED_CUTOFF_OFFSET_DAYS
            )
            keep = (
                md[MATDISP_SLED].notna()
                & md[MATDISP_LAST_SELL_DATE].notna()
                & (md[MATDISP_SLED] <= cutoff_ts)
                & (md[MATDISP_LAST_SELL_DATE] <= cutoff_ts)
            )
            md = md[keep].reset_index(drop=True)

        plant_col = _clean_text(md[MATDISP_PLANT])
        for plant in MATERIALS_PLANTS:
            sheets[plant] = md[plant_col == plant].reset_index(drop=True)

    return sheets


# ---------------------------------------------------------------------------
# Excel export (matches the finished workbook: Arial 10, plain header)
# ---------------------------------------------------------------------------
_FONT = Font(name="Arial", size=10)
_DATE_FMT = "mm-dd-yy"
_TIME_FMT = r"[$-F400]h:mm:ss\ AM/PM"
_INT_FMT = "#,##0"
_DECIMAL_FMT = "#,##0.000"

# Column widths copied from the finished workbook, keyed by header.
_COL_WIDTHS = {
    "Storage Type": 14.0, "Storage Bin": 13.0, "Logical Position": 18.0,
    "Resource": 10.0, "Internal Number of Transp. Unit": 33.0,
    "Transportation Unit": 21.0, "Carrier": 9.0, "Handling Unit": 20.0,
    "Product": 10.0, "Product Short Description": 37.0, "Quantity": 10.0,
    "Base Unit of Measure": 13.0, "Stock Type": 12.0,
    "Description of Stock Type": 28.0, "Batch": 12.0, "BDM": 26.71,
    "Stock Segment": 15.0, "Batch in restr.-use": 21.0, "Owner": 8.0,
    "Usage": 7.0, "Type": 6.0, "Stock Reference Document": 26.0,
    "Sales Order Item": 11.0, "Inspection ID Type": 20.0,
    "Shelf Life Expiration Date": 13.0, "Goods Receipt Time": 20.0,
    "Inventory Active": 18.0, "Storage Bin Improvable": 24.0,
    "Certificate Number": 20.0, "Higher-Level HU": 17.0,
    "Highest-Level HU": 20.0, "Packed Qty (AUoM)": 6.0,
    "Alt. Unit of Measure": 13.0, "Overall Valuation Quantity": 18.0,
    "Valuation Unit": 10.0, "Valuation Measured": 12.0,
    "Stock Identification": 22.0, "Document Category": 12.0, "Document": 10.0,
    "Item Number": 13.0, "Loading Weight": 16.0, "Weight Unit": 13.0,
    "Loading Volume": 16.0, "Volume Unit": 8.0, "Capacity Consumption": 7.0,
    "Consolidation Group": 21.0, "Serial No. Requiremt": 22.0,
    "Production Supply Area": 24.0, "Order Item Reduced": 20.0,
    "WIP Number": 12.0, "Latest Delivery Date": 22.0,
    # Materials-dispose sheets (2925 / 2935) — headers unique to that shape.
    "Material": 13.0, "Material Description": 36.0, "Plant": 6.14,
    "Plant Name": 16.0, "Storage Location": 18.57,
    "Description of Storage Location": 34.0, "Last sell day": 13.57,
    "Last sell date": 14.57, "Special Stock Type Description": 33.29,
    "Unrestricted Stock": 20.29, "Stock in Quality Inspection": 28.43,
    "Blocked Stock": 15.71,
}


def _write_sheet(ws, df: pd.DataFrame) -> None:
    headers = list(df.columns)
    for c_idx, col in enumerate(headers, 1):
        ws.cell(row=1, column=c_idx, value=col).font = _FONT

    fmt_for = {}
    for cols, fmt in (
        (DATE_COLS, _DATE_FMT), (TIME_COLS, _TIME_FMT),
        (INT_COLS, _INT_FMT), (DECIMAL_COLS, _DECIMAL_FMT),
    ):
        for c in cols:
            if c in headers:
                fmt_for[headers.index(c) + 1] = fmt

    for r_off, (_, row) in enumerate(df.iterrows()):
        r_idx = r_off + 2
        for c_idx, col in enumerate(headers, 1):
            val = row[col]
            if pd.isna(val):
                val = None
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = _FONT
            fmt = fmt_for.get(c_idx)
            if fmt and val is not None:
                cell.number_format = fmt

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(df) + 1, 1)}"
    for c_idx, col in enumerate(headers, 1):
        width = _COL_WIDTHS.get(col)
        if width:
            ws.column_dimensions[get_column_letter(c_idx)].width = width


def generate_excel(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Render the per-plant DataFrames into a formatted workbook (bytes)."""
    wb = Workbook()
    wb.remove(wb.active)
    for plant, df in sheets.items():      # already in PLANTS order
        _write_sheet(wb.create_sheet(plant), df)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
