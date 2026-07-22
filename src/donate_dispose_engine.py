"""Processing engine for the weekly Donate / Dispose list.

This is the mirror image of the Overstock report (see ``overstock_engine``).
Both read the same two SAPUI5 exports (Materials inventory + Last Sell / BDM
master) and both split into Mississauga / Calgary / Surrey sheets, but the date
window is flipped:

* **Overstock** keeps stock whose Shelf Life Expiration Date is *far enough out*
  to still be sold (and restricts to the main warehouse).
* **Donate / Dispose** keeps stock that is *at, near, or past* expiry — too late
  to sell, so it must be donated or disposed of — across *every* storage
  location (main warehouse, overstock, clearance, rework hold, consignment).

The supplied finished workbook ``DonateDispose list - June 24 2026 P2.xlsx`` is
the golden specification. Rules reverse-engineered against it (all three sheets
reproduced exactly — 48 / 92 / 19 rows):

1. Total stock = Unrestricted + Quality Inspection + Blocked > 0.
2. Plant belongs to the sheet's region (see ``REGION_PLANTS``).
3. Material number does NOT start with "40" (display / shipper / label / sample
   packaging, never sellable stock).
4. The Material matches a master ``Product Number`` (so it has a Last Sell Day).
5. NOT the RANA retail brand handled by Sandra; NOT the Sweet Street ("SSD …")
   brand.
6. Shelf Life Expiration Date present and on/before the cutoff
   (report date + ``SLED_CUTOFF_OFFSET_DAYS``).

There is **no** storage-location restriction and **no** last-sell-date filter
(the SLED cutoff alone defines the window). Rows are sorted by Shelf Life
Expiration Date ascending within each sheet.

Storage bins (optional third input)
-----------------------------------
Mirrors the Overstock report. An **EWM dispose export** may be supplied per
plant (2910 / 2920 / 2930) — ``Mo - EWM 29xx dispose.xlsx``. It carries one row
per Product / Batch / Storage Bin and names the bin each expiring batch is
sitting in, replacing a manual ``VLOOKUP(Material&Batch, EWM!I:J, 2, 0)``.

* The key is ``Material`` & ``Batch`` concatenated with no separator.
* EWM repeats a Product/Batch across many bins; the VLOOKUP returned the
  **first** match in file order, so the lookup keeps first and file order is
  load-bearing.
* An empty Bin and an ``#N/A`` Bin mean different things: **blank** = that
  plant's export was searched and this batch wasn't in it; **#N/A** = there was
  nothing to search. #N/A is always the case for plants 2925 / 2935, which have
  no export, and for any of 2910 / 2920 / 2930 whose file wasn't uploaded.
* The bins are a lookup only. They never add, drop, or reorder a row.

Bins are optional: with no EWM file the list is built exactly as before.

This module is intentionally self-contained (it does not import the overstock
engine) — every engine in this app stands alone to keep imports simple. The bin
helpers below are therefore duplicated from ``overstock_engine`` rather than
shared; importing across engines once deadlocked on Streamlit Cloud.
"""
from __future__ import annotations

import io
from datetime import date, datetime, timedelta
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


class DonateDisposeError(Exception):
    """Raised when an uploaded file isn't a usable Donate/Dispose source export."""


# ---------------------------------------------------------------------------
# Source columns (exact export headers)
# ---------------------------------------------------------------------------
MAT_MATERIAL      = "Material"
MAT_DESCRIPTION   = "Material Description"
MAT_PLANT         = "Plant"
MAT_PLANT_NAME    = "Plant Name"
MAT_STORAGE_LOC   = "Storage Location"
MAT_STORAGE_DESC  = "Description of Storage Location"
MAT_BATCH         = "Batch"
MAT_SLED          = "Shelf Life Expiration Date"
MAT_SPECIAL_STOCK = "Special Stock Type Description"
MAT_UNRESTRICTED  = "Unrestricted Stock"
MAT_QUALITY       = "Stock in Quality Inspection"
MAT_BLOCKED       = "Blocked Stock"

STOCK_COLS = [MAT_UNRESTRICTED, MAT_QUALITY, MAT_BLOCKED]
MAT_TEXT_COLS = [MAT_MATERIAL, MAT_PLANT, MAT_STORAGE_LOC, MAT_BATCH]

MASTER_PRODUCT   = "Product Number"
MASTER_BDM_NAME  = "Brand Manager Name"
MASTER_LAST_SELL = "Last Sell Day"

# EWM dispose export (one row per Product / Batch / Storage Bin).
EWM_PRODUCT = "Product"
EWM_BATCH   = "Batch"
EWM_BIN     = "Storage Bin"
EWM_OWNER   = "Party Entitled to Dispose"   # "BP2910" — lets a file name its plant

# Output columns. NOTE: this report uses "Last sell day" / "Last sell date"
# (without the "by" the overstock report uses).
OUT_BDM           = "BDM"
OUT_LAST_SELL_DAY = "Last sell day"
OUT_LAST_SELL_DT  = "Last sell date"
OUT_MATERIALBATCH = "Materialbatch"
OUT_BIN           = "Bin"

OUTPUT_COLUMNS = [
    MAT_MATERIAL,
    MAT_DESCRIPTION,
    MAT_PLANT,
    MAT_PLANT_NAME,
    OUT_BDM,
    MAT_STORAGE_LOC,
    OUT_MATERIALBATCH,
    OUT_BIN,
    MAT_STORAGE_DESC,
    MAT_BATCH,
    MAT_SLED,
    OUT_LAST_SELL_DAY,
    OUT_LAST_SELL_DT,
    MAT_SPECIAL_STOCK,
    MAT_UNRESTRICTED,
    MAT_QUALITY,
    MAT_BLOCKED,
]

# Region -> plant codes (shared logic with overstock, duplicated to stay
# self-contained). Sheet order is preserved on output.
REGION_PLANTS: dict[str, list[str]] = {
    "Mississauga": ["2910"],
    "Calgary": ["2920", "2925"],
    "Surrey": ["2930", "2935"],
}

# ---------------------------------------------------------------------------
# Business-rule constants (auditable; change here, not in the UI)
# ---------------------------------------------------------------------------
EXCLUDED_MATERIAL_PREFIXES = ("40",)        # packaging / display / promo
SWEET_STREET_DESC_PREFIX = "SSD"            # Sweet Street desserts brand
RANA_DESC_PREFIX = "RANA"                   # Giovanni Rana
RANA_EXCLUDED_BDM = "SANDRA GAGANIARAS GB"  # her RANA retail line is dropped

# Date window: stock is in scope when its Shelf Life Expiration Date is
# on/before report_date + this many days.
SLED_CUTOFF_OFFSET_DAYS = 4

# Plants with an EWM dispose export. Region plants absent here (2925, 2935) have
# no bin data at all.
EWM_PLANTS: list[str] = ["2910", "2920", "2930"]

# Shown when there is no EWM export to look the row up in, so a reader can tell
# "checked, this batch has no bin" (blank) from "no bin data exists for this
# plant" (#N/A). openpyxl binds this string to a real Excel error value, which
# is what the manual VLOOKUP produced.
NO_LOOKUP_MARKER = "#N/A"


# ---------------------------------------------------------------------------
# Importers
# ---------------------------------------------------------------------------
def _read_excel_str(file_obj: Any) -> pd.DataFrame:
    file_obj.seek(0)
    df = pd.read_excel(file_obj, dtype=str, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _require(df: pd.DataFrame, cols: list[str], what: str) -> None:
    missing = [c for c in cols if c not in df.columns]
    if missing:
        raise DonateDisposeError(
            f"This doesn't look like a {what} export — missing column(s): "
            + ", ".join(missing)
        )


def load_materials(file_obj: Any) -> pd.DataFrame:
    """Read the Materials inventory export. Stock buckets become numeric, the
    SLED becomes a datetime, and id columns are kept as clean text strings."""
    df = _read_excel_str(file_obj)
    _require(df, [MAT_MATERIAL, MAT_PLANT, MAT_SLED] + STOCK_COLS, "Materials")

    for c in STOCK_COLS:
        df[c] = pd.to_numeric(df[c], errors="coerce").fillna(0)

    df[MAT_SLED] = pd.to_datetime(df[MAT_SLED], errors="coerce")

    for c in MAT_TEXT_COLS:
        if c in df.columns:
            # Strip only ASCII whitespace so any preserved batch padding
            # (trailing non-breaking spaces) survives verbatim.
            df[c] = (
                df[c].astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.strip(" \t\r\n")
                .replace({"nan": "", "None": "", "NaT": ""})
            )
    return df


def load_master(file_obj: Any) -> pd.DataFrame:
    """Read the Last Sell / BDM master and reduce it to one row per Product
    Number (the master repeats products across vendors)."""
    df = _read_excel_str(file_obj)
    _require(df, [MASTER_PRODUCT, MASTER_BDM_NAME, MASTER_LAST_SELL],
             "Last Sell / BDM Material Master")

    df[MASTER_PRODUCT] = (
        df[MASTER_PRODUCT].astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )
    df[MASTER_LAST_SELL] = pd.to_numeric(df[MASTER_LAST_SELL], errors="coerce")
    df = df.drop_duplicates(subset=MASTER_PRODUCT, keep="first")
    return df[[MASTER_PRODUCT, MASTER_BDM_NAME, MASTER_LAST_SELL]].reset_index(drop=True)


def _clean_text(s: pd.Series) -> pd.Series:
    r"""Blank-safe text: NA becomes "", float artifacts are dropped, and padding
    is trimmed.

    ``.str.strip()`` rather than a ``\s+`` regex on purpose. Batches carry
    **non-breaking** space padding — ``load_materials`` preserves it
    deliberately — and pandas runs ``.str.replace(regex=True)`` on PyArrow's
    RE2 engine, whose ``\s`` is ASCII-only and leaves U+00A0 in place.
    ``str.strip`` is Unicode-aware. Likewise ``fillna`` rather than
    ``astype(str)``: the latter leaves NA as NA on the string dtype, so a
    "nan" -> "" mapping silently misses it.
    """
    return (
        s.fillna("")
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .replace({"nan": "", "None": "", "NaT": ""})
    )


def _bin_key(material: pd.Series, batch: pd.Series) -> pd.Series:
    """Material & Batch, trimmed on both sides — the join key. Batches arrive
    padded inconsistently between the two exports, so the key strips while the
    displayed Materialbatch keeps whatever padding it came with."""
    return _clean_text(material) + _clean_text(batch)


def load_ewm_bins(file_obj: Any, expected_plant: str | None = None) -> pd.Series:
    """Read one plant's EWM dispose export and reduce it to a
    ``Material+Batch -> Storage Bin`` lookup.

    EWM lists a Product/Batch once per bin it occupies; the manual VLOOKUP took
    the first hit, so this keeps the first row in file order. Pass
    ``expected_plant`` to reject a file uploaded into the wrong plant's box —
    each export names its own plant in ``Party Entitled to Dispose`` ("BP2910").
    """
    df = _read_excel_str(file_obj)
    _require(df, [EWM_PRODUCT, EWM_BATCH, EWM_BIN],
             f"EWM dispose ({expected_plant})" if expected_plant else "EWM dispose")

    if expected_plant and EWM_OWNER in df.columns:
        owners = (
            df[EWM_OWNER].dropna().astype(str).str.strip()
            .str.upper().str.removeprefix("BP")
        )
        found = sorted({o for o in owners if o})
        if found and expected_plant not in found:
            raise DonateDisposeError(
                f"This file is the plant {', '.join(found)} EWM dispose export, "
                f"but it was uploaded in the {expected_plant} box. Swap the "
                "files and try again."
            )

    keys = _bin_key(df[EWM_PRODUCT], df[EWM_BATCH])
    bins = _clean_text(df[EWM_BIN])

    lut = pd.DataFrame({"key": keys, "bin": bins})
    lut = lut[(lut["key"] != "") & (lut["bin"] != "")]
    lut = lut.drop_duplicates(subset="key", keep="first")   # VLOOKUP first match
    return lut.set_index("key")["bin"]


# ---------------------------------------------------------------------------
# Core build
# ---------------------------------------------------------------------------
def _coerce_report_date(report_date: date | datetime | None) -> date:
    if report_date is None:
        return date.today()
    if isinstance(report_date, datetime):
        return report_date.date()
    return report_date


def default_cutoff(report_date: date | datetime | None = None) -> date:
    """The default SLED cutoff for a run date. Used to pre-fill the page; the
    user can override it freely."""
    return _coerce_report_date(report_date) + timedelta(days=SLED_CUTOFF_OFFSET_DAYS)


def build_donate_dispose(
    materials: pd.DataFrame,
    master: pd.DataFrame,
    sled_cutoff: date | datetime | None = None,
    report_date: date | datetime | None = None,
    ewm_bins: dict[str, pd.Series] | None = None,
) -> dict[str, pd.DataFrame]:
    """Apply the donate/dispose selection rules and return one DataFrame per
    region in finished-workbook column order.

    Pass ``sled_cutoff`` (include Shelf Life Expiration on/before) explicitly;
    left as ``None`` it falls back to report_date + ``SLED_CUTOFF_OFFSET_DAYS``
    (today if ``report_date`` is also ``None``).

    ``ewm_bins`` maps a plant code to that plant's ``load_ewm_bins`` lookup. It
    only fills the Bin column — row selection is identical with or without it.
    Plants with no entry get ``#N/A``; plants with one get the bin, or blank
    where the batch isn't in their export."""
    cutoff = pd.Timestamp(
        sled_cutoff if sled_cutoff is not None else default_cutoff(report_date)
    )

    m = materials.copy()
    lut = master.rename(columns={
        MASTER_BDM_NAME: OUT_BDM,
        MASTER_LAST_SELL: OUT_LAST_SELL_DAY,
    })
    m = m.merge(lut, how="left", left_on=MAT_MATERIAL, right_on=MASTER_PRODUCT)

    m[OUT_LAST_SELL_DT] = m[MAT_SLED] - pd.to_timedelta(m[OUT_LAST_SELL_DAY], unit="D")
    total_stock = m[STOCK_COLS].sum(axis=1)

    desc = m[MAT_DESCRIPTION].astype(str).str.strip().str.upper()
    bdm = m[OUT_BDM].astype(str).str.strip().str.upper()

    is_packaging = m[MAT_MATERIAL].astype(str).str.startswith(EXCLUDED_MATERIAL_PREFIXES)
    is_sweet_street = desc.str.startswith(SWEET_STREET_DESC_PREFIX)
    is_rana_sandra = (bdm == RANA_EXCLUDED_BDM) & desc.str.startswith(RANA_DESC_PREFIX)

    keep = (
        (total_stock > 0)
        & m[MAT_SLED].notna()
        & m[OUT_LAST_SELL_DAY].notna()
        & ~is_packaging
        & ~is_sweet_street
        & ~is_rana_sandra
        & (m[MAT_SLED] <= cutoff)
    )
    selected = m[keep].copy()

    ewm_bins = ewm_bins or {}

    sheets: dict[str, pd.DataFrame] = {}
    for sheet_name, plants in REGION_PLANTS.items():
        sub = selected[selected[MAT_PLANT].isin(plants)].copy()
        sub = sub.sort_values(
            [MAT_SLED, MAT_MATERIAL, MAT_BATCH], kind="mergesort"
        ).reset_index(drop=True)

        # Displayed key keeps the batch's own padding; the join key strips it.
        # fillna before astype — on the string dtype astype(str) leaves NA as
        # NA, which would turn the whole key blank for a batch-less row.
        sub[OUT_MATERIALBATCH] = (
            sub[MAT_MATERIAL].fillna("").astype(str)
            + sub[MAT_BATCH].fillna("").astype(str)
        )

        # A plant with no lookup gets #N/A — nothing was checked. A plant with a
        # lookup gets the bin, or blank when the batch simply isn't in it.
        join_key = _bin_key(sub[MAT_MATERIAL], sub[MAT_BATCH])
        sub[OUT_BIN] = pd.Series(
            [NO_LOOKUP_MARKER] * len(sub), index=sub.index, dtype=object
        )
        for plant in plants:
            lut = ewm_bins.get(plant)
            if lut is None or lut.empty:
                continue
            at_plant = sub[MAT_PLANT] == plant
            sub.loc[at_plant, OUT_BIN] = join_key[at_plant].map(lut)

        out = pd.DataFrame({col: sub.get(col) for col in OUTPUT_COLUMNS})
        out[MAT_SPECIAL_STOCK] = out[MAT_SPECIAL_STOCK].where(
            out[MAT_SPECIAL_STOCK].notna(), None
        )
        sheets[sheet_name] = out

    return sheets


# ---------------------------------------------------------------------------
# Excel export (matches the finished workbook: mm-dd-yy dates, plain stock)
# ---------------------------------------------------------------------------
_HEADER_FILL = PatternFill("solid", fgColor="FFF7F7F7")
_HEADER_FONT = Font(name="Arial", size=11, bold=True)
_BODY_FONT = Font(name="Arial", size=11)
_DATE_FMT = "mm-dd-yy"

_COL_WIDTHS = {
    MAT_MATERIAL: 13.0,
    MAT_DESCRIPTION: 36.0,
    MAT_PLANT: 6.14,
    MAT_PLANT_NAME: 16.0,
    OUT_BDM: 23.14,
    MAT_STORAGE_LOC: 18.57,
    OUT_MATERIALBATCH: 13.0,
    OUT_BIN: 20.0,
    MAT_STORAGE_DESC: 34.0,
    MAT_BATCH: 11.57,
    MAT_SLED: 27.43,
    OUT_LAST_SELL_DAY: 13.57,
    OUT_LAST_SELL_DT: 14.57,
    MAT_SPECIAL_STOCK: 33.29,
    MAT_UNRESTRICTED: 20.29,
    MAT_QUALITY: 28.43,
    MAT_BLOCKED: 15.71,
}

# Materialbatch is written as text so long numeric-looking keys aren't mangled
# into floats. Bin is left General, matching the manual workbook.
_TEXT_OUT_COLS = set(MAT_TEXT_COLS) | {OUT_MATERIALBATCH}


def _write_sheet(ws, df: pd.DataFrame) -> None:
    headers = list(df.columns)
    for c_idx, col in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")

    date_idx = {headers.index(c) + 1 for c in (MAT_SLED, OUT_LAST_SELL_DT) if c in headers}
    text_idx = {headers.index(c) + 1 for c in _TEXT_OUT_COLS if c in headers}

    for r_off, (_, row) in enumerate(df.iterrows()):
        r_idx = r_off + 2
        for c_idx, col in enumerate(headers, 1):
            val = row[col]
            if pd.isna(val):
                val = None
            if c_idx in text_idx and val is not None:
                val = str(val)
            if isinstance(val, pd.Timestamp):
                val = val.to_pydatetime()
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = _BODY_FONT
            if c_idx in date_idx and val is not None:
                cell.number_format = _DATE_FMT
            elif c_idx in text_idx:
                cell.number_format = "@"

    last_col = get_column_letter(len(headers))
    ws.auto_filter.ref = f"A1:{last_col}{max(len(df) + 1, 1)}"
    for c_idx, col in enumerate(headers, 1):
        width = _COL_WIDTHS.get(col)
        if width:
            ws.column_dimensions[get_column_letter(c_idx)].width = width


def generate_excel(sheets: dict[str, pd.DataFrame]) -> bytes:
    """Render the region DataFrames into a formatted workbook (bytes)."""
    wb = Workbook()
    wb.remove(wb.active)
    for sheet_name in REGION_PLANTS:
        ws = wb.create_sheet(sheet_name)
        _write_sheet(ws, sheets[sheet_name])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
