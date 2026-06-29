"""One-time generator for src/templates/risky_inventory_template.xlsx.

Derives the Risky Inventory PivotTable template from a golden export, adds a
'Bucket' page field, and scrubs all embedded supplier data (sharedItems +
cached records emptied; Excel rebuilds them on open via refreshOnLoad).

Usage:
    python scripts/build_risky_inventory_template.py ["path/to/golden 90D.xlsx"]
"""
from __future__ import annotations

import io
import sys
import zipfile
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.pivot.cache import CacheField, SharedItems
from openpyxl.pivot.table import FieldItem, PageField, PivotField
from openpyxl.utils import get_column_letter

GOLDEN = sys.argv[1] if len(sys.argv) > 1 else (
    r"C:/Users/melgh/Downloads/Risky Inventory June 24 P2 - 90D.xlsx"
)
OUT = Path(__file__).resolve().parents[1] / "src" / "templates" / "risky_inventory_template.xlsx"

EMPTY_RECORDS = (
    b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    b'<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" count="0"/>'
)


def main() -> None:
    wb = load_workbook(GOLDEN)
    ws1, ws2 = wb["Sheet1"], wb["Sheet2"]
    piv = ws2._pivots[0]
    cache = piv.cache

    ucol = ws1.max_column + 1            # 21 -> column U
    ws1.cell(1, ucol, "Bucket")

    # Insert Bucket as cache/pivot field 20 (after the 20 base columns, before the
    # grouped date fields). Append it as a page filter set to (All).
    cache.cacheFields.insert(20, CacheField(name="Bucket", sharedItems=SharedItems()))
    piv.pivotFields.insert(20, PivotField(axis="axisPage", showAll=False,
                                          items=[FieldItem(t="default")]))
    piv.pageFields.append(PageField(fld=20))

    # Rename sheets and repoint the cache at the (empty) Detail header.
    ws1.title, ws2.title = "Detail", "Summary"
    last = get_column_letter(ucol)
    cache.cacheSource.worksheetSource.sheet = "Detail"
    cache.cacheSource.worksheetSource.ref = f"A1:{last}1"
    cache.refreshOnLoad = True
    cache.recordCount = 0

    # Scrub embedded data: empty every field's shared items, then drop data rows.
    for cf in cache.cacheFields:
        cf.sharedItems = SharedItems()
    if ws1.max_row > 1:
        ws1.delete_rows(2, ws1.max_row - 1)

    buf = io.BytesIO()
    wb.save(buf)

    # Replace the cached records part with an empty one (no supplier data).
    OUT.parent.mkdir(parents=True, exist_ok=True)
    zin = zipfile.ZipFile(io.BytesIO(buf.getvalue()))
    with zipfile.ZipFile(OUT, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            content = zin.read(item.filename)
            if item.filename.endswith("pivotCacheRecords1.xml"):
                content = EMPTY_RECORDS
            zout.writestr(item, content)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    main()
